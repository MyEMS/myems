import base64
import gettext
import os
import uuid

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import column_index_from_string, get_column_letter


########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file from the report data
# Step 3: Encode the excel file to Base64
########################################################################################################################

def export(result, space_name, reporting_start_datetime_local, reporting_end_datetime_local, language):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if result is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(result,
                              space_name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              language)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        print(str(ex))

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        print(str(ex))
    return base64_message


def generate_excel(report, space_name, reporting_start_datetime_local, reporting_end_datetime_local, language):

    locale_path = './i18n/'
    if language == 'zh_CN':
        trans = gettext.translation('myems', locale_path, languages=['zh_CN'])
    elif language == 'de':
        trans = gettext.translation('myems', locale_path, languages=['de'])
    elif language == 'en':
        trans = gettext.translation('myems', locale_path, languages=['en'])
    else:
        trans = gettext.translation('myems', locale_path, languages=['en'])
    trans.install()
    _ = trans.gettext

    wb = Workbook()
    ws = wb.active
    ws.title = "TenantBatch"

    # Row height
    ws.row_dimensions[1].height = 102
    for i in range(2, 5 + 1):
        ws.row_dimensions[i].height = 42

    for i in range(6, len(report['tenants']) + 15):
        ws.row_dimensions[i].height = 60

    # Col width
    ws.column_dimensions['A'].width = 1.5

    ws.column_dimensions['B'].width = 25.0

    for i in range(ord('C'), ord('L')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Arial', size=15, bold=True)
    title_font = Font(name='Arial', size=15, bold=True)

    table_fill = PatternFill(fill_type='solid', fgColor='90ee90')
    f_border = Border(left=Side(border_style='medium'),
                      right=Side(border_style='medium'),
                      bottom=Side(border_style='medium'),
                      top=Side(border_style='medium')
                      )
    b_border = Border(
        bottom=Side(border_style='medium'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)

    # Img
    img = Image("excelexporters/myems.png")
    ws.add_image(img, 'A1')

    # Title
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = _('Space') + ':'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'] = space_name

    ws['B4'].alignment = b_r_alignment
    ws['B4'] = _('Reporting Start Datetime') + ':'
    ws['C4'].border = b_border
    ws['C4'].alignment = b_c_alignment
    ws['C4'] = reporting_start_datetime_local

    ws['B5'].alignment = b_r_alignment
    ws['B5'] = _('Reporting End Datetime') + ':'
    ws['C5'].border = b_border
    ws['C5'].alignment = b_c_alignment
    ws['C5'] = reporting_end_datetime_local

    # Title
    ws['B7'].border = f_border
    ws['B7'].font = name_font
    ws['B7'].alignment = c_c_alignment
    ws['B7'].fill = table_fill
    ws['B7'] = _('Name') + ':'

    ws['C7'].border = f_border
    ws['C7'].alignment = c_c_alignment
    ws['C7'].font = name_font
    ws['C7'].fill = table_fill
    ws['C7'] = _('Space') + ':'

    ca_len = len(report['energycategories'])

    for i in range(0, ca_len):
        col = get_column_letter(column_index_from_string('D') + i * 2)
        ws[col + '7'].fill = table_fill
        ws[col + '7'].font = name_font
        ws[col + '7'].alignment = c_c_alignment
        ws[col + '7'] = report['energycategories'][i]['name'] + \
            " (" + report['energycategories'][i]['unit_of_measure'] + ")"
        ws[col + '7'].border = f_border

        col = get_column_letter(column_index_from_string(col) + 1)
        ws[col + '7'].fill = table_fill
        ws[col + '7'].font = name_font
        ws[col + '7'].alignment = c_c_alignment
        ws[col + '7'] = report['energycategories'][i]['name'] + \
            " " + _('Maximum Load') + " (" + report['energycategories'][i]['unit_of_measure'] + ")"
        ws[col + '7'].border = f_border

    current_row_number = 8
    for i in range(0, len(report['tenants'])):

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)].border = f_border
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)] = report['tenants'][i]['tenant_name']

        ws['C' + str(current_row_number)].font = title_font
        ws['C' + str(current_row_number)].border = f_border
        ws['C' + str(current_row_number)].alignment = c_c_alignment
        ws['C' + str(current_row_number)] = report['tenants'][i]['space_name']

        ca_len = len(report['tenants'][i]['values'])
        for j in range(0, ca_len):
            col = get_column_letter(column_index_from_string('D') + j * 2)
            ws[col + str(current_row_number)].font = title_font
            ws[col + str(current_row_number)].border = f_border
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)] = report['tenants'][i]['values'][j]

            col = get_column_letter(column_index_from_string(col) + 1)
            ws[col + str(current_row_number)].font = title_font
            ws[col + str(current_row_number)].border = f_border
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)] = report['tenants'][i]['maximum'][j]
        current_row_number += 1

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename
