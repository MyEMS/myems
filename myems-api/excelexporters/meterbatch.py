"""
Meter Batch Excel Exporter

This module provides functionality to export meter batch data to Excel format.
It generates comprehensive reports showing energy consumption data for multiple meters
within a specific time period.

Key Features:
- Multi-meter energy consumption comparison
- Energy category breakdown with units
- Daily energy consumption breakdown
- Formatted Excel output with proper styling
- Multi-language support
- Base64 encoding for file transmission

The exported Excel file includes:
- Meter names and associated spaces
- Energy consumption by category
- Daily consumption breakdown
- Proper formatting and borders
- Logo and header information
"""

import base64
from core.utilities import get_translation
import os
import uuid
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file from the report data
# Step 3: Encode the excel file to Base64
########################################################################################################################

def export(result, space_name, reporting_start_datetime_local, reporting_end_datetime_local, language):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if result is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(result,
                              space_name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              language)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        print(str(ex))

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        print(str(ex))
    return base64_message


def generate_excel(report, space_name, reporting_start_datetime_local, reporting_end_datetime_local, language):

    trans = get_translation(language)
    trans.install()
    _ = trans.gettext

    wb = Workbook()
    ws = wb.active
    ws.title = "MeterBatch"

    date_list = list(report.get('date_list') or [])
    if len(date_list) == 0:
        first_meter = next(iter(report.get('meters', [])), None)
        if first_meter is not None:
            date_list = [item.get('date') for item in first_meter.get('daily_values', []) if item.get('date')]

    # Calculate required columns
    ca_len = 1
    date_list_len = len(date_list)
    total_cols = 3 + ca_len + date_list_len + 1  # ID, Name, Space + categories + dates + total

    # Col width
    for col_number in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_number)].width = 15.0

    # Set wider width for date columns
    if date_list_len > 0:
        for col_idx in range(3 + ca_len, 3 + ca_len + date_list_len):
            ws.column_dimensions[get_column_letter(col_idx + 1)].width = 12.0

    # Head image
    ws.row_dimensions[1].height = 105
    img = Image("excelexporters/myems.png")
    ws.add_image(img, 'A1')

    # Query Parameters
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    ws['A3'].alignment = b_r_alignment
    ws['A3'] = _('Space') + ':'
    ws['B3'] = space_name
    ws['A4'].alignment = b_r_alignment
    ws['A4'] = _('Start Datetime') + ':'
    ws['B4'] = reporting_start_datetime_local
    ws['A5'].alignment = b_r_alignment
    ws['A5'] = _('End Datetime') + ':'
    ws['B5'] = reporting_end_datetime_local

    # Title
    title_font = Font(size=12, bold=True)
    ws['A6'].font = title_font
    ws['A6'] = _('ID')
    ws['B6'].font = title_font
    ws['B6'] = _('Name')
    ws['C6'].font = title_font
    ws['C6'] = _('Space')
    ws['D6'].font = title_font
    ws['D6'] = _('Energy Category')

    # Energy category headers
    # for i in range(0, ca_len):
    #     col = chr(ord('D') + i)
    #     ws[col + '6'].font = title_font
    #     ws[col + '6'] = report['energycategories'][i]['name'] + " (" + \
    #                     report['energycategories'][i]['unit_of_measure'] + ")"

    # Date headers
    for i in range(0, len(date_list)):
        col_idx = 3 + ca_len + i
        col = get_column_letter(col_idx + 1)
        ws[col + '6'].font = title_font
        ws[col + '6'] = str(date_list[i])
    total_col_idx = 3 + ca_len + len(date_list)
    total_col = get_column_letter(total_col_idx + 1)
    ws[total_col + '6'].font = title_font
    ws[total_col + '6'] = _('Total')

    current_row_number = 7
    for i in range(0, len(report['meters'])):
        ws['A' + str(current_row_number)] = str(report['meters'][i]['id'])
        ws['B' + str(current_row_number)] = report['meters'][i]['meter_name']
        ws['C' + str(current_row_number)] = report['meters'][i]['space_name']
        ws['D' + str(current_row_number)] = report['meters'][i].get('energy_category_name', '')
        # Daily values
        daily_values = report['meters'][i].get('daily_values', [])
        daily_value_map = {item.get('date'): item.get('value') for item in daily_values if item.get('date')}
        for j, day in enumerate(date_list):
            col_idx = 3 + ca_len + j
            col = get_column_letter(col_idx + 1)
            ws[col + str(current_row_number)] = daily_value_map.get(day, None)
        ws[total_col + str(current_row_number)] = report['meters'][i].get('subtotal', None)

        current_row_number += 1

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename

