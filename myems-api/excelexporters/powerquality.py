"""
Meter Trend Excel Exporter

This module provides functionality to export meter trend data to Excel format.
It generates comprehensive reports showing energy consumption trends for meters
with detailed analysis and visualizations.

Key Features:
- Meter energy consumption trend analysis
- Base period vs reporting period comparison
- Trend breakdown by energy categories
- Detailed data with line charts
- Multi-language support
- Base64 encoding for file transmission

The exported Excel file includes:
- Meter trend summary
- Base period comparison data
- Trend breakdown by energy categories
- Detailed time-series data with line charts
- Parameter data (if available)
"""

import base64
from core.utilities import get_translation
import os
import re
import uuid
import openpyxl.utils.cell as format_cell
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from core.utilities import round2

########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file
# Step 3: Encode the excel file bytes to Base64
########################################################################################################################


def export(result,
           name,
           reporting_start_datetime_local,
           reporting_end_datetime_local,
           period_type,
           language):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if result is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(result,
                              name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              period_type,
                              language)

    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        print(str(ex))

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        print(str(ex))
    return base64_message


def generate_excel(report,
                   name,
                   reporting_start_datetime_local,
                   reporting_end_datetime_local,
                   period_type,
                   language):

    trans = get_translation(language)
    trans.install()
    _ = trans.gettext

    wb = Workbook()
    ws = wb.active
    ws.title = "MeterTrend"

    # Row height
    ws.row_dimensions[1].height = 102
    for i in range(2, 8):
        ws.row_dimensions[i].height = 42

    # Col width
    ws.column_dimensions['A'].width = 1.5

    ws.column_dimensions['B'].width = 25.0

    for i in range(ord('C'), ord('V')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Arial', size=15, bold=True)
    title_font = Font(name='Arial', size=15, bold=True)

    table_fill = PatternFill(fill_type='solid', fgColor='90ee90')
    f_border = Border(left=Side(border_style='medium'),
                      right=Side(border_style='medium'),
                      bottom=Side(border_style='medium'),
                      top=Side(border_style='medium')
                      )
    b_border = Border(
        bottom=Side(border_style='medium'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)

    # Img
    img = Image("excelexporters/myems.png")
    ws.add_image(img, 'A1')

    # Title
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = _('Name') + ':'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'] = name

    ws['B4'].alignment = b_r_alignment
    ws['B4'] = _('Reporting Start Datetime') + ':'
    ws['C4'].border = b_border
    ws['C4'].alignment = b_c_alignment
    ws['C4'] = reporting_start_datetime_local

    ws['B5'].alignment = b_r_alignment
    ws['B5'] = _('Reporting End Datetime') + ':'
    ws['C5'].border = b_border
    ws['C5'].alignment = b_c_alignment
    ws['C5'] = reporting_end_datetime_local

    if "reporting_period" not in report.keys() or \
            "names" not in report['reporting_period'].keys() or len(report['reporting_period']['names']) == 0:
        filename = str(uuid.uuid4()) + '.xlsx'
        wb.save(filename)

        return filename
    ####################################################################################################################
    # First: Trend
    # 6: title
    # 7: table title
    # 8~ table_data
    ####################################################################################################################
    has_data_flag = True
    report['reporting_period'] = report['reporting_period']
    if "names" not in report['reporting_period'].keys() or \
            report['reporting_period']['names'] is None or \
            len(report['reporting_period']['names']) == 0:
        has_data_flag = False

    if "timestamps" not in report['reporting_period'].keys() or \
            report['reporting_period']['timestamps'] is None or \
            len(report['reporting_period']['timestamps']) == 0:
        has_data_flag = False

    if "values" not in report['reporting_period'].keys() or \
            report['reporting_period']['values'] is None or \
            len(report['reporting_period']['values']) == 0:
        has_data_flag = False

    ca_len = len(report['reporting_period']['names'])
    times = report['reporting_period']['timestamps']
    category = report['meter']['energy_category_name']
    start_detail_data_row_num = 9 + ca_len * 6
    if has_data_flag:
        time = times[0]
        for time in times:
            if len(time) > 0:
                break
        has_data = False
        current_sheet_parameters_row_number = 7
        for i in range(8, len(time) + 6 + ca_len * 6 + len(category) * 6 + 2):
            ws.row_dimensions[i].height = 42
        if len(time) > 0:
            has_data = True
            current_sheet_parameters_row_number = 7 + ca_len * 6
        if has_data:

            max_row = start_detail_data_row_num + len(time)
            ws['B6'].font = title_font
            ws['B6'] = name + ' ' + _('Trend')

            ws.row_dimensions[start_detail_data_row_num - 1].height = 60
            ws['B' + str(start_detail_data_row_num - 1)].fill = table_fill
            ws['B' + str(start_detail_data_row_num - 1)].font = title_font
            ws['B' + str(start_detail_data_row_num - 1)].border = f_border
            ws['B' + str(start_detail_data_row_num - 1)].alignment = c_c_alignment
            ws['B' + str(start_detail_data_row_num - 1)] = _('Datetime')

            for i in range(0, len(time)):
                col = 'B'
                row = str(start_detail_data_row_num + i)
                ws[col + row].font = title_font
                ws[col + row].alignment = c_c_alignment
                ws[col + row] = time[i]
                ws[col + row].border = f_border

            for i in range(0, ca_len):
                # 38 title
                col = format_cell.get_column_letter(3 + i)

                ws[col + str(start_detail_data_row_num - 1)].fill = table_fill
                ws[col + str(start_detail_data_row_num - 1)].font = title_font
                ws[col + str(start_detail_data_row_num - 1)].alignment = c_c_alignment
                ws[col + str(start_detail_data_row_num - 1)] = report['reporting_period']['names'][i]
                ws[col + str(start_detail_data_row_num - 1)].border = f_border

                for j in range(0, len(time)):

                    row = str(start_detail_data_row_num + j)
                    ws[col + row].font = title_font
                    ws[col + row].alignment = c_c_alignment
                    try:
                        ws[col + row] = round2(report['reporting_period']['values'][i][j], 3) if \
                            len(report['reporting_period']['values'][i]) > 0 and \
                            len(report['reporting_period']['values'][i]) > j and \
                            report['reporting_period']['values'][i][j] is not None else ''
                    except Exception as e:
                        print('error 1 in excelexporters\\metertrend: ' + str(e))

                    ws[col + row].border = f_border
            # line
            # 39~: line
                line = LineChart()
                line.title = report['reporting_period']['names'][i]
                labels = Reference(ws, min_col=2, min_row=start_detail_data_row_num, max_row=max_row-1)
                line_data = Reference(ws, min_col=3 + i, min_row=start_detail_data_row_num+1, max_row=max_row-1)
                line.add_data(line_data, titles_from_data=True)
                line.set_categories(labels)
                line_data.smooth = True
                line.x_axis.crosses = 'min'
                line.height = 8.25
                line.width = 36
                chart_col = chr(ord('B'))
                chart_cell = chart_col + str(7 + 6*i)

                ws.add_chart(line, chart_cell)

    ####################################################################################################################
    # Power Quality Analysis sheet
    ####################################################################################################################
    if 'analysis' in report and isinstance(report['analysis'], list) and len(report['analysis']) > 0:
        file_name = (re.sub(r'[^A-Z]', '', ws.title))+'_'
        analysis_ws = wb.create_sheet(file_name + _('Power Quality Analysis'))

        # Row height
        analysis_ws.row_dimensions[1].height = 102
        for i in range(2, 7 + 1):
            analysis_ws.row_dimensions[i].height = 42

        for i in range(8, 200):
            analysis_ws.row_dimensions[i].height = 24

        # Col width
        analysis_ws.column_dimensions['A'].width = 1.5
        analysis_ws.column_dimensions['B'].width = 28.0
        for i in range(3, 20):
            analysis_ws.column_dimensions[format_cell.get_column_letter(i)].width = 18.0

        # Img
        img = Image("excelexporters/myems.png")
        analysis_ws.add_image(img, 'A1')

        # Title
        analysis_ws['B3'].alignment = b_r_alignment
        analysis_ws['B3'] = _('Name') + ':'
        analysis_ws['C3'].border = b_border
        analysis_ws['C3'].alignment = b_c_alignment
        analysis_ws['C3'] = name

        analysis_ws['B4'].alignment = b_r_alignment
        analysis_ws['B4'] = _('Reporting Start Datetime') + ':'
        analysis_ws['C4'].border = b_border
        analysis_ws['C4'].alignment = b_c_alignment
        analysis_ws['C4'] = reporting_start_datetime_local

        analysis_ws['B5'].alignment = b_r_alignment
        analysis_ws['B5'] = _('Reporting End Datetime') + ':'
        analysis_ws['C5'].border = b_border
        analysis_ws['C5'].alignment = b_c_alignment
        analysis_ws['C5'] = reporting_end_datetime_local

        current_row = 6
        analysis_ws['B' + str(current_row)].font = title_font
        analysis_ws['B' + str(current_row)] = name + ' ' + _('Power Quality Analysis')
        current_row += 1

        # Header
        headers = [
            _('Point'), _('Category'), _('Type'), _('Unit'),
            _('Limit'), _('Normal Limit'), _('Severe Limit'), _('Compliance'),
            _('Worst Deviation'), _('Worst Time')
        ]
        # plus dynamic metrics
        metrics_names = set()
        for item in report['analysis']:
            if isinstance(item.get('metrics'), list):
                for m in item['metrics']:
                    metrics_names.add(m.get('name'))
        metrics_names = list(metrics_names)
        table_headers = headers + metrics_names

        analysis_ws.row_dimensions[current_row].height = 28
        for idx, h in enumerate(table_headers):
            col = format_cell.get_column_letter(2 + idx)
            analysis_ws[col + str(current_row)].fill = table_fill
            analysis_ws[col + str(current_row)].border = f_border
            analysis_ws[col + str(current_row)].font = name_font
            analysis_ws[col + str(current_row)].alignment = c_c_alignment
            analysis_ws[col + str(current_row)] = h
        current_row += 1

        # Rows
        for item in report['analysis']:
            row_values = [
                item.get('point_name', ''),
                item.get('category', ''),
                item.get('type', ''),
                item.get('unit', ''),
                item.get('limit_pct', ''),
                item.get('limit_normal_hz', ''),
                item.get('limit_severe_hz', ''),
                item.get('compliance_pct', ''),
                item.get('worst_abs_deviation_pct', item.get('worst_unbalance_pct', item.get('worst_deviation_hz', ''))),
                item.get('worst_time', ''),
            ]
            # metrics map
            metrics_map = {}
            if isinstance(item.get('metrics'), list):
                for m in item['metrics']:
                    metrics_map[m.get('name')] = m.get('value')
            for mn in metrics_names:
                row_values.append(metrics_map.get(mn, ''))

            for idx, v in enumerate(row_values):
                col = format_cell.get_column_letter(2 + idx)
                analysis_ws[col + str(current_row)].border = f_border
                analysis_ws[col + str(current_row)].font = title_font
                analysis_ws[col + str(current_row)].alignment = c_c_alignment
                analysis_ws[col + str(current_row)] = v
            current_row += 1

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename


def timestamps_data_all_equal_0(lists):
    for i, value in enumerate(list(lists)):
        if len(value) > 0:
            return False

    return True


def get_parameters_timestamps_lists_max_len(parameters_timestamps_lists):
    max_len = 0
    for i, value in enumerate(list(parameters_timestamps_lists)):
        if len(value) > max_len:
            max_len = len(value)

    return max_len


def timestamps_data_not_equal_0(lists):
    number = 0
    for i, value in enumerate(list(lists)):
        if len(value) > 0:
            number += 1
    return number
