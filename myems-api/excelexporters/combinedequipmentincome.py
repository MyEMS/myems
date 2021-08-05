import base64
import uuid
import os
from decimal import Decimal
from openpyxl.chart import PieChart, LineChart, Reference
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl.chart.label import DataLabelList
import openpyxl.utils.cell as format_cell


########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file
# Step 3: Encode the excel file bytes to Base64
########################################################################################################################


def export(report,
           name,
           reporting_start_datetime_local,
           reporting_end_datetime_local,
           period_type):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if report is None:
        return None
    print(report)

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(report,
                              name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              period_type)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        pass

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        pass
    return base64_message


def generate_excel(report,
                   name,
                   reporting_start_datetime_local,
                   reporting_end_datetime_local,
                   period_type):
    wb = Workbook()
    ws = wb.active
    ws.title = "CombinedEquipmentIncome"

    # Row height
    ws.row_dimensions[1].height = 102
    for i in range(2, 2000 + 1):
        ws.row_dimensions[i].height = 42

    # Col width
    ws.column_dimensions['A'].width = 1.5

    ws.column_dimensions['B'].width = 25.0

    for i in range(ord('C'), ord('L')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Constantia', size=15, bold=True)
    title_font = Font(name='宋体', size=15, bold=True)

    table_fill = PatternFill(fill_type='solid', fgColor='1F497D')
    f_border = Border(left=Side(border_style='medium', color='00000000'),
                      right=Side(border_style='medium', color='00000000'),
                      bottom=Side(border_style='medium', color='00000000'),
                      top=Side(border_style='medium', color='00000000')
                      )
    b_border = Border(
        bottom=Side(border_style='medium', color='00000000'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)

    # Img
    img = Image("excelexporters/myems.png")
    img.width = img.width * 0.85
    img.height = img.height * 0.85
    ws.add_image(img, 'B1')

    # Title
    ws.row_dimensions[3].height = 60

    ws['B3'].font = name_font
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = 'Name:'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'].font = name_font
    ws['C3'] = name

    ws['D3'].font = name_font
    ws['D3'].alignment = b_r_alignment
    ws['D3'] = 'Period:'
    ws['E3'].border = b_border
    ws['E3'].alignment = b_c_alignment
    ws['E3'].font = name_font
    ws['E3'] = period_type

    ws['F3'].font = name_font
    ws['F3'].alignment = b_r_alignment
    ws['F3'] = 'Date:'
    ws['G3'].border = b_border
    ws['G3'].alignment = b_c_alignment
    ws['G3'].font = name_font
    ws['G3'] = reporting_start_datetime_local + "__" + reporting_end_datetime_local
    ws.merge_cells("G3:H3")

    if "reporting_period" not in report.keys() or \
            "names" not in report['reporting_period'].keys() or len(report['reporting_period']['names']) == 0:
        filename = str(uuid.uuid4()) + '.xlsx'
        wb.save(filename)

        return filename

    ####################################################################################################################

    reporting_period_data = report['reporting_period']

    has_cost_data_flag = True

    if "names" not in reporting_period_data.keys() or \
            reporting_period_data['names'] is None or \
            len(reporting_period_data['names']) == 0:
        has_cost_data_flag = False

    if has_cost_data_flag:
        ws['B5'].font = title_font
        ws['B5'] = name + ' 报告期收入'
        category = reporting_period_data['names']
        ca_len = len(category)

        ws.row_dimensions[7].height = 60
        ws['B6'].fill = table_fill
        ws['B6'].border = f_border

        ws['B7'].font = title_font
        ws['B7'].alignment = c_c_alignment
        ws['B7'] = '报告期收入'
        ws['B7'].border = f_border

        ws['B8'].font = title_font
        ws['B8'].alignment = c_c_alignment
        ws['B8'] = '环比'
        ws['B8'].border = f_border

        col = ''

        for i in range(0, ca_len):
            col = chr(ord('C') + i)

            ws[col + '6'].fill = table_fill
            ws[col + '6'].font = name_font
            ws[col + '6'].alignment = c_c_alignment
            ws[col + '6'] = reporting_period_data['names'][i] + " (" + reporting_period_data['units'][i] + ")"
            ws[col + '6'].border = f_border

            ws[col + '7'].font = name_font
            ws[col + '7'].alignment = c_c_alignment
            ws[col + '7'] = round(reporting_period_data['subtotals'][i], 2)
            ws[col + '7'].border = f_border

            ws[col + '8'].font = name_font
            ws[col + '8'].alignment = c_c_alignment
            ws[col + '8'] = str(round(reporting_period_data['increment_rates'][i] * 100, 2)) + "%" \
                if reporting_period_data['increment_rates'][i] is not None else "-"
            ws[col + '8'].border = f_border

        col = chr(ord(col) + 1)

        ws[col + '6'].fill = table_fill
        ws[col + '6'].font = name_font
        ws[col + '6'].alignment = c_c_alignment
        ws[col + '6'] = "总计 (" + reporting_period_data['total_unit'] + ")"
        ws[col + '6'].border = f_border

        ws[col + '7'].font = name_font
        ws[col + '7'].alignment = c_c_alignment
        ws[col + '7'] = round(reporting_period_data['total'], 2)
        ws[col + '7'].border = f_border

        ws[col + '8'].font = name_font
        ws[col + '8'].alignment = c_c_alignment
        ws[col + '8'] = str(round(reporting_period_data['total_increment_rate'] * 100, 2)) + "%" \
            if reporting_period_data['total_increment_rate'] is not None else "-"
        ws[col + '8'].border = f_border

    else:
        for i in range(6, 8 + 1):
            ws.row_dimensions[i].height = 0.1
    ####################################################################################################################
    current_row_number = 10
    has_subtotals_data_flag = True
    if "subtotals" not in reporting_period_data.keys() or \
            reporting_period_data['subtotals'] is None or \
            len(reporting_period_data['subtotals']) == 0:
        has_subtotals_data_flag = False

    if has_subtotals_data_flag:
        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)] = name + ' 收入占比'

        current_row_number += 1

        table_start_row_number = current_row_number

        ws['B' + str(current_row_number)].fill = table_fill
        ws['B' + str(current_row_number)].font = name_font
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)].border = f_border

        ws['C' + str(current_row_number)].fill = table_fill
        ws['C' + str(current_row_number)].font = name_font
        ws['C' + str(current_row_number)].alignment = c_c_alignment
        ws['C' + str(current_row_number)].border = f_border
        ws['C' + str(current_row_number)] = '收入'

        ws['D' + str(current_row_number)].fill = table_fill
        ws['D' + str(current_row_number)].font = name_font
        ws['D' + str(current_row_number)].alignment = c_c_alignment
        ws['D' + str(current_row_number)].border = f_border
        ws['D' + str(current_row_number)] = '收入占比'

        current_row_number += 1

        ca_len = len(reporting_period_data['names'])
        income_sum = Decimal(0.0)
        for i in range(0, ca_len):
            income_sum = round(reporting_period_data['subtotals'][i], 2) + income_sum
        for i in range(0, ca_len):
            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)].alignment = c_c_alignment
            ws['B' + str(current_row_number)] = reporting_period_data['names'][i]
            ws['B' + str(current_row_number)].border = f_border

            ws['C' + str(current_row_number)].font = title_font
            ws['C' + str(current_row_number)].alignment = c_c_alignment
            ws['C' + str(current_row_number)].border = f_border
            ws['C' + str(current_row_number)] = round(reporting_period_data['subtotals'][i], 2)

            ws['D' + str(current_row_number)].font = title_font
            ws['D' + str(current_row_number)].alignment = c_c_alignment
            ws['D' + str(current_row_number)].border = f_border
            ws['D' + str(current_row_number)] = '{:.2%}'.format(round(
                reporting_period_data['subtotals'][i], 2) / income_sum) if income_sum is not None and \
                income_sum != Decimal(0.0) else " "

            current_row_number += 1

        table_end_row_number = current_row_number - 1

        pie = PieChart()
        pie.title = name + ' 收入占比'
        labels = Reference(ws, min_col=2, min_row=table_start_row_number + 1, max_row=table_end_row_number)
        pie_data = Reference(ws, min_col=3, min_row=table_start_row_number, max_row=table_end_row_number)
        pie.add_data(pie_data, titles_from_data=True)
        pie.set_categories(labels)
        pie.height = 6.6
        pie.width = 9
        s1 = pie.series[0]
        s1.dLbls = DataLabelList()
        s1.dLbls.showCatName = False
        s1.dLbls.showVal = True
        s1.dLbls.showPercent = True
        table_cell = 'F' + str(table_start_row_number - 1)
        ws.add_chart(pie, table_cell)

        if ca_len < 4:
            current_row_number = current_row_number - ca_len + 4

    else:
        for i in range(13, 22 + 1):
            ws.row_dimensions[i].height = 0.1

    ####################################################################################################################
    current_row_number = 14
    reporting_period_data = report['reporting_period']
    times = reporting_period_data['timestamps']
    has_detail_data_flag = True
    ca_len = len(report['reporting_period']['names'])
    real_timestamps_len = timestamps_data_not_equal_0(report['parameters']['timestamps'])
    table_row = (current_row_number + 1) + ca_len * 6 + real_timestamps_len * 7 + 1
    if "timestamps" not in reporting_period_data.keys() or \
            reporting_period_data['timestamps'] is None or \
            len(reporting_period_data['timestamps']) == 0:
        has_detail_data_flag = False

    if has_detail_data_flag:
        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)] = name + ' 详细数据'

        ws.row_dimensions[table_row].height = 60
        ws['B' + str(table_row)].fill = table_fill
        ws['B' + str(table_row)].font = title_font
        ws['B' + str(table_row)].border = f_border
        ws['B' + str(table_row)].alignment = c_c_alignment
        ws['B' + str(table_row)] = '日期时间'
        time = times[0]
        has_data = False
        max_row = 0
        if len(time) > 0:
            has_data = True
            max_row = table_row + len(time)

        if has_data:
            for i in range(0, len(time)):
                col = 'B'
                row = str(table_row + 1 + i)
                ws[col + row].font = title_font
                ws[col + row].alignment = c_c_alignment
                ws[col + row] = time[i]
                ws[col + row].border = f_border

            for i in range(0, ca_len):

                col = chr(ord('C') + i)

                ws[col + str(table_row)].fill = table_fill
                ws[col + str(table_row)].font = title_font
                ws[col + str(table_row)].alignment = c_c_alignment
                ws[col + str(table_row)] = reporting_period_data['names'][i] + " (" + reporting_period_data['units'][
                    i] + ")"
                ws[col + str(table_row)].border = f_border

                # 39 data
                time = times[i]
                time_len = len(time)

                for j in range(0, time_len):
                    row = str(table_row + 1 + j)
                    ws[col + row].font = title_font
                    ws[col + row].alignment = c_c_alignment
                    ws[col + row] = round(reporting_period_data['values'][i][j], 2)
                    ws[col + row].border = f_border

                line = LineChart()
                line.title = \
                    '报告期收入 - ' + reporting_period_data['names'][i] + " (" + reporting_period_data['units'][i] + ")"
                labels = Reference(ws, min_col=2, min_row=table_row + 1, max_row=max_row)
                line_data = Reference(ws, min_col=3 + i, min_row=table_row, max_row=max_row)
                line.add_data(line_data, titles_from_data=True)
                line.set_categories(labels)
                line_data = line.series[0]
                line_data.marker.symbol = "circle"
                line_data.smooth = True
                line.x_axis.crosses = 'min'
                line.height = 8.25
                line.width = 24
                line.dLbls = DataLabelList()
                line.dLbls.dLblPos = 't'
                line.dLbls.showVal = True
                line.dLbls.showPercent = False
                chart_col = 'B'
                chart_cell = chart_col + str(current_row_number + 1 + 6 * i)
                chart_start_row_number = current_row_number
                ws.add_chart(line, chart_cell)

            row = str(max_row + 1)

            ws['B' + row].font = title_font
            ws['B' + row].alignment = c_c_alignment
            ws['B' + row] = '小计'
            ws['B' + row].border = f_border

            col = ''

            for i in range(0, ca_len):
                col = chr(ord('C') + i)
                row = str(max_row + 1)
                ws[col + row].font = title_font
                ws[col + row].alignment = c_c_alignment
                ws[col + row] = round(reporting_period_data['subtotals'][i], 2)
                ws[col + row].border = f_border

            col = chr(ord(col) + 1)

            ws[col + str(table_row)].fill = table_fill
            ws[col + str(table_row)].font = title_font
            ws[col + str(table_row)].alignment = c_c_alignment
            ws[col + str(table_row)] = '总计 (' + report['reporting_period']['total_unit'] + ')'
            ws[col + str(table_row)].border = f_border

            total_sum = 0

            for j in range(0, len(time)):
                row = str(table_row + 1 + j)
                ws[col + row].font = title_font
                ws[col + row].alignment = c_c_alignment
                periodic_sum = reporting_period_values_periodic_sum(reporting_period_data, j, ca_len)
                total_sum += periodic_sum
                ws[col + row] = round(periodic_sum, 2)
                ws[col + row].border = f_border

            row = str(table_row + 1 + len(time))
            ws[col + row].font = title_font
            ws[col + row].alignment = c_c_alignment
            ws[col + row] = round(total_sum, 2)
            ws[col + row].border = f_border

    else:
        for i in range(37, 69 + 1):
            ws.row_dimensions[i].height = 0.1

    ####################################################################################################################

    has_associated_equipment_flag = True

    time_len = len(times[0])
    current_row_number = time_len + table_row + 3
    if "associated_equipment" not in report.keys() or \
            "energy_category_names" not in report['associated_equipment'].keys() or \
            len(report['associated_equipment']["energy_category_names"]) == 0 \
            or 'associated_equipment_names_array' not in report['associated_equipment'].keys() \
            or report['associated_equipment']['associated_equipment_names_array'] is None \
            or len(report['associated_equipment']['associated_equipment_names_array']) == 0 \
            or len(report['associated_equipment']['associated_equipment_names_array'][0]) == 0:
        has_associated_equipment_flag = False

    if has_associated_equipment_flag:
        associated_equipment = report['associated_equipment']

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)] = name + ' 相关设备数据'

        current_row_number += 1

        ws.row_dimensions[current_row_number].height = 60
        ws['B' + str(current_row_number)].fill = table_fill
        ws['B' + str(current_row_number)].font = name_font
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)].border = f_border
        ws['B' + str(current_row_number)] = '相关设备'
        ca_len = len(associated_equipment['energy_category_names'])

        for i in range(0, ca_len):
            col = chr(ord('C') + i)
            ws[col + str(current_row_number)].fill = table_fill
            ws[col + str(current_row_number)].font = name_font
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)].border = f_border
            ws[col + str(current_row_number)] = \
                reporting_period_data['names'][i] + " (" + reporting_period_data['units'][i] + ")"

        col_subtotal = chr(ord('C') + ca_len)
        ws[col_subtotal + str(current_row_number)].fill = table_fill
        ws[col_subtotal + str(current_row_number)].font = name_font
        ws[col_subtotal + str(current_row_number)].alignment = c_c_alignment
        ws[col_subtotal + str(current_row_number)].border = f_border
        ws[col_subtotal + str(current_row_number)] = '总计 (' + report['reporting_period']['total_unit'] + ')'

        associated_equipment_len = len(associated_equipment['associated_equipment_names_array'][0])

        for i in range(0, associated_equipment_len):
            current_row_number += 1
            row = str(current_row_number)

            ws['B' + row].font = title_font
            ws['B' + row].alignment = c_c_alignment
            ws['B' + row] = associated_equipment['associated_equipment_names_array'][0][i]
            ws['B' + row].border = f_border

            subtotal = Decimal(0.0)
            for j in range(0, ca_len):
                col = chr(ord('C') + j)
                ws[col + row].font = title_font
                ws[col + row].alignment = c_c_alignment
                ws[col + row] = round(associated_equipment['subtotals_array'][j][i], 2)
                ws[col + row].border = f_border

                subtotal += associated_equipment['subtotals_array'][j][i]

            ws[col_subtotal + row].font = title_font
            ws[col_subtotal + row].alignment = c_c_alignment
            ws[col_subtotal + row] = round(subtotal, 2)
            ws[col_subtotal + row].border = f_border
            print(subtotal)
    ####################################################################################################################
    current_sheet_parameters_row_number = chart_start_row_number + ca_len * 6 + 1
    has_parameters_names_and_timestamps_and_values_data = True
    if 'parameters' not in report.keys() or \
            report['parameters'] is None or \
            'names' not in report['parameters'].keys() or \
            report['parameters']['names'] is None or \
            len(report['parameters']['names']) == 0 or \
            'timestamps' not in report['parameters'].keys() or \
            report['parameters']['timestamps'] is None or \
            len(report['parameters']['timestamps']) == 0 or \
            'values' not in report['parameters'].keys() or \
            report['parameters']['values'] is None or \
            len(report['parameters']['values']) == 0 or \
            timestamps_data_all_equal_0(report['parameters']['timestamps']):

        has_parameters_names_and_timestamps_and_values_data = False
    if has_parameters_names_and_timestamps_and_values_data:

        ################################################################################################################
        # new worksheet
        ################################################################################################################

        parameters_data = report['parameters']
        parameters_names_len = len(parameters_data['names'])

        file_name = __file__.split('/')[-1].replace(".py", "")
        parameters_ws = wb.create_sheet(file_name + 'Parameters')

        parameters_timestamps_data_max_len = \
            get_parameters_timestamps_lists_max_len(list(parameters_data['timestamps']))

        # Row height
        parameters_ws.row_dimensions[1].height = 102
        for i in range(2, 7 + 1):
            parameters_ws.row_dimensions[i].height = 42

        for i in range(8, parameters_timestamps_data_max_len + 10):
            parameters_ws.row_dimensions[i].height = 60

        # Col width
        parameters_ws.column_dimensions['A'].width = 1.5

        parameters_ws.column_dimensions['B'].width = 25.0

        for i in range(3, 12+parameters_names_len*3):
            parameters_ws.column_dimensions[format_cell.get_column_letter(i)].width = 15.0

        # Img
        img = Image("excelexporters/myems.png")
        img.width = img.width * 0.85
        img.height = img.height * 0.85
        parameters_ws.add_image(img, 'B1')

        # Title
        parameters_ws.row_dimensions[3].height = 60

        parameters_ws['B3'].font = name_font
        parameters_ws['B3'].alignment = b_r_alignment
        parameters_ws['B3'] = 'Name:'
        parameters_ws['C3'].border = b_border
        parameters_ws['C3'].alignment = b_c_alignment
        parameters_ws['C3'].font = name_font
        parameters_ws['C3'] = name

        parameters_ws['D3'].font = name_font
        parameters_ws['D3'].alignment = b_r_alignment
        parameters_ws['D3'] = 'Period:'
        parameters_ws['E3'].border = b_border
        parameters_ws['E3'].alignment = b_c_alignment
        parameters_ws['E3'].font = name_font
        parameters_ws['E3'] = period_type

        parameters_ws['F3'].font = name_font
        parameters_ws['F3'].alignment = b_r_alignment
        parameters_ws['F3'] = 'Date:'
        parameters_ws['G3'].border = b_border
        parameters_ws['G3'].alignment = b_c_alignment
        parameters_ws['G3'].font = name_font
        parameters_ws['G3'] = reporting_start_datetime_local + "__" + reporting_end_datetime_local
        parameters_ws.merge_cells("G3:H3")

        parameters_ws_current_row_number = 6

        parameters_ws['B' + str(parameters_ws_current_row_number)].font = title_font
        parameters_ws['B' + str(parameters_ws_current_row_number)] = name + ' Parameters'

        parameters_ws_current_row_number += 1

        parameters_table_start_row_number = parameters_ws_current_row_number

        parameters_ws.row_dimensions[parameters_ws_current_row_number].height = 80

        parameters_ws_current_row_number += 1

        table_current_col_number = 2

        for i in range(0, parameters_names_len):

            if len(parameters_data['timestamps'][i]) == 0:
                continue

            col = format_cell.get_column_letter(table_current_col_number)

            parameters_ws[col + str(parameters_ws_current_row_number-1)].fill = table_fill
            parameters_ws[col + str(parameters_ws_current_row_number-1)].border = f_border

            col = format_cell.get_column_letter(table_current_col_number + 1)

            parameters_ws[col + str(parameters_ws_current_row_number-1)].fill = table_fill
            parameters_ws[col + str(parameters_ws_current_row_number-1)].border = f_border
            parameters_ws[col + str(parameters_ws_current_row_number-1)].font = name_font
            parameters_ws[col + str(parameters_ws_current_row_number-1)].alignment = c_c_alignment
            parameters_ws[col + str(parameters_ws_current_row_number-1)] = parameters_data['names'][i]

            table_current_row_number = parameters_ws_current_row_number

            for j, value in enumerate(list(parameters_data['timestamps'][i])):
                col = format_cell.get_column_letter(table_current_col_number)

                parameters_ws[col + str(table_current_row_number)].border = f_border
                parameters_ws[col + str(table_current_row_number)].font = title_font
                parameters_ws[col + str(table_current_row_number)].alignment = c_c_alignment
                parameters_ws[col + str(table_current_row_number)] = value

                col = format_cell.get_column_letter(table_current_col_number + 1)

                parameters_ws[col + str(table_current_row_number)].border = f_border
                parameters_ws[col + str(table_current_row_number)].font = title_font
                parameters_ws[col + str(table_current_row_number)].alignment = c_c_alignment
                parameters_ws[col + str(table_current_row_number)] = round(parameters_data['values'][i][j], 2)

                table_current_row_number += 1

            table_current_col_number = table_current_col_number + 3

        ################################################################################################################
        # parameters chart and parameters table
        ################################################################################################################

        ws['B' + str(current_sheet_parameters_row_number)].font = title_font
        ws['B' + str(current_sheet_parameters_row_number)] = name + ' Parameters'

        current_sheet_parameters_row_number += 1

        chart_start_row_number = current_sheet_parameters_row_number

        col_index = 0

        for i in range(0, parameters_names_len):

            if len(parameters_data['timestamps'][i]) == 0:
                continue

            line = LineChart()
            data_col = 3+col_index*3
            labels_col = 2+col_index*3
            col_index += 1
            line.title = 'Parameters - ' + \
                         parameters_ws.cell(row=parameters_table_start_row_number, column=data_col).value
            labels = Reference(parameters_ws, min_col=labels_col, min_row=parameters_table_start_row_number + 1,
                               max_row=(len(parameters_data['timestamps'][i])+parameters_table_start_row_number))
            line_data = Reference(parameters_ws, min_col=data_col, min_row=parameters_table_start_row_number,
                                  max_row=(len(parameters_data['timestamps'][i])+parameters_table_start_row_number))
            line.add_data(line_data, titles_from_data=True)
            line.set_categories(labels)
            line_data = line.series[0]
            line_data.marker.symbol = "circle"
            line_data.smooth = True
            line.x_axis.crosses = 'min'
            line.height = 8.25
            line.width = 24
            line.dLbls = DataLabelList()
            line.dLbls.dLblPos = 't'
            line.dLbls.showVal = False
            line.dLbls.showPercent = False
            chart_col = 'B'
            chart_cell = chart_col + str(chart_start_row_number)
            chart_start_row_number += 6
            ws.add_chart(line, chart_cell)

        current_sheet_parameters_row_number = chart_start_row_number

        current_sheet_parameters_row_number += 1
    ####################################################################################################################
    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename


def reporting_period_values_periodic_sum(reporting_period_data, periodic_index, ca_len):
    periodic_sum = 0
    for i in range(0, ca_len):
        periodic_sum += reporting_period_data['values'][i][periodic_index]

    return periodic_sum


def get_parameters_timestamps_lists_max_len(parameters_timestamps_lists):
    max_len = 0
    for i, value in enumerate(list(parameters_timestamps_lists)):
        if len(value) > max_len:
            max_len = len(value)

    return max_len


def timestamps_data_all_equal_0(lists):
    for i, value in enumerate(list(lists)):
        if len(value) > 0:
            return False

    return True


def timestamps_data_not_equal_0(lists):
    number = 0
    for i, value in enumerate(list(lists)):
        if len(value) > 0:
            number += 1
    return number
