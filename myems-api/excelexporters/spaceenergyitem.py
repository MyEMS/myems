import base64
import uuid
import os
from openpyxl.chart import (
    PieChart,
    LineChart,
    BarChart,
    Reference,
)
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl.chart.label import DataLabelList


####################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file
# Step 3: Encode the excel file bytes to Base64
####################################################################################################################


def export(report,
           name,
           reporting_start_datetime_local,
           reporting_end_datetime_local,
           period_type):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if report is None:
        return None
    print(report)

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(report,
                              name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              period_type)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        pass

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        pass
    return base64_message


def generate_excel(report,
                   name,
                   reporting_start_datetime_local,
                   reporting_end_datetime_local,
                   period_type):
    wb = Workbook()
    ws = wb.active

    # Row height
    ws.row_dimensions[1].height = 102
    for i in range(2, 2000 + 1):
        ws.row_dimensions[i].height = 42

    # Col width
    ws.column_dimensions['A'].width = 1.5

    ws.column_dimensions['B'].width = 25.0

    for i in range(ord('C'), ord('L')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Constantia', size=15, bold=True)
    title_font = Font(name='宋体', size=15, bold=True)
    data_font = Font(name='Franklin Gothic Book', size=11)

    table_fill = PatternFill(fill_type='solid', fgColor='1F497D')
    f_border = Border(left=Side(border_style='medium', color='00000000'),
                      right=Side(border_style='medium', color='00000000'),
                      bottom=Side(border_style='medium', color='00000000'),
                      top=Side(border_style='medium', color='00000000')
                      )
    b_border = Border(
        bottom=Side(border_style='medium', color='00000000'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_r_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    # Img
    img = Image("excelexporters/myems.png")
    img.width = img.width * 0.85
    img.height = img.height * 0.85
    # img = Image("myems.png")
    ws.add_image(img, 'B1')

    # Title
    ws.row_dimensions[3].height = 60

    ws['B3'].font = name_font
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = 'Name:'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'].font = name_font
    ws['C3'] = name

    ws['D3'].font = name_font
    ws['D3'].alignment = b_r_alignment
    ws['D3'] = 'Period:'
    ws['E3'].border = b_border
    ws['E3'].alignment = b_c_alignment
    ws['E3'].font = name_font
    ws['E3'] = period_type

    ws['F3'].font = name_font
    ws['F3'].alignment = b_r_alignment
    ws['F3'] = 'Date:'
    ws['G3'].border = b_border
    ws['G3'].alignment = b_c_alignment
    ws['G3'].font = name_font
    ws['G3'] = reporting_start_datetime_local + "__" + reporting_end_datetime_local
    ws.merge_cells("G3:H3")

    if "reporting_period" not in report.keys() or \
            "names" not in report['reporting_period'].keys() or len(report['reporting_period']['names']) == 0:
        filename = str(uuid.uuid4()) + '.xlsx'
        wb.save(filename)

        return filename

    ##################################

    current_row_number = 6

    reporting_period_data = report['reporting_period']

    has_names_data_flag = True

    if "names" not in reporting_period_data.keys() or \
            reporting_period_data['names'] is None or \
            len(reporting_period_data['names']) == 0:
        has_names_data_flag = False

    if has_names_data_flag:
        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)] = name + ' 报告期消耗'

        current_row_number += 1

        category = reporting_period_data['names']
        ca_len = len(category)

        ws.row_dimensions[current_row_number].height = 60
        ws['B' + str(current_row_number)].fill = table_fill
        ws['B' + str(current_row_number)].border = f_border

        col = 'C'

        for i in range(0, ca_len):
            ws[col + str(current_row_number)].fill = table_fill
            ws[col + str(current_row_number)].font = name_font
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)].border = f_border
            ws[col + str(current_row_number)] = \
                reporting_period_data['names'][i] + " " + reporting_period_data['energy_category_names'][i] + \
                " (" + reporting_period_data['units'][i] + ")"

            col = chr(ord(col) + 1)

        current_row_number += 1

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)].border = f_border
        ws['B' + str(current_row_number)] = '消耗'

        col = 'C'

        for i in range(0, ca_len):
            ws[col + str(current_row_number)].font = name_font
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)].border = f_border
            ws[col + str(current_row_number)] = round(reporting_period_data['subtotals'][i], 2)

            col = chr(ord(col) + 1)

        current_row_number += 1

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)].border = f_border
        ws['B' + str(current_row_number)] = '单位面积值'

        col = 'C'

        for i in range(0, ca_len):
            ws[col + str(current_row_number)].font = name_font
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)].border = f_border
            ws[col + str(current_row_number)] = round(reporting_period_data['subtotals_per_unit_area'][i], 2)

            col = chr(ord(col) + 1)

        current_row_number += 1

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)].border = f_border
        ws['B' + str(current_row_number)] = '环比'

        col = 'C'

        for i in range(0, ca_len):
            ws[col + str(current_row_number)].font = name_font
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)].border = f_border
            ws[col + str(current_row_number)] = str(
                round(reporting_period_data['increment_rates'][i] * 100, 2)) + '%' \
                if reporting_period_data['increment_rates'][i] is not None else '-'

            col = chr(ord(col) + 1)

        current_row_number += 2

        category_dict = group_by_category(reporting_period_data['energy_category_names'])

        for category_dict_name, category_dict_values in category_dict.items():

            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)] = \
                name + ' ' + category_dict_name + ' (' + reporting_period_data['units'][category_dict_values[0]] + \
                ') 分项消耗占比'

            current_row_number += 1
            table_start_row_number = current_row_number

            ws['B' + str(current_row_number)].fill = table_fill
            ws['B' + str(current_row_number)].border = f_border

            ws['C' + str(current_row_number)].font = name_font
            ws['C' + str(current_row_number)].fill = table_fill
            ws['C' + str(current_row_number)].alignment = c_c_alignment
            ws['C' + str(current_row_number)].border = f_border
            ws['C' + str(current_row_number)] = '消耗'

            current_row_number += 1

            for i in category_dict_values:
                ws['B' + str(current_row_number)].font = title_font
                ws['B' + str(current_row_number)].alignment = c_c_alignment
                ws['B' + str(current_row_number)].border = f_border
                ws['B' + str(current_row_number)] = \
                    reporting_period_data['names'][i] + " (" + reporting_period_data['units'][i] + ")"
                ws['C' + str(current_row_number)].font = name_font
                ws['C' + str(current_row_number)].alignment = c_c_alignment
                ws['C' + str(current_row_number)].border = f_border
                ws['C' + str(current_row_number)] = round(reporting_period_data['subtotals'][i], 3)

                current_row_number += 1

            table_end_row_number = current_row_number - 1

            pie = PieChart()
            pie.title = \
                name + ' ' + category_dict_name + ' (' + reporting_period_data['units'][category_dict_values[0]] + \
                ') 分项消耗占比'
            labels = Reference(ws, min_col=2, min_row=table_start_row_number + 1, max_row=table_end_row_number)
            pie_data = Reference(ws, min_col=3, min_row=table_start_row_number, max_row=table_end_row_number)
            pie.add_data(pie_data, titles_from_data=True)
            pie.set_categories(labels)
            pie.height = 6.6
            pie.width = 9
            s1 = pie.series[0]
            s1.dLbls = DataLabelList()
            s1.dLbls.showCatName = False
            s1.dLbls.showVal = True
            s1.dLbls.showPercent = True
            ws.add_chart(pie, 'D' + str(table_start_row_number))

            if len(category_dict_values) < 4:
                current_row_number = current_row_number - len(category_dict_values) + 4

            current_row_number += 1

        #####################################

        has_child_flag = True

        if "child_space" not in report.keys() or "energy_item_names" not in report['child_space'].keys() or \
                len(report['child_space']["energy_item_names"]) == 0 \
                or 'child_space_names_array' not in report['child_space'].keys() \
                or report['child_space']['child_space_names_array'] is None \
                or len(report['child_space']['child_space_names_array']) == 0 \
                or len(report['child_space']['child_space_names_array'][0]) == 0:
            has_child_flag = False

        if has_child_flag:
            child = report['child_space']

            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)] = name + ' 子空间数据'

            current_row_number += 1
            table_start_row_number = current_row_number

            ws.row_dimensions[current_row_number].height = 60
            ws['B' + str(current_row_number)].fill = table_fill
            ws['B' + str(current_row_number)].border = f_border
            ca_len = len(child['energy_item_names'])

            for i in range(0, ca_len):
                row = chr(ord('C') + i)
                ws[row + str(current_row_number)].fill = table_fill
                ws[row + str(current_row_number)].font = name_font
                ws[row + str(current_row_number)].alignment = c_c_alignment
                ws[row + str(current_row_number)].border = f_border
                ws[row + str(current_row_number)] = \
                    reporting_period_data['names'][i] + " (" + reporting_period_data['units'][i] + ")"

            space_len = len(child['child_space_names_array'][0])

            for i in range(0, space_len):
                current_row_number += 1
                row = str(current_row_number)

                ws['B' + row].font = name_font
                ws['B' + row].alignment = c_c_alignment
                ws['B' + row] = child['child_space_names_array'][0][i]
                ws['B' + row].border = f_border

                for j in range(0, ca_len):
                    col = chr(ord('C') + j)
                    ws[col + row].font = name_font
                    ws[col + row].alignment = c_c_alignment
                    ws[col + row] = round(child['subtotals_array'][j][i], 2)
                    ws[col + row].border = f_border

            table_end_row_number = current_row_number
            current_row_number += 1
            pie_start_row_number = current_row_number

            # Pie
            for i in range(0, ca_len):
                pie = PieChart()
                labels = Reference(ws, min_col=2, min_row=table_start_row_number + 1,
                                   max_row=table_end_row_number)
                pie_data = Reference(ws, min_col=3 + i, min_row=table_start_row_number,
                                     max_row=table_end_row_number)
                pie.add_data(pie_data, titles_from_data=True)
                pie.set_categories(labels)
                pie.height = 6.6
                pie.width = 8
                pie.title = ws.cell(column=3 + i, row=table_start_row_number).value
                s1 = pie.series[0]
                s1.dLbls = DataLabelList()
                s1.dLbls.showCatName = False
                s1.dLbls.showVal = True
                s1.dLbls.showPercent = True
                chart_cell = ''
                if i % 2 == 0:
                    chart_cell = 'B' + str(pie_start_row_number)
                else:
                    chart_cell = 'E' + str(pie_start_row_number)
                    pie_start_row_number += 5
                ws.add_chart(pie, chart_cell)

            current_row_number = pie_start_row_number
            if ca_len % 2 == 1:
                current_row_number += 5

            current_row_number += 1
        #######################

        has_values_data = True
        has_timestamps_data = True

        if 'values' not in reporting_period_data.keys() or \
                reporting_period_data['values'] is None or \
                len(reporting_period_data['values']) == 0:
            has_values_data = False

        if 'timestamps' not in reporting_period_data.keys() or \
                reporting_period_data['timestamps'] is None or \
                len(reporting_period_data['timestamps']) == 0 or \
                len(reporting_period_data['timestamps'][0]) == 0:
            has_timestamps_data = False

        if has_values_data and has_timestamps_data:
            ca_len = len(reporting_period_data['names'])
            time = reporting_period_data['timestamps'][0]

            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)] = name + ' 详细数据'

            current_row_number += 1

            chart_start_row_number = current_row_number

            current_row_number += ca_len * 6
            table_start_row_number = current_row_number

            ws.row_dimensions[current_row_number].height = 60
            ws['B' + str(current_row_number)].fill = table_fill
            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)].alignment = c_c_alignment
            ws['B' + str(current_row_number)].border = f_border
            ws['B' + str(current_row_number)] = '日期时间'

            col = 'C'

            for i in range(0, ca_len):
                ws[col + str(current_row_number)].fill = table_fill
                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)].border = f_border
                ws[col + str(current_row_number)] = \
                    reporting_period_data['names'][i] + " (" + reporting_period_data['units'][i] + ")"
                col = chr(ord(col) + 1)

            current_row_number += 1

            for i in range(0, len(time)):
                ws['B' + str(current_row_number)].font = title_font
                ws['B' + str(current_row_number)].alignment = c_c_alignment
                ws['B' + str(current_row_number)].border = f_border
                ws['B' + str(current_row_number)] = time[i]

                col = 'C'
                for j in range(0, ca_len):
                    ws[col + str(current_row_number)].font = title_font
                    ws[col + str(current_row_number)].alignment = c_c_alignment
                    ws[col + str(current_row_number)].border = f_border
                    ws[col + str(current_row_number)] = round(reporting_period_data['values'][j][i], 2) \
                        if reporting_period_data['values'][j][i] is not None else 0.00
                    col = chr(ord(col) + 1)

                current_row_number += 1

            table_end_row_number = current_row_number - 1

            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)].alignment = c_c_alignment
            ws['B' + str(current_row_number)].border = f_border
            ws['B' + str(current_row_number)] = '小计'

            col = 'C'

            for i in range(0, ca_len):
                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)].border = f_border
                ws[col + str(current_row_number)] = round(reporting_period_data['subtotals'][i], 2)
                col = chr(ord(col) + 1)

            current_row_number += 2

            format_time_width_number = 1.0
            min_len_number = 1.0
            min_width_number = 11.0  # format_time_width_number * min_len_number + 4 and min_width_number > 11.0

            if period_type == 'hourly':
                format_time_width_number = 4.0
                min_len_number = 2
                min_width_number = 12.0
            elif period_type == 'daily':
                format_time_width_number = 2.5
                min_len_number = 4
                min_width_number = 14.0
            elif period_type == 'monthly':
                format_time_width_number = 2.1
                min_len_number = 4
                min_width_number = 12.4
            elif period_type == 'yearly':
                format_time_width_number = 1.5
                min_len_number = 5
                min_width_number = 11.5

            for i in range(0, ca_len):
                line = LineChart()
                line.title = '报告期消耗 - ' + \
                             reporting_period_data['names'][i] + " (" + reporting_period_data['units'][i] + ")"
                labels = Reference(ws, min_col=2, min_row=table_start_row_number + 1, max_row=table_end_row_number)
                line_data = Reference(ws, min_col=3 + i, min_row=table_start_row_number, max_row=table_end_row_number)
                line.add_data(line_data, titles_from_data=True)
                line.set_categories(labels)
                line_data = line.series[0]
                line_data.marker.symbol = "circle"
                line_data.smooth = True
                line.x_axis.crosses = 'min'
                line.height = 8.25
                line.width = format_time_width_number * len(time) if len(time) > min_len_number else min_width_number
                if line.width > 24:
                    line.width = 24
                line.dLbls = DataLabelList()
                line.dLbls.dLblPos = 't'
                line.dLbls.showVal = True
                line.dLbls.showPercent = False
                chart_col = 'B'
                chart_cell = chart_col + str(chart_start_row_number)
                chart_start_row_number += 6
                ws.add_chart(line, chart_cell)

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename


def group_by_category(category_list):
    category_dict = dict()
    for i, value in enumerate(category_list):
        if value not in category_dict.keys():
            category_dict[value] = list()
        category_dict[value].append(i)
    return category_dict
