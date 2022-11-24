import base64
import gettext
import os
import uuid

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font


########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file from the report data
# Step 3: Encode the excel file to Base64
########################################################################################################################

def export(result, space_name, reporting_start_datetime_local, reporting_end_datetime_local, language):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if result is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(result,
                              space_name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              language)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        print(str(ex))

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        print(str(ex))
    return base64_message


def generate_excel(report, space_name, reporting_start_datetime_local, reporting_end_datetime_local, language):

    locale_path = './i18n/'
    if language == 'zh_CN':
        trans = gettext.translation('myems', locale_path, languages=['zh_CN'])
    elif language == 'de':
        trans = gettext.translation('myems', locale_path, languages=['de'])
    elif language == 'en':
        trans = gettext.translation('myems', locale_path, languages=['en'])
    else:
        trans = gettext.translation('myems', locale_path, languages=['en'])
    trans.install()
    _ = trans.gettext

    wb = Workbook()
    ws = wb.active
    ws.title = "OfflineMeterBatch"

    # Col width
    for i in range(ord('A'), ord('L')):
        ws.column_dimensions[chr(i)].width = 20.0

    # Head image
    ws.row_dimensions[1].height = 105
    img = Image("excelexporters/myems.png")
    ws.add_image(img, 'A1')

    # Query Parameters
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    ws['A3'].alignment = b_r_alignment
    ws['A3'] = _('Space') + ':'
    ws['B3'] = space_name
    ws['A4'].alignment = b_r_alignment
    ws['A4'] = _('Start Datetime') + ':'
    ws['B4'] = reporting_start_datetime_local
    ws['A5'].alignment = b_r_alignment
    ws['A5'] = _('End Datetime') + ':'
    ws['B5'] = reporting_end_datetime_local

    # Title
    title_font = Font(size=12, bold=True)
    ws['A6'].font = title_font
    ws['A6'] = _('ID')
    ws['B6'].font = title_font
    ws['B6'] = _('Name')
    ws['C6'].font = title_font
    ws['C6'] = _('Space')

    ca_len = len(report['energycategories'])
    for i in range(0, ca_len):
        col = chr(ord('D') + i)
        ws[col + '6'].font = title_font
        ws[col + '6'] = report['energycategories'][i]['name'] + " (" + \
            report['energycategories'][i]['unit_of_measure'] + ")"

    current_row_number = 7
    for i in range(0, len(report['offline_meters'])):
        ws['A' + str(current_row_number)] = str(report['offline_meters'][i]['id'])
        ws['B' + str(current_row_number)] = report['offline_meters'][i]['offline_meter_name']
        ws['C' + str(current_row_number)] = report['offline_meters'][i]['space_name']
        ca_len = len(report['offline_meters'][i]['values'])
        for j in range(0, ca_len):
            col = chr(ord('D') + j)
            ws[col + str(current_row_number)] = report['offline_meters'][i]['values'][j]

        current_row_number += 1

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename
