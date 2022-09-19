import base64
import os
import uuid
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import openpyxl.utils.cell as format_cell
import gettext

########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file
# Step 3: Encode the excel file to Base64
########################################################################################################################


def export(report,
           name,
           reporting_start_datetime_local,
           reporting_end_datetime_local,
           period_type,
           language):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if report is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(report,
                              name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              period_type,
                              language)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        pass

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        pass
    return base64_message


def generate_excel(report,
                   name,
                   reporting_start_datetime_local,
                   reporting_end_datetime_local,
                   period_type,
                   language):

    locale_path = './i18n/'
    if language == 'zh_CN':
        trans = gettext.translation('myems', locale_path, languages=['zh_CN'])
    elif language == 'de':
        trans = gettext.translation('myems', locale_path, languages=['de'])
    elif language == 'en':
        trans = gettext.translation('myems', locale_path, languages=['en'])
    else:
        trans = gettext.translation('myems', locale_path, languages=['en'])
    trans.install()
    _ = trans.gettext

    wb = Workbook()
    ws = wb.active
    ws.title = "StoreStatistics"

    # Row height
    ws.row_dimensions[1].height = 102

    for i in range(2, 2000 + 1):
        ws.row_dimensions[i].height = 42

    # Col width
    ws.column_dimensions['A'].width = 1.5
    ws.column_dimensions['B'].width = 25.0

    for i in range(ord('C'), ord('L')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Arial', size=15, bold=True)
    title_font = Font(name='Arial', size=15, bold=True)

    table_fill = PatternFill(fill_type='solid', fgColor='1F497D')
    f_border = Border(left=Side(border_style='medium', color='00000000'),
                      right=Side(border_style='medium', color='00000000'),
                      bottom=Side(border_style='medium', color='00000000'),
                      top=Side(border_style='medium', color='00000000')
                      )
    b_border = Border(
        bottom=Side(border_style='medium', color='00000000'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)

    # Img
    img = Image("excelexporters/myems.png")
    ws.add_image(img, 'A1')

    # Title
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = _('Name') + ':'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'] = name

    ws['D3'].alignment = b_r_alignment
    ws['D3'] = _('Period Type') + ':'
    ws['E3'].border = b_border
    ws['E3'].alignment = b_c_alignment
    ws['E3'] = period_type

    ws['B4'].alignment = b_r_alignment
    ws['B4'] = _('Reporting Start Datetime') + ':'
    ws['C4'].border = b_border
    ws['C4'].alignment = b_c_alignment
    ws['C4'] = reporting_start_datetime_local

    ws['D4'].alignment = b_r_alignment
    ws['D4'] = _('Reporting End Datetime') + ':'
    ws['E4'].border = b_border
    ws['E4'].alignment = b_c_alignment
    ws['E4'] = reporting_end_datetime_local

    if "reporting_period" not in report.keys() or \
            "names" not in report['reporting_period'].keys() or len(report['reporting_period']['names']) == 0:
        filename = str(uuid.uuid4()) + '.xlsx'
        wb.save(filename)

        return filename
    ####################################################################################################################
    # First: Statistics
    # 6: title
    # 7: table title
    # 8~ca_len table_data
    ####################################################################################################################
    reporting_period_data = report['reporting_period']

    if "names" not in reporting_period_data.keys() or \
            reporting_period_data['names'] is None or \
            len(reporting_period_data['names']) == 0:
        filename = str(uuid.uuid4()) + '.xlsx'
        wb.save(filename)
        return filename

    ws['B6'].font = title_font
    ws['B6'] = name + ' ' + _('Statistics')

    category = reporting_period_data['names']

    # table_title
    ws['B7'].fill = table_fill
    ws['B7'].font = title_font
    ws['B7'].alignment = c_c_alignment
    ws['B7'] = _('Reporting Period')
    ws['B7'].border = f_border

    ws['C7'].font = title_font
    ws['C7'].alignment = c_c_alignment
    ws['C7'] = _('Arithmetic Mean')
    ws['C7'].border = f_border

    ws['D7'].font = title_font
    ws['D7'].alignment = c_c_alignment
    ws['D7'] = _('Median (Middle Value)')
    ws['D7'].border = f_border

    ws['E7'].font = title_font
    ws['E7'].alignment = c_c_alignment
    ws['E7'] = _('Minimum Value')
    ws['E7'].border = f_border

    ws['F7'].font = title_font
    ws['F7'].alignment = c_c_alignment
    ws['F7'] = _('Maximum Value')
    ws['F7'].border = f_border

    ws['G7'].font = title_font
    ws['G7'].alignment = c_c_alignment
    ws['G7'] = _('Sample Standard Deviation')
    ws['G7'].border = f_border

    ws['H7'].font = title_font
    ws['H7'].alignment = c_c_alignment
    ws['H7'] = _('Sample Variance')
    ws['H7'].border = f_border

    # table_data

    for i, value in enumerate(category):
        row = i * 2 + 8
        ws['B' + str(row)].font = name_font
        ws['B' + str(row)].alignment = c_c_alignment
        ws['B' + str(row)] = reporting_period_data['names'][i] + " (" + reporting_period_data['units'][i] + " )"
        ws['B' + str(row)].border = f_border

        ws['B' + str(row + 1)].font = name_font
        ws['B' + str(row + 1)].alignment = c_c_alignment
        ws['B' + str(row + 1)] = _('Increment Rate')
        ws['B' + str(row + 1)].border = f_border

        ws['C' + str(row)].font = name_font
        ws['C' + str(row)].alignment = c_c_alignment
        ws['C' + str(row)] = round(reporting_period_data['means'][i], 2) \
            if reporting_period_data['means'][i] is not None else ''
        ws['C' + str(row)].border = f_border
        ws['C' + str(row)].number_format = '0.00'

        ws['C' + str(row + 1)].font = name_font
        ws['C' + str(row + 1)].alignment = c_c_alignment
        ws['C' + str(row + 1)] = str(round(reporting_period_data['means_increment_rate'][i] * 100, 2)) + "%" \
            if reporting_period_data['means_increment_rate'][i] is not None else '0.00%'
        ws['C' + str(row + 1)].border = f_border

        ws['D' + str(row)].font = name_font
        ws['D' + str(row)].alignment = c_c_alignment
        ws['D' + str(row)] = round(reporting_period_data['medians'][i], 2) \
            if reporting_period_data['medians'][i] is not None else ''
        ws['D' + str(row)].border = f_border
        ws['D' + str(row)].number_format = '0.00'

        ws['D' + str(row + 1)].font = name_font
        ws['D' + str(row + 1)].alignment = c_c_alignment
        ws['D' + str(row + 1)] = str(round(reporting_period_data['medians_increment_rate'][i] * 100, 2)) + "%" \
            if reporting_period_data['medians_increment_rate'][i] is not None else '0.00%'
        ws['D' + str(row + 1)].border = f_border

        ws['E' + str(row)].font = name_font
        ws['E' + str(row)].alignment = c_c_alignment
        ws['E' + str(row)] = round(reporting_period_data['minimums'][i], 2) \
            if reporting_period_data['minimums'][i] is not None else ''
        ws['E' + str(row)].border = f_border
        ws['E' + str(row)].number_format = '0.00'

        ws['E' + str(row + 1)].font = name_font
        ws['E' + str(row + 1)].alignment = c_c_alignment
        ws['E' + str(row + 1)] = str(round(reporting_period_data['minimums_increment_rate'][i] * 100, 2)) + "%" \
            if reporting_period_data['minimums_increment_rate'][i] is not None else '0.00%'
        ws['E' + str(row + 1)].border = f_border

        ws['F' + str(row)].font = name_font
        ws['F' + str(row)].alignment = c_c_alignment
        ws['F' + str(row)] = round(reporting_period_data['maximums'][i], 2) \
            if reporting_period_data['maximums'][i] is not None else ''
        ws['F' + str(row)].border = f_border
        ws['F' + str(row)].number_format = '0.00'

        ws['F' + str(row + 1)].font = name_font
        ws['F' + str(row + 1)].alignment = c_c_alignment
        ws['F' + str(row + 1)] = str(round(reporting_period_data['maximums_increment_rate'][i] * 100, 2)) + "%" \
            if reporting_period_data['maximums_increment_rate'][i] is not None else '0.00%'
        ws['F' + str(row + 1)].border = f_border

        ws['G' + str(row)].font = name_font
        ws['G' + str(row)].alignment = c_c_alignment
        ws['G' + str(row)] = round(reporting_period_data['stdevs'][i], 2) \
            if reporting_period_data['stdevs'][i] is not None else ''
        ws['G' + str(row)].border = f_border
        ws['G' + str(row)].number_format = '0.00'

        ws['G' + str(row + 1)].font = name_font
        ws['G' + str(row + 1)].alignment = c_c_alignment
        ws['G' + str(row + 1)] = str(round(reporting_period_data['stdevs_increment_rate'][i] * 100, 2)) + "%" \
            if reporting_period_data['stdevs_increment_rate'][i] is not None else '0.00%'
        ws['G' + str(row + 1)].border = f_border

        ws['H' + str(row)].font = name_font
        ws['H' + str(row)].alignment = c_c_alignment
        ws['H' + str(row)] = round(reporting_period_data['variances'][i], 2) \
            if reporting_period_data['variances'][i] is not None else ''
        ws['H' + str(row)].border = f_border
        ws['H' + str(row)].number_format = '0.00'

        ws['H' + str(row + 1)].font = name_font
        ws['H' + str(row + 1)].alignment = c_c_alignment
        ws['H' + str(row + 1)] = str(round(reporting_period_data['variances_increment_rate'][i] * 100, 2)) + "%" \
            if reporting_period_data['variances_increment_rate'][i] is not None else '0.00%'
        ws['H' + str(row + 1)].border = f_border
    ####################################################################################################################
    # Second: Reporting Period Consumption
    # 9 + ca_len * 2: title
    # 10 + ca_len * 2: table title
    # per_unit_area_start_row_number + 2 ~ per_unit_area_start_row_number + 2 + ca_len :  table_data
    ####################################################################################################################

    names = reporting_period_data['names']
    ca_len = len(names)

    per_unit_area_start_row_number = 9 + ca_len * 2

    ws['B' + str(per_unit_area_start_row_number)].font = title_font
    ws['B' + str(per_unit_area_start_row_number)] = name + ' ' + _('Per Unit Area') + str(report['store']['area']) + 'M²'

    category = reporting_period_data['names']

    # table_title
    ws['B' + str(per_unit_area_start_row_number + 1)].fill = table_fill
    ws['B' + str(per_unit_area_start_row_number + 1)].font = title_font
    ws['B' + str(per_unit_area_start_row_number + 1)].alignment = c_c_alignment
    ws['B' + str(per_unit_area_start_row_number + 1)] = _('Reporting Period')
    ws['B' + str(per_unit_area_start_row_number + 1)].border = f_border

    ws['C' + str(per_unit_area_start_row_number + 1)].font = title_font
    ws['C' + str(per_unit_area_start_row_number + 1)].alignment = c_c_alignment
    ws['C' + str(per_unit_area_start_row_number + 1)] = _('Arithmetic Mean')
    ws['C' + str(per_unit_area_start_row_number + 1)].border = f_border

    ws['D' + str(per_unit_area_start_row_number + 1)].font = title_font
    ws['D' + str(per_unit_area_start_row_number + 1)].alignment = c_c_alignment
    ws['D' + str(per_unit_area_start_row_number + 1)] = _('Median (Middle Value)')
    ws['D' + str(per_unit_area_start_row_number + 1)].border = f_border

    ws['E' + str(per_unit_area_start_row_number + 1)].font = title_font
    ws['E' + str(per_unit_area_start_row_number + 1)].alignment = c_c_alignment
    ws['E' + str(per_unit_area_start_row_number + 1)] = _('Minimum Value')
    ws['E' + str(per_unit_area_start_row_number + 1)].border = f_border

    ws['F' + str(per_unit_area_start_row_number + 1)].font = title_font
    ws['F' + str(per_unit_area_start_row_number + 1)].alignment = c_c_alignment
    ws['F' + str(per_unit_area_start_row_number + 1)] = _('Maximum Value')
    ws['F' + str(per_unit_area_start_row_number + 1)].border = f_border

    ws['G' + str(per_unit_area_start_row_number + 1)].font = title_font
    ws['G' + str(per_unit_area_start_row_number + 1)].alignment = c_c_alignment
    ws['G' + str(per_unit_area_start_row_number + 1)] = _('Sample Standard Deviation')
    ws['G' + str(per_unit_area_start_row_number + 1)].border = f_border

    ws['H' + str(per_unit_area_start_row_number + 1)].font = title_font
    ws['H' + str(per_unit_area_start_row_number + 1)].alignment = c_c_alignment
    ws['H' + str(per_unit_area_start_row_number + 1)] = _('Sample Variance')
    ws['H' + str(per_unit_area_start_row_number + 1)].border = f_border

    # table_data

    for i, value in enumerate(category):
        row_data = per_unit_area_start_row_number + 2 + i
        ws['B' + str(row_data)].font = name_font
        ws['B' + str(row_data)].alignment = c_c_alignment
        ws['B' + str(row_data)] = reporting_period_data['names'][i] + " (" + reporting_period_data['units'][
            i] + "/M²)"
        ws['B' + str(row_data)].border = f_border

        ws['C' + str(row_data)].font = name_font
        ws['C' + str(row_data)].alignment = c_c_alignment
        if reporting_period_data['means_per_unit_area'][i] \
                or reporting_period_data['means_per_unit_area'][i] == 0:
            ws['C' + str(row_data)] = round(reporting_period_data['means_per_unit_area'][i], 2)
        ws['C' + str(row_data)].border = f_border
        ws['C' + str(row_data)].number_format = '0.00'

        ws['D' + str(row_data)].font = name_font
        ws['D' + str(row_data)].alignment = c_c_alignment
        if reporting_period_data['medians_per_unit_area'][i] \
                or reporting_period_data['medians_per_unit_area'][i] == 0:
            ws['D' + str(row_data)] = round(reporting_period_data['medians_per_unit_area'][i], 2)
        ws['D' + str(row_data)].border = f_border
        ws['D' + str(row_data)].number_format = '0.00'

        ws['E' + str(row_data)].font = name_font
        ws['E' + str(row_data)].alignment = c_c_alignment
        if reporting_period_data['minimums_per_unit_area'][i] \
                or reporting_period_data['minimums_per_unit_area'][i] == 0:
            ws['E' + str(row_data)] = round(reporting_period_data['minimums_per_unit_area'][i], 2)
        ws['E' + str(row_data)].border = f_border
        ws['E' + str(row_data)].number_format = '0.00'

        ws['F' + str(row_data)].font = name_font
        ws['F' + str(row_data)].alignment = c_c_alignment
        if reporting_period_data['maximums_per_unit_area'][i] \
                or reporting_period_data['maximums_per_unit_area'][i] == 0:
            ws['F' + str(row_data)] = round(reporting_period_data['maximums_per_unit_area'][i], 2)
        ws['F' + str(row_data)].border = f_border
        ws['F' + str(row_data)].number_format = '0.00'

        ws['G' + str(row_data)].font = name_font
        ws['G' + str(row_data)].alignment = c_c_alignment
        if (reporting_period_data['stdevs_per_unit_area'][i]) \
                or reporting_period_data['stdevs_per_unit_area'][i] == 0:
            ws['G' + str(row_data)] = round(reporting_period_data['stdevs_per_unit_area'][i], 2)
        ws['G' + str(row_data)].border = f_border
        ws['G' + str(row_data)].number_format = '0.00'

        ws['H' + str(row_data)].font = name_font
        ws['H' + str(row_data)].alignment = c_c_alignment
        if reporting_period_data['variances_per_unit_area'][i] \
                or reporting_period_data['variances_per_unit_area'][i] == 0:
            ws['H' + str(row_data)] = round(reporting_period_data['variances_per_unit_area'][i], 2)
        ws['H' + str(row_data)].border = f_border
        ws['H' + str(row_data)].number_format = '0.00'

    ####################################################################################################################
    # Third: Detailed Data
    # detailed_start_row_number~ detailed_start_row_number+time_len: line
    # detailed_start_row_number+1: table title
    # i + analysis_end_row_number + 2 + 6 * ca_len~: table_data
    ####################################################################################################################
    has_timestamps_flag = True
    if "timestamps" not in reporting_period_data.keys() or \
            reporting_period_data['timestamps'] is None or \
            len(reporting_period_data['timestamps']) == 0:
        has_timestamps_flag = False

    timestamps = reporting_period_data['timestamps'][0]
    names = reporting_period_data['names']
    ca_len = len(names)
    time_len = len(timestamps)
    parameters_names_len = len(report['parameters']['names'])
    parameters_parameters_datas_len = 0
    analysis_end_row_number = 12 + 3 * ca_len
    current_row_number = analysis_end_row_number
    values = reporting_period_data['values']
    if has_timestamps_flag:
        for i in range(0, parameters_names_len):
            if len(report['parameters']['timestamps'][i]) == 0:
                continue
            parameters_parameters_datas_len += 1
        detail_data_table_start_row_number = current_row_number + (ca_len + parameters_parameters_datas_len) * 6 + 2

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)] = name + ' ' + _('Detailed Data')
        # table_title
        ws['B' + str(detail_data_table_start_row_number)].fill = table_fill
        ws['B' + str(detail_data_table_start_row_number)].font = name_font
        ws['B' + str(detail_data_table_start_row_number)].alignment = c_c_alignment
        ws['B' + str(detail_data_table_start_row_number)] = _('Datetime')
        ws['B' + str(detail_data_table_start_row_number)].border = f_border

        current_row_number += 1

        for i in range(0, ca_len):
            col = chr(ord('C') + i)

            ws[col + str(detail_data_table_start_row_number)].font = name_font
            ws[col + str(detail_data_table_start_row_number)].alignment = c_c_alignment
            ws[col + str(detail_data_table_start_row_number)] = \
                names[i] + " - (" + reporting_period_data['units'][i] + ")"
            ws[col + str(detail_data_table_start_row_number)].border = f_border
        # table_date
        for i in range(0, time_len):
            rows = i + detail_data_table_start_row_number + 1

            ws['B' + str(rows)].font = name_font
            ws['B' + str(rows)].alignment = c_c_alignment
            ws['B' + str(rows)] = timestamps[i]
            ws['B' + str(rows)].border = f_border

            for index in range(0, ca_len):
                col = chr(ord('C') + index)

                ws[col + str(rows)].font = name_font
                ws[col + str(rows)].alignment = c_c_alignment
                ws[col + str(rows)] = round(values[index][i], 2)
                ws[col + str(rows)].number_format = '0.00'
                ws[col + str(rows)].border = f_border

        # Subtotal
        row_subtotals = detail_data_table_start_row_number + 1 + time_len
        ws['B' + str(row_subtotals)].font = name_font
        ws['B' + str(row_subtotals)].alignment = c_c_alignment
        ws['B' + str(row_subtotals)] = _('Subtotal')
        ws['B' + str(row_subtotals)].border = f_border

        for i in range(0, ca_len):
            col = chr(ord('C') + i)

            ws[col + str(row_subtotals)].font = name_font
            ws[col + str(row_subtotals)].alignment = c_c_alignment
            ws[col + str(row_subtotals)] = round(reporting_period_data['subtotals'][i], 2)
            ws[col + str(row_subtotals)].border = f_border
            ws[col + str(row_subtotals)].number_format = '0.00'

    ####################################################################################################################
    # third: LineChart
    # LineChart requires data from the detailed data table in the Excel file
    # so print the detailed data table first and then print LineChart
    ####################################################################################################################
        for i in range(0, ca_len):
            line = LineChart()
            line.title = _("Reporting Period Consumption") + " - " + names[i] + \
                "(" + reporting_period_data['units'][i] + ")"
            line.style = 10
            line.x_axis.majorTickMark = 'in'
            line.y_axis.majorTickMark = 'in'
            line.smooth = True
            line.x_axis.crosses = 'min'
            line.height = 8.25
            line.width = 24
            line.dLbls = DataLabelList()
            line.dLbls.dLblPos = 't'
            line.dLbls.showVal = True
            times = Reference(ws, min_col=2, min_row=detail_data_table_start_row_number + 1,
                              max_row=detail_data_table_start_row_number + 1 + time_len)
            line_data = Reference(ws, min_col=3 + i, min_row=detail_data_table_start_row_number,
                                  max_row=detail_data_table_start_row_number + time_len)
            line.add_data(line_data, titles_from_data=True)
            line.set_categories(times)
            ser = line.series[0]
            ser.marker.symbol = "diamond"
            ser.marker.size = 5
            ws.add_chart(line, 'B' + str(current_row_number + 6 * i))

    ####################################################################################################################
    ca_len = len(report['reporting_period']['names'])
    current_sheet_parameters_row_number = current_row_number + ca_len * 6
    if 'parameters' not in report.keys() or \
            report['parameters'] is None or \
            'names' not in report['parameters'].keys() or \
            report['parameters']['names'] is None or \
            len(report['parameters']['names']) == 0 or \
            'timestamps' not in report['parameters'].keys() or \
            report['parameters']['timestamps'] is None or \
            len(report['parameters']['timestamps']) == 0 or \
            'values' not in report['parameters'].keys() or \
            report['parameters']['values'] is None or \
            len(report['parameters']['values']) == 0 or \
            timestamps_data_all_equal_0(report['parameters']['timestamps']):
        pass
    else:

        ################################################################################################################
        # new worksheet
        ################################################################################################################

        parameters_data = report['parameters']

        parameters_names_len = len(parameters_data['names'])

        file_name = ws.title
        parameters_ws = wb.create_sheet(file_name + _('Parameters'))

        parameters_timestamps_data_max_len = \
            get_parameters_timestamps_lists_max_len(list(parameters_data['timestamps']))

        # Row height
        parameters_ws.row_dimensions[1].height = 102
        for i in range(2, 7 + 1):
            parameters_ws.row_dimensions[i].height = 42

        for i in range(8, parameters_timestamps_data_max_len + 10):
            parameters_ws.row_dimensions[i].height = 60

        # Col width
        parameters_ws.column_dimensions['A'].width = 1.5

        parameters_ws.column_dimensions['B'].width = 25.0

        for i in range(3, 12 + parameters_names_len * 3):
            parameters_ws.column_dimensions[format_cell.get_column_letter(i)].width = 15.0

        # Img
        img = Image("excelexporters/myems.png")
        parameters_ws.add_image(img, 'A1')

        # Title
        parameters_ws['B3'].alignment = b_r_alignment
        parameters_ws['B3'] = _('Name') + ':'
        parameters_ws['C3'].border = b_border
        parameters_ws['C3'].alignment = b_c_alignment
        parameters_ws['C3'] = name

        parameters_ws['D3'].alignment = b_r_alignment
        parameters_ws['D3'] = _('Period Type') + ':'
        parameters_ws['E3'].border = b_border
        parameters_ws['E3'].alignment = b_c_alignment
        parameters_ws['E3'] = period_type

        parameters_ws['B4'].alignment = b_r_alignment
        parameters_ws['B4'] = _('Reporting Start Datetime') + ':'
        parameters_ws['C4'].border = b_border
        parameters_ws['C4'].alignment = b_c_alignment
        parameters_ws['C4'] = reporting_start_datetime_local

        parameters_ws['D4'].alignment = b_r_alignment
        parameters_ws['D4'] = _('Reporting End Datetime') + ':'
        parameters_ws['E4'].border = b_border
        parameters_ws['E4'].alignment = b_c_alignment
        parameters_ws['E4'] = reporting_end_datetime_local

        parameters_ws_current_row_number = 6

        parameters_ws['B' + str(parameters_ws_current_row_number)].font = title_font
        parameters_ws['B' + str(parameters_ws_current_row_number)] = name + ' ' + _('Parameters')

        parameters_ws_current_row_number += 1

        parameters_table_start_row_number = parameters_ws_current_row_number

        parameters_ws.row_dimensions[parameters_ws_current_row_number].height = 80

        parameters_ws_current_row_number += 1

        table_current_col_number = 2

        for i in range(0, parameters_names_len):

            if len(parameters_data['timestamps'][i]) == 0:
                continue

            col = format_cell.get_column_letter(table_current_col_number)

            parameters_ws[col + str(parameters_ws_current_row_number - 1)].fill = table_fill
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].border = f_border

            col = format_cell.get_column_letter(table_current_col_number + 1)

            parameters_ws[col + str(parameters_ws_current_row_number - 1)].fill = table_fill
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].border = f_border
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].font = name_font
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].alignment = c_c_alignment
            parameters_ws[col + str(parameters_ws_current_row_number - 1)] = parameters_data['names'][i]

            table_current_row_number = parameters_ws_current_row_number

            for j, value in enumerate(list(parameters_data['timestamps'][i])):
                col = format_cell.get_column_letter(table_current_col_number)

                parameters_ws[col + str(table_current_row_number)].border = f_border
                parameters_ws[col + str(table_current_row_number)].font = title_font
                parameters_ws[col + str(table_current_row_number)].alignment = c_c_alignment
                parameters_ws[col + str(table_current_row_number)] = value

                col = format_cell.get_column_letter(table_current_col_number + 1)

                parameters_ws[col + str(table_current_row_number)].border = f_border
                parameters_ws[col + str(table_current_row_number)].font = title_font
                parameters_ws[col + str(table_current_row_number)].alignment = c_c_alignment
                parameters_ws[col + str(table_current_row_number)] = round(parameters_data['values'][i][j], 2)

                table_current_row_number += 1

            table_current_col_number = table_current_col_number + 3

        ################################################################################################################
        # parameters chart and parameters table
        ################################################################################################################

        ws['B' + str(current_sheet_parameters_row_number)].font = title_font
        ws['B' + str(current_sheet_parameters_row_number)] = name + ' ' + _('Parameters')

        current_sheet_parameters_row_number += 1

        chart_start_row_number = current_sheet_parameters_row_number

        col_index = 0

        for i in range(0, parameters_names_len):

            if len(parameters_data['timestamps'][i]) == 0:
                continue

            line = LineChart()
            data_col = 3 + col_index * 3
            labels_col = 2 + col_index * 3
            col_index += 1
            line.title = _('Parameters') + ' - ' + \
                         parameters_ws.cell(row=parameters_table_start_row_number, column=data_col).value
            labels = Reference(parameters_ws, min_col=labels_col, min_row=parameters_table_start_row_number + 1,
                               max_row=(len(parameters_data['timestamps'][i]) + parameters_table_start_row_number))
            line_data = Reference(parameters_ws, min_col=data_col, min_row=parameters_table_start_row_number,
                                  max_row=(len(parameters_data['timestamps'][i]) + parameters_table_start_row_number))
            line.add_data(line_data, titles_from_data=True)
            line.set_categories(labels)
            line_data = line.series[0]
            line_data.marker.symbol = "circle"
            line_data.smooth = True
            line.x_axis.crosses = 'min'
            line.height = 8.25
            line.width = 24
            line.dLbls = DataLabelList()
            line.dLbls.dLblPos = 't'
            line.dLbls.showVal = False
            line.dLbls.showPercent = False
            chart_col = 'B'
            chart_cell = chart_col + str(chart_start_row_number)
            chart_start_row_number += 6
            ws.add_chart(line, chart_cell)

        current_sheet_parameters_row_number = chart_start_row_number

        current_sheet_parameters_row_number += 1

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename


def get_parameters_timestamps_lists_max_len(parameters_timestamps_lists):
    max_len = 0
    for i, value in enumerate(list(parameters_timestamps_lists)):
        if len(value) > max_len:
            max_len = len(value)

    return max_len


def timestamps_data_all_equal_0(lists):
    for i, value in enumerate(list(lists)):
        if len(value) > 0:
            return False

    return True
