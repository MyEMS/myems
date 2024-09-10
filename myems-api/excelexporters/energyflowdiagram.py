import base64
from core.utilities import get_translation
import os
import uuid
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from core.utilities import round2
from decimal import Decimal
import plotly.graph_objects as go


####################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excelexporters file
# Step 3: Encode the excelexporters file to Base64
####################################################################################################################
def export(result, name, reporting_start_datetime_local, reporting_end_datetime_local, language):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if result is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(result,
                              name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              language)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        print(str(ex))

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        print(str(ex))
    return base64_message


def generate_excel(report, name, reporting_start_datetime_local, reporting_end_datetime_local, language):

    trans = get_translation(language)
    trans.install()
    _ = trans.gettext

    wb = Workbook()
    ws = wb.active
    ws.title = "EnergyFlowDiagram"

    # Energy Flow Diagram Data Structure
    nodes = list()
    target = list()
    values = list()
    node_value = list()
    source = list()
    labels = list()
    color = ['rgba(25,202,173, 0.8)', 'rgba(140,199,181, 0.8)', 'rgba(160,238,225, 0.8)', 'rgba(190,231,233, 0.8)',
             'rgba(190,237,199,0.8)', 'rgba(109,91,7,0.8)', 'rgba(27,167,4,0.8)', 'rgba(108,234,12,0.8)',
             'rgba(202,217,14,0.8)', 'rgba(70,6,12,0.8)', 'rgba(246,10,15,0.8)', 'rgba(240,5,15,0.8)']
    node_color_list = list()
    for key, node in enumerate(report['nodes']):
        nodes.append(node['name'])
        node_value.append(0)
        node_color_list.append(color[key % len(color)])
    for link in report['links']:
        source.append(nodes.index(link['source']))
        target.append(nodes.index(link['target']))
        link_value = Decimal(link['value']) if link['value'] is not None else Decimal(0.0)
        values.append(link_value)
        node_value[nodes.index(link['source'])] = node_value[nodes.index(link['source'])] - link_value
        node_value[nodes.index(link['target'])] = node_value[nodes.index(link['target'])] + link_value
    for key, value in enumerate(nodes):
        labels.append(value + ':' + str(node_value[key].copy_abs()))

    # Row height
    ws.row_dimensions[1].height = 102
    for i in range(2, 5 + 1):
        ws.row_dimensions[i].height = 42

    for i in range(6, len(report['links']) + 15):
        ws.row_dimensions[i].height = 60

    # Col width
    ws.column_dimensions['A'].width = 1.5

    ws.column_dimensions['B'].width = 25.0

    for i in range(ord('C'), ord('L')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Arial', size=15, bold=True)
    title_font = Font(name='Arial', size=15, bold=True)

    table_fill = PatternFill(fill_type='solid', fgColor='90ee90')
    f_border = Border(left=Side(border_style='medium'),
                      right=Side(border_style='medium'),
                      bottom=Side(border_style='medium'),
                      top=Side(border_style='medium')
                      )
    b_border = Border(
        bottom=Side(border_style='medium'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)

    # Img
    img = Image("excelexporters/myems.png")
    ws.add_image(img, 'A1')

    # Title
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = _('Name') + ':'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'] = name

    ws['B4'].alignment = b_r_alignment
    ws['B4'] = _('Reporting Start Datetime') + ':'
    ws['C4'].border = b_border
    ws['C4'].alignment = b_c_alignment
    ws['C4'] = reporting_start_datetime_local

    ws['B5'].alignment = b_r_alignment
    ws['B5'] = _('Reporting End Datetime') + ':'
    ws['C5'].border = b_border
    ws['C5'].alignment = b_c_alignment
    ws['C5'] = reporting_end_datetime_local

    # Title
    ws['B6'].border = f_border
    ws['B6'].font = name_font
    ws['B6'].alignment = c_c_alignment
    ws['B6'].fill = table_fill
    ws['B6'] = _('Name')

    ws['C6'].border = f_border
    ws['C6'].alignment = c_c_alignment
    ws['C6'].font = name_font
    ws['C6'].fill = table_fill
    ws['C6'] = _('Name')

    ws['D6'].border = f_border
    ws['D6'].alignment = c_c_alignment
    ws['D6'].font = name_font
    ws['D6'].fill = table_fill
    ws['D6'] = _('Value')

    current_row_number = 7
    for i in range(0, len(report['links'])):

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)].border = f_border
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)] = report['links'][i]['source']

        ws['C' + str(current_row_number)].font = title_font
        ws['C' + str(current_row_number)].border = f_border
        ws['C' + str(current_row_number)].alignment = c_c_alignment
        ws['C' + str(current_row_number)] = report['links'][i]['target']

        ws['D' + str(current_row_number)].font = title_font
        ws['D' + str(current_row_number)].border = f_border
        ws['D' + str(current_row_number)].alignment = c_c_alignment
        ws['D' + str(current_row_number)] = round2(report['links'][i]['value'],2)

        current_row_number += 1

    fig = go.Figure(data=[go.Sankey(
        valueformat=".0f",
        valuesuffix="TWh",
        # Define nodes
        node=dict(pad=15, thickness=15, line=dict(color="black", width=0.5), label=labels, color=node_color_list),
        # Add links
        link=dict(source=source, target=target, value=values)
    )])

    fig.update_layout(title_text=name, font_size=10)
    
    # Save image file
    fig.write_image("sankey.png")  

    # Insert image
    img = Image("sankey.png")
    ws.add_image(img, 'B' + str(current_row_number))

    current_row_number += 1

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    # Delete image file
    os.remove("sankey.png")
    return filename
