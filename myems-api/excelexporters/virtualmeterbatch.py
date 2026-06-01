"""
Virtual Meter Batch Excel Exporter

This module provides functionality to export virtual meter batch data to Excel format.
It generates comprehensive reports showing energy consumption data for multiple virtual meters
within a specific time period.

Key Features:
- Multi-virtual meter energy consumption comparison
- Energy category breakdown with units
- Formatted Excel output with proper styling
- Multi-language support
- Base64 encoding for file transmission

The exported Excel file includes:
- Virtual meter names and associated spaces
- Energy consumption by category
- Proper formatting and borders
- Logo and header information
"""

import base64
from core.utilities import get_translation
import os
import uuid
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font


########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file from the report data
# Step 3: Encode the excel file to Base64
########################################################################################################################

def export(result, space_name, reporting_start_datetime_local, reporting_end_datetime_local, language):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if result is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(result,
                              space_name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              language)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        print(str(ex))

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        print(str(ex))
    return base64_message


def generate_excel(report, space_name, reporting_start_datetime_local, reporting_end_datetime_local, language):

    trans = get_translation(language)
    trans.install()
    _ = trans.gettext

    wb = Workbook()
    ws = wb.active
    ws.title = "VirtualMeterBatch"

    date_list_len = len(report.get('date_list', []))
    total_cols = 4 + date_list_len + 1  # ID, Name, Space, Energy Category + dates + total

    for i in range(ord('A'), ord('A') + min(total_cols, 26)):
        ws.column_dimensions[chr(i)].width = 15.0

    if date_list_len > 0:
        for i in range(4, 4 + date_list_len):
            col_letter = chr(ord('A') + i) if i < 26 else chr(ord('A') + (i // 26) - 1) + chr(ord('A') + (i % 26))
            ws.column_dimensions[col_letter].width = 12.0

    # Head image
    ws.row_dimensions[1].height = 105
    img = Image("excelexporters/myems.png")
    ws.add_image(img, 'A1')

    # Query Parameters
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    ws['A3'].alignment = b_r_alignment
    ws['A3'] = _('Space') + ':'
    ws['B3'] = space_name
    ws['A4'].alignment = b_r_alignment
    ws['A4'] = _('Start Datetime') + ':'
    ws['B4'] = reporting_start_datetime_local
    ws['A5'].alignment = b_r_alignment
    ws['A5'] = _('End Datetime') + ':'
    ws['B5'] = reporting_end_datetime_local

    # Title
    title_font = Font(size=12, bold=True)
    ws['A6'].font = title_font
    ws['A6'] = _('ID')
    ws['B6'].font = title_font
    ws['B6'] = _('Name')
    ws['C6'].font = title_font
    ws['C6'] = _('Space')
    ws['D6'].font = title_font
    ws['D6'] = _('Energy Category')

    date_list = report.get('date_list', [])
    for i in range(0, len(date_list)):
        col_idx = 4 + i
        if col_idx < 26:
            col = chr(ord('A') + col_idx)
        else:
            col = chr(ord('A') + (col_idx // 26) - 1) + chr(ord('A') + (col_idx % 26))
        ws[col + '6'].font = title_font
        ws[col + '6'] = date_list[i]

    total_col_idx = 4 + len(date_list)
    if total_col_idx < 26:
        total_col = chr(ord('A') + total_col_idx)
    else:
        total_col = chr(ord('A') + (total_col_idx // 26) - 1) + chr(ord('A') + (total_col_idx % 26))
    ws[total_col + '6'].font = title_font
    ws[total_col + '6'] = _('Total')

    current_row_number = 7
    for i in range(0, len(report['virtual_meters'])):
        ws['A' + str(current_row_number)] = str(report['virtual_meters'][i]['id'])
        ws['B' + str(current_row_number)] = report['virtual_meters'][i]['virtual_meter_name']
        ws['C' + str(current_row_number)] = report['virtual_meters'][i]['space_name']
        ws['D' + str(current_row_number)] = report['virtual_meters'][i].get('energy_category_name', '')

        daily_values = report['virtual_meters'][i].get('daily_values', [])
        for j in range(0, len(daily_values)):
            col_idx = 4 + j
            if col_idx < 26:
                col = chr(ord('A') + col_idx)
            else:
                col = chr(ord('A') + (col_idx // 26) - 1) + chr(ord('A') + (col_idx % 26))
            ws[col + str(current_row_number)] = daily_values[j].get('value', None)
        ws[total_col + str(current_row_number)] = report['virtual_meters'][i].get('subtotal', None)

        current_row_number += 1

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename
