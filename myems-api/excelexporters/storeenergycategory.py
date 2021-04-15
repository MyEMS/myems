import base64
import uuid
import os
from openpyxl.chart import (
        PieChart,
        LineChart,
        Reference,
    )
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl.chart.label import DataLabelList
import openpyxl.utils.cell as format_cell

####################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file
# Step 3: Encode the excel file bytes to Base64
####################################################################################################################


def export(report,
           name,
           reporting_start_datetime_local,
           reporting_end_datetime_local,
           period_type):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if report is None:
        return None
    print(report)

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(report,
                              name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              period_type)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        pass

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        pass
    return base64_message


def decimal_to_column(num=65):
    string = ''
    num = num - 64
    # The column number is not greater than 90
    if num <= 26:
        return chr(num + 64)
    # The column number is greater than 90
    while num // 26 > 0:
        if num % 26 == 0:
            string += 'Z'
            num = num // 26 - 1
        else:
            string += chr(num % 26 + 64)
            num //= 26
    # Avoid conversion errors that might occur between 741 and 766
    if num > 0:
        string += chr(num + 64)

    return string[::-1]


def column_to_decimal(string='A'):
    num = 0
    for index, key in enumerate(string[::-1]):
        num += (ord(key) - 64) * (26 ** index)

    return num + 64


def get_parameters_timestamps_lists_max_len(parameters_timestamps_lists):
    max_len = 0
    for i, value in enumerate(list(parameters_timestamps_lists)):
        if len(value) > max_len:
            max_len = len(value)

    return max_len


def timestamps_data_all_equal_0(lists):
    for i, value in enumerate(list(lists)):
        if len(value) > 0:
            return False

    return True


def generate_excel(report,
                   name,
                   reporting_start_datetime_local,
                   reporting_end_datetime_local,
                   period_type):

    wb = Workbook()
    ws = wb.active

    # Row height
    ws.row_dimensions[1].height = 102

    for i in range(2, 2000 + 1):
        ws.row_dimensions[i].height = 42

    # Col width
    ws.column_dimensions['A'].width = 1.5

    ws.column_dimensions['B'].width = 25.0

    for i in range(ord('C'), ord('L')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Constantia', size=15, bold=True)
    title_font = Font(name='宋体', size=15, bold=True)
    data_font = Font(name='Franklin Gothic Book', size=11)

    table_fill = PatternFill(fill_type='solid', fgColor='1F497D')
    f_border = Border(left=Side(border_style='medium', color='00000000'),
                      right=Side(border_style='medium', color='00000000'),
                      bottom=Side(border_style='medium', color='00000000'),
                      top=Side(border_style='medium', color='00000000')
                      )
    b_border = Border(
        bottom=Side(border_style='medium', color='00000000'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_r_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)

    # Img
    img = Image("excelexporters/myems.png")
    # img = Image("myems.png")
    img.width = img.width * 1.06
    img.height = img.height * 1.06
    ws.add_image(img, 'B1')

    # Title
    ws.row_dimensions[3].height = 60

    ws['B3'].font = name_font
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = 'Name:'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'].font = name_font
    ws['C3'] = name

    ws['D3'].font = name_font
    ws['D3'].alignment = b_r_alignment
    ws['D3'] = 'Period:'
    ws['E3'].border = b_border
    ws['E3'].alignment = b_c_alignment
    ws['E3'].font = name_font
    ws['E3'] = period_type

    ws['F3'].font = name_font
    ws['F3'].alignment = b_r_alignment
    ws['F3'] = 'Date:'
    ws['G3'].border = b_border
    ws['G3'].alignment = b_c_alignment
    ws['G3'].font = name_font
    ws['G3'] = reporting_start_datetime_local[:10] + "__" + reporting_end_datetime_local[:10]
    ws.merge_cells("G3:H3")

    if "reporting_period" not in report.keys() or \
            "names" not in report['reporting_period'].keys() or len(report['reporting_period']['names']) == 0:
        filename = str(uuid.uuid4()) + '.xlsx'
        wb.save(filename)

        return filename
    #################################################
    # First: 能耗分析
    # 6: title
    # 7: table title
    # 8~10 table_data
    # Total: 5 rows
    # if has not energy data: set low height for rows
    #################################################
    reporting_period_data = report['reporting_period']

    has_energy_data_flag = True
    if "names" not in reporting_period_data.keys() or \
            reporting_period_data['names'] is None or \
            len(reporting_period_data['names']) == 0:
        has_energy_data_flag = False

    if has_energy_data_flag:
        ws['B6'].font = title_font
        ws['B6'] = name+' 能耗分析'

        category = reporting_period_data['names']
        ca_len = len(category)

        ws.row_dimensions[7].height = 60
        ws['B7'].fill = table_fill
        ws['B7'].border = f_border

        ws['B8'].font = title_font
        ws['B8'].alignment = c_c_alignment
        ws['B8'] = '能耗'
        ws['B8'].border = f_border

        ws['B9'].font = title_font
        ws['B9'].alignment = c_c_alignment
        ws['B9'] = '单位面积能耗'
        ws['B9'].border = f_border

        ws['B10'].font = title_font
        ws['B10'].alignment = c_c_alignment
        ws['B10'] = '环比'
        ws['B10'].border = f_border

        col = ''

        for i in range(0, ca_len):
            col = chr(ord('C') + i)
            row = '7'
            cell = col + row
            ws[col + '7'].fill = table_fill
            ws[col + '7'].font = name_font
            ws[col + '7'].alignment = c_c_alignment
            ws[col + '7'] = reporting_period_data['names'][i] + " (" + reporting_period_data['units'][i] + ")"
            ws[col + '7'].border = f_border

            ws[col + '8'].font = name_font
            ws[col + '8'].alignment = c_c_alignment
            ws[col + '8'] = round(reporting_period_data['subtotals'][i], 2)
            ws[col + '8'].border = f_border

            ws[col + '9'].font = name_font
            ws[col + '9'].alignment = c_c_alignment
            ws[col + '9'] = round(reporting_period_data['subtotals_per_unit_area'][i], 2)
            ws[col + '9'].border = f_border

            ws[col + '10'].font = name_font
            ws[col + '10'].alignment = c_c_alignment
            ws[col + '10'] = str(round(reporting_period_data['increment_rates'][i] * 100, 2)) + "%" \
                if reporting_period_data['increment_rates'][i] is not None else "-"
            ws[col + '10'].border = f_border

        # TCE TCO2E
        end_col = col
        # TCE
        tce_col = chr(ord(end_col) + 1)
        ws[tce_col + '7'].fill = table_fill
        ws[tce_col + '7'].font = name_font
        ws[tce_col + '7'].alignment = c_c_alignment
        ws[tce_col + '7'] = "吨标准煤 (TCE)"
        ws[tce_col + '7'].border = f_border

        ws[tce_col + '8'].font = name_font
        ws[tce_col + '8'].alignment = c_c_alignment
        ws[tce_col + '8'] = round(reporting_period_data['total_in_kgce'] / 1000, 2)
        ws[tce_col + '8'].border = f_border

        ws[tce_col + '9'].font = name_font
        ws[tce_col + '9'].alignment = c_c_alignment
        ws[tce_col + '9'] = round(reporting_period_data['total_in_kgce_per_unit_area'] / 1000, 2)
        ws[tce_col + '9'].border = f_border

        ws[tce_col + '10'].font = name_font
        ws[tce_col + '10'].alignment = c_c_alignment
        ws[tce_col + '10'] = str(round(reporting_period_data['increment_rate_in_kgce'] * 100, 2)) + "%" \
            if reporting_period_data['increment_rate_in_kgce'] is not None else "-"
        ws[tce_col + '10'].border = f_border

        # TCO2E
        tco2e_col = chr(ord(end_col) + 2)
        ws[tco2e_col + '7'].fill = table_fill
        ws[tco2e_col + '7'].font = name_font
        ws[tco2e_col + '7'].alignment = c_c_alignment
        ws[tco2e_col + '7'] = "吨二氧化碳排放 (TCO2E)"
        ws[tco2e_col + '7'].border = f_border

        ws[tco2e_col + '8'].font = name_font
        ws[tco2e_col + '8'].alignment = c_c_alignment
        ws[tco2e_col + '8'] = round(reporting_period_data['total_in_kgco2e'] / 1000, 2)
        ws[tco2e_col + '8'].border = f_border

        ws[tco2e_col + '9'].font = name_font
        ws[tco2e_col + '9'].alignment = c_c_alignment
        ws[tco2e_col + '9'] = round(reporting_period_data['total_in_kgco2e_per_unit_area'] / 1000, 2)
        ws[tco2e_col + '9'].border = f_border

        ws[tco2e_col + '10'].font = name_font
        ws[tco2e_col + '10'].alignment = c_c_alignment
        ws[tco2e_col + '10'] = str(round(reporting_period_data['increment_rate_in_kgco2e'] * 100, 2)) + "%" \
            if reporting_period_data['increment_rate_in_kgco2e'] is not None else "-"
        ws[tco2e_col + '10'].border = f_border
    else:
        for i in range(6, 10 + 1):
            ws.row_dimensions[i].height = 0.1
    #################################################
    # Second: 分时电耗
    # 12: title
    # 13: table title
    # 14~17 table_data
    # Total: 6 rows
    ################################################
    has_ele_peak_flag = True
    if "toppeaks" not in reporting_period_data.keys() or \
            reporting_period_data['toppeaks'] is None or \
            len(reporting_period_data['toppeaks']) == 0:
        has_ele_peak_flag = False

    if has_ele_peak_flag:
        ws['B12'].font = title_font
        ws['B12'] = name+' 分时电耗'

        ws.row_dimensions[13].height = 60
        ws['B13'].fill = table_fill
        ws['B13'].font = name_font
        ws['B13'].alignment = c_c_alignment
        ws['B13'].border = f_border

        ws['C13'].fill = table_fill
        ws['C13'].font = name_font
        ws['C13'].alignment = c_c_alignment
        ws['C13'].border = f_border
        ws['C13'] = '分时电耗'

        ws['B14'].font = title_font
        ws['B14'].alignment = c_c_alignment
        ws['B14'] = '尖'
        ws['B14'].border = f_border

        ws['C14'].font = title_font
        ws['C14'].alignment = c_c_alignment
        ws['C14'].border = f_border
        ws['C14'] = round(reporting_period_data['toppeaks'][0], 2)

        ws['B15'].font = title_font
        ws['B15'].alignment = c_c_alignment
        ws['B15'] = '峰'
        ws['B15'].border = f_border

        ws['C15'].font = title_font
        ws['C15'].alignment = c_c_alignment
        ws['C15'].border = f_border
        ws['C15'] = round(reporting_period_data['onpeaks'][0], 2)

        ws['B16'].font = title_font
        ws['B16'].alignment = c_c_alignment
        ws['B16'] = '平'
        ws['B16'].border = f_border

        ws['C16'].font = title_font
        ws['C16'].alignment = c_c_alignment
        ws['C16'].border = f_border
        ws['C16'] = round(reporting_period_data['midpeaks'][0], 2)

        ws['B17'].font = title_font
        ws['B17'].alignment = c_c_alignment
        ws['B17'] = '谷'
        ws['B17'].border = f_border

        ws['C17'].font = title_font
        ws['C17'].alignment = c_c_alignment
        ws['C17'].border = f_border
        ws['C17'] = round(reporting_period_data['offpeaks'][0], 2)

        pie = PieChart()
        pie.title = name+' 分时电耗'
        labels = Reference(ws, min_col=2, min_row=14, max_row=17)
        pie_data = Reference(ws, min_col=3, min_row=13, max_row=17)
        pie.add_data(pie_data, titles_from_data=True)
        pie.set_categories(labels)
        pie.height = 7.25  # cm 1.05*5 1.05cm = 30 pt
        pie.width = 9
        # pie.title = "Pies sold by category"
        s1 = pie.series[0]
        s1.dLbls = DataLabelList()
        s1.dLbls.showCatName = False  # 标签显示
        s1.dLbls.showVal = True  # 数量显示
        s1.dLbls.showPercent = True  # 百分比显示
        # s1 = CharacterProperties(sz=1800)     # 图表中字体大小 *100

        ws.add_chart(pie, "D13")

    else:
        for i in range(12, 18 + 1):
            ws.row_dimensions[i].height = 0.1

    ################################################
    # Fourth: 能耗详情
    # current_row_number: title
    # current_row_number+1 ~ current_row_number+1+ca_len*6-1: line
    # current_row_number+1+ca_len*6: table title
    # current_row_number+1+ca_len*6~: table_data
    ################################################
    current_row_number = 19
    reporting_period_data = report['reporting_period']
    times = reporting_period_data['timestamps']
    has_detail_data_flag = True
    ca_len = len(report['reporting_period']['names'])
    parameters_names_len = len(report['parameters']['names'])
    parameters_parameters_datas_len = 0
    for i in range(0, parameters_names_len):
        if len(report['parameters']['timestamps'][i]) == 0:
            continue
        parameters_parameters_datas_len += 1
    table_row = current_row_number + (ca_len + parameters_parameters_datas_len) * 6 + 2
    chart_start_row_number = current_row_number + 1
    if "timestamps" not in reporting_period_data.keys() or \
            reporting_period_data['timestamps'] is None or \
            len(reporting_period_data['timestamps']) == 0:
        has_detail_data_flag = False

    if has_detail_data_flag:
        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)] = name+' 详细数据'

        ws.row_dimensions[table_row].height = 60
        ws['B'+str(table_row)].fill = table_fill
        ws['B' + str(table_row)].font = title_font
        ws['B'+str(table_row)].border = f_border
        ws['B'+str(table_row)].alignment = c_c_alignment
        ws['B'+str(table_row)] = '日期时间'
        time = times[0]
        has_data = False
        max_row = 0
        if len(time) > 0:
            has_data = True
            max_row = table_row + len(time)
            print("max_row", max_row)

        if has_data:
            for i in range(0, len(time)):
                col = 'B'
                row = str(table_row+1 + i)
                # col = chr(ord('B') + i)
                ws[col + row].font = title_font
                ws[col + row].alignment = c_c_alignment
                ws[col + row] = time[i]
                ws[col + row].border = f_border

            for i in range(0, ca_len):
                # 38 title
                col = chr(ord('C') + i)

                ws[col + str(table_row)].fill = table_fill
                ws[col + str(table_row)].font = title_font
                ws[col + str(table_row)].alignment = c_c_alignment
                ws[col + str(table_row)] = reporting_period_data['names'][i] + \
                    " (" + reporting_period_data['units'][i] + ")"
                ws[col + str(table_row)].border = f_border

                # 39 data
                time = times[i]
                time_len = len(time)

                for j in range(0, time_len):
                    row = str(table_row+1 + j)
                    # col = chr(ord('B') + i)
                    ws[col + row].font = title_font
                    ws[col + row].alignment = c_c_alignment
                    ws[col + row] = round(reporting_period_data['values'][i][j], 2)
                    ws[col + row].border = f_border

            current_row_number = table_row + 1 + len(times[0])

            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)].alignment = c_c_alignment
            ws['B' + str(current_row_number)].border = f_border
            ws['B' + str(current_row_number)] = '小计'

            for i in range(0, ca_len):
                col = chr(ord('C') + i)
                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)].border = f_border
                ws[col + str(current_row_number)] = round(reporting_period_data['subtotals'][i], 2)

                # line
                # 39~: line
                line = LineChart()
                line.title = '报告期消耗 - ' + ws.cell(column=3+i, row=table_row).value
                labels = Reference(ws, min_col=2, min_row=table_row+1, max_row=max_row)
                line_data = Reference(ws, min_col=3 + i, min_row=table_row, max_row=max_row)  # openpyxl bug
                line.add_data(line_data, titles_from_data=True)
                line.set_categories(labels)
                line_data = line.series[0]
                line_data.marker.symbol = "circle"
                line_data.smooth = True
                line.x_axis.crosses = 'min'
                line.height = 8.25  # cm 1.05*5 1.05cm = 30 pt
                line.width = 24
                # pie.title = "Pies sold by category"
                line.dLbls = DataLabelList()
                line.dLbls.dLblPos = 't'
                # line.dLbls.showCatName = True  # label show
                line.dLbls.showVal = True  # val show
                line.dLbls.showPercent = True  # percent show
                # s1 = CharacterProperties(sz=1800)     # font size *100
                chart_col = 'B'
                chart_cell = chart_col + str(chart_start_row_number + 6*i)
                ws.add_chart(line, chart_cell)

    ##########################################
    has_parameters_names_and_timestamps_and_values_data = True
    # 12 is the starting line number of the last line chart in the report period

    ca_len = len(report['reporting_period']['names'])
    current_sheet_parameters_row_number = chart_start_row_number + ca_len * 6
    if 'parameters' not in report.keys() or \
            report['parameters'] is None or \
            'names' not in report['parameters'].keys() or \
            report['parameters']['names'] is None or \
            len(report['parameters']['names']) == 0 or \
            'timestamps' not in report['parameters'].keys() or \
            report['parameters']['timestamps'] is None or \
            len(report['parameters']['timestamps']) == 0 or \
            'values' not in report['parameters'].keys() or \
            report['parameters']['values'] is None or \
            len(report['parameters']['values']) == 0 or \
            timestamps_data_all_equal_0(report['parameters']['timestamps']):
        has_parameters_names_and_timestamps_and_values_data = False
    if has_parameters_names_and_timestamps_and_values_data:

        ###############################
        # new worksheet
        ###############################

        parameters_data = report['parameters']

        parameters_names_len = len(parameters_data['names'])

        parameters_ws = wb.create_sheet('相关参数')

        parameters_timestamps_data_max_len = \
            get_parameters_timestamps_lists_max_len(list(parameters_data['timestamps']))

        # Row height
        parameters_ws.row_dimensions[1].height = 102
        for i in range(2, 7 + 1):
            parameters_ws.row_dimensions[i].height = 42

        for i in range(8, parameters_timestamps_data_max_len + 10):
            parameters_ws.row_dimensions[i].height = 60

        # Col width
        parameters_ws.column_dimensions['A'].width = 1.5

        parameters_ws.column_dimensions['B'].width = 25.0

        for i in range(3, 12 + parameters_names_len * 3):
            parameters_ws.column_dimensions[format_cell.get_column_letter(i)].width = 15.0

        # Img
        img = Image("excelexporters/myems.png")
        img.width = img.width * 0.85
        img.height = img.height * 0.85
        # img = Image("myems.png")
        parameters_ws.add_image(img, 'B1')

        # Title
        parameters_ws.row_dimensions[3].height = 60

        parameters_ws['B3'].font = name_font
        parameters_ws['B3'].alignment = b_r_alignment
        parameters_ws['B3'] = 'Name:'
        parameters_ws['C3'].border = b_border
        parameters_ws['C3'].alignment = b_c_alignment
        parameters_ws['C3'].font = name_font
        parameters_ws['C3'] = name

        parameters_ws['D3'].font = name_font
        parameters_ws['D3'].alignment = b_r_alignment
        parameters_ws['D3'] = 'Period:'
        parameters_ws['E3'].border = b_border
        parameters_ws['E3'].alignment = b_c_alignment
        parameters_ws['E3'].font = name_font
        parameters_ws['E3'] = period_type

        parameters_ws['F3'].font = name_font
        parameters_ws['F3'].alignment = b_r_alignment
        parameters_ws['F3'] = 'Date:'
        parameters_ws['G3'].border = b_border
        parameters_ws['G3'].alignment = b_c_alignment
        parameters_ws['G3'].font = name_font
        parameters_ws['G3'] = reporting_start_datetime_local + "__" + reporting_end_datetime_local
        parameters_ws.merge_cells("G3:H3")

        parameters_ws_current_row_number = 6

        parameters_ws['B' + str(parameters_ws_current_row_number)].font = title_font
        parameters_ws['B' + str(parameters_ws_current_row_number)] = name + ' 相关参数'

        parameters_ws_current_row_number += 1

        parameters_table_start_row_number = parameters_ws_current_row_number

        parameters_ws.row_dimensions[parameters_ws_current_row_number].height = 80

        parameters_ws_current_row_number += 1

        table_current_col_number = 'B'

        for i in range(0, parameters_names_len):

            if len(parameters_data['timestamps'][i]) == 0:
                continue

            parameters_ws[table_current_col_number + str(parameters_ws_current_row_number - 1)].fill = table_fill
            parameters_ws[table_current_col_number + str(parameters_ws_current_row_number - 1)].border = f_border

            col = decimal_to_column(column_to_decimal(table_current_col_number) + 1)

            parameters_ws[col + str(parameters_ws_current_row_number - 1)].fill = table_fill
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].border = f_border
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].font = name_font
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].alignment = c_c_alignment
            parameters_ws[col + str(parameters_ws_current_row_number - 1)] = parameters_data['names'][i]

            table_current_row_number = parameters_ws_current_row_number

            for j, value in enumerate(list(parameters_data['timestamps'][i])):
                col = table_current_col_number

                parameters_ws[col + str(table_current_row_number)].border = f_border
                parameters_ws[col + str(table_current_row_number)].font = title_font
                parameters_ws[col + str(table_current_row_number)].alignment = c_c_alignment
                parameters_ws[col + str(table_current_row_number)] = value

                col = decimal_to_column(column_to_decimal(col) + 1)

                parameters_ws[col + str(table_current_row_number)].border = f_border
                parameters_ws[col + str(table_current_row_number)].font = title_font
                parameters_ws[col + str(table_current_row_number)].alignment = c_c_alignment
                parameters_ws[col + str(table_current_row_number)] = round(parameters_data['values'][i][j], 2)

                table_current_row_number += 1

            table_current_col_number = decimal_to_column(column_to_decimal(table_current_col_number) + 3)

        ########################################################
        # parameters chart and parameters table
        ########################################################

        ws['B' + str(current_sheet_parameters_row_number)].font = title_font
        ws['B' + str(current_sheet_parameters_row_number)] = name + ' 相关参数'

        current_sheet_parameters_row_number += 1

        chart_start_row_number = current_sheet_parameters_row_number

        col_index = 0

        for i in range(0, parameters_names_len):

            if len(parameters_data['timestamps'][i]) == 0:
                continue

            line = LineChart()
            data_col = 3 + col_index * 3
            labels_col = 2 + col_index * 3
            col_index += 1
            line.title = '相关参数 - ' + \
                         parameters_ws.cell(row=parameters_table_start_row_number, column=data_col).value
            labels = Reference(parameters_ws, min_col=labels_col, min_row=parameters_table_start_row_number + 1,
                               max_row=(len(parameters_data['timestamps'][i]) + parameters_table_start_row_number))
            line_data = Reference(parameters_ws, min_col=data_col, min_row=parameters_table_start_row_number,
                                  max_row=(len(parameters_data['timestamps'][i]) + parameters_table_start_row_number))
            line.add_data(line_data, titles_from_data=True)
            line.set_categories(labels)
            line_data = line.series[0]
            line_data.marker.symbol = "circle"
            line_data.smooth = True
            line.x_axis.crosses = 'min'
            line.height = 8.25
            line.width = 24
            line.dLbls = DataLabelList()
            line.dLbls.dLblPos = 't'
            line.dLbls.showVal = False
            line.dLbls.showPercent = False
            chart_col = 'B'
            chart_cell = chart_col + str(chart_start_row_number)
            chart_start_row_number += 6
            ws.add_chart(line, chart_cell)

        current_sheet_parameters_row_number = chart_start_row_number

        current_sheet_parameters_row_number += 1

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename
