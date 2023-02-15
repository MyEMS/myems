import base64
import gettext
import os
import re
import uuid

import openpyxl.utils.cell as format_cell
from openpyxl import Workbook
from openpyxl.chart import PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font


########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file
# Step 3: Encode the excel file to Base64
########################################################################################################################


def export(report,
           name,
           base_period_start_datetime_local,
           base_period_end_datetime_local,
           reporting_start_datetime_local,
           reporting_end_datetime_local,
           period_type,
           language):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if report is None:
        return None
    print(report)

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(report,
                              name,
                              base_period_start_datetime_local,
                              base_period_end_datetime_local,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              period_type,
                              language)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        print(str(ex))

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        print(str(ex))
    return base64_message


def generate_excel(report,
                   name,
                   base_period_start_datetime_local,
                   base_period_end_datetime_local,
                   reporting_start_datetime_local,
                   reporting_end_datetime_local,
                   period_type,
                   language):

    locale_path = './i18n/'
    if language == 'zh_CN':
        trans = gettext.translation('myems', locale_path, languages=['zh_CN'])
    elif language == 'de':
        trans = gettext.translation('myems', locale_path, languages=['de'])
    elif language == 'en':
        trans = gettext.translation('myems', locale_path, languages=['en'])
    else:
        trans = gettext.translation('myems', locale_path, languages=['en'])
    trans.install()
    _ = trans.gettext

    wb = Workbook()
    ws = wb.active
    ws.title = "CombinedEquipmentEnergyItem"

    # Row height
    ws.row_dimensions[1].height = 102
    for i in range(2, 2000 + 1):
        ws.row_dimensions[i].height = 42

    # Col width
    ws.column_dimensions['A'].width = 1.5

    ws.column_dimensions['B'].width = 25.0

    for i in range(ord('C'), ord('Z')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Arial', size=15, bold=True)
    title_font = Font(name='Arial', size=15, bold=True)

    table_fill = PatternFill(fill_type='solid', fgColor='90ee90')
    f_border = Border(left=Side(border_style='medium'),
                      right=Side(border_style='medium'),
                      bottom=Side(border_style='medium'),
                      top=Side(border_style='medium')
                      )
    b_border = Border(
        bottom=Side(border_style='medium'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    # Img
    img = Image("excelexporters/myems.png")
    ws.add_image(img, 'A1')

    # Title
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = _('Name') + ':'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'] = name

    ws['D3'].alignment = b_r_alignment
    ws['D3'] = _('Period Type') + ':'
    ws['E3'].border = b_border
    ws['E3'].alignment = b_c_alignment
    ws['E3'] = period_type

    ws['B4'].alignment = b_r_alignment
    ws['B4'] = _('Reporting Start Datetime') + ':'
    ws['C4'].border = b_border
    ws['C4'].alignment = b_c_alignment
    ws['C4'] = reporting_start_datetime_local

    ws['D4'].alignment = b_r_alignment
    ws['D4'] = _('Reporting End Datetime') + ':'
    ws['E4'].border = b_border
    ws['E4'].alignment = b_c_alignment
    ws['E4'] = reporting_end_datetime_local

    is_base_period_timestamp_exists_flag = is_base_period_timestamp_exists(report['base_period'])

    if is_base_period_timestamp_exists_flag:
        ws['B5'].alignment = b_r_alignment
        ws['B5'] = _('Base Period Start Datetime') + ':'
        ws['C5'].border = b_border
        ws['C5'].alignment = b_c_alignment
        ws['C5'] = base_period_start_datetime_local

        ws['D5'].alignment = b_r_alignment
        ws['D5'] = _('Base Period End Datetime') + ':'
        ws['E5'].border = b_border
        ws['E5'].alignment = b_c_alignment
        ws['E5'] = base_period_end_datetime_local

    if "reporting_period" not in report.keys() or \
            "names" not in report['reporting_period'].keys() or len(report['reporting_period']['names']) == 0:
        filename = str(uuid.uuid4()) + '.xlsx'
        wb.save(filename)

        return filename

    ####################################################################################################################

    current_row_number = 7

    reporting_period_data = report['reporting_period']
    if "names" not in reporting_period_data.keys() or \
            reporting_period_data['names'] is None or \
            len(reporting_period_data['names']) == 0:
        pass
    else:
        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)] = name + ' ' + _('Reporting Period Consumption')

        current_row_number += 1

        category = reporting_period_data['names']
        ca_len = len(category)

        ws.row_dimensions[current_row_number].height = 60
        ws['B' + str(current_row_number)].fill = table_fill
        ws['B' + str(current_row_number)].border = f_border

        col = 'C'

        for i in range(0, ca_len):
            ws[col + str(current_row_number)].fill = table_fill
            ws[col + str(current_row_number)].font = name_font
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)].border = f_border
            ws[col + str(current_row_number)] = \
                reporting_period_data['names'][i] + " " + reporting_period_data['energy_category_names'][i] + \
                " (" + reporting_period_data['units'][i] + ")"

            col = chr(ord(col) + 1)

        current_row_number += 1

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)].border = f_border
        ws['B' + str(current_row_number)] = _('Per Unit Area')

        col = 'C'

        for i in range(0, ca_len):
            ws[col + str(current_row_number)].font = name_font
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)].border = f_border
            ws[col + str(current_row_number)] = round(reporting_period_data['subtotals'][i], 2)

            col = chr(ord(col) + 1)

        current_row_number += 1

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)].border = f_border
        ws['B' + str(current_row_number)] = _('Increment Rate')

        col = 'C'

        for i in range(0, ca_len):
            ws[col + str(current_row_number)].font = name_font
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)].border = f_border
            ws[col + str(current_row_number)] = str(
                round(reporting_period_data['increment_rates'][i] * 100, 2)) + '%' \
                if reporting_period_data['increment_rates'][i] is not None else '-'

            col = chr(ord(col) + 1)

        current_row_number += 2

        category_dict = group_by_category(reporting_period_data['energy_category_names'])

        for category_dict_name, category_dict_values in category_dict.items():

            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)] = \
                name + ' ' + category_dict_name + ' ' + \
                '(' + reporting_period_data['units'][category_dict_values[0]] + ') by Energy Item'

            current_row_number += 1
            table_start_row_number = current_row_number

            ws['B' + str(current_row_number)].fill = table_fill
            ws['B' + str(current_row_number)].border = f_border

            ws['C' + str(current_row_number)].font = name_font
            ws['C' + str(current_row_number)].fill = table_fill
            ws['C' + str(current_row_number)].alignment = c_c_alignment
            ws['C' + str(current_row_number)].border = f_border
            ws['C' + str(current_row_number)] = _('Per Unit Area')

            current_row_number += 1

            for i in category_dict_values:
                ws['B' + str(current_row_number)].font = title_font
                ws['B' + str(current_row_number)].alignment = c_c_alignment
                ws['B' + str(current_row_number)].border = f_border
                ws['B' + str(current_row_number)] = \
                    reporting_period_data['names'][i] + " (" + reporting_period_data['units'][i] + ")"
                ws['C' + str(current_row_number)].font = name_font
                ws['C' + str(current_row_number)].alignment = c_c_alignment
                ws['C' + str(current_row_number)].border = f_border
                ws['C' + str(current_row_number)] = round(reporting_period_data['subtotals'][i], 3)

                current_row_number += 1

            table_end_row_number = current_row_number - 1

            pie = PieChart()
            pie.title = \
                name + ' ' + category_dict_name + ' ' + \
                '(' + reporting_period_data['units'][category_dict_values[0]] + ') by Energy Item'
            labels = Reference(ws, min_col=2, min_row=table_start_row_number + 1, max_row=table_end_row_number)
            pie_data = Reference(ws, min_col=3, min_row=table_start_row_number, max_row=table_end_row_number)
            pie.add_data(pie_data, titles_from_data=True)
            pie.set_categories(labels)
            pie.height = 6.6
            pie.width = 9
            s1 = pie.series[0]
            s1.dLbls = DataLabelList()
            s1.dLbls.showCatName = False
            s1.dLbls.showVal = True
            s1.dLbls.showPercent = True
            ws.add_chart(pie, 'D' + str(table_start_row_number))

            if len(category_dict_values) < 4:
                current_row_number = current_row_number - len(category_dict_values) + 4

            current_row_number += 1

    ####################################################################################################################
    chart_start_row_number = current_row_number

    has_values_data = True
    has_timestamps_data = True

    if 'values' not in reporting_period_data.keys() or \
            reporting_period_data['values'] is None or \
            len(reporting_period_data['values']) == 0:
        has_values_data = False

    if 'timestamps' not in reporting_period_data.keys() or \
            reporting_period_data['timestamps'] is None or \
            len(reporting_period_data['timestamps']) == 0 or \
            len(reporting_period_data['timestamps'][0]) == 0:
        has_timestamps_data = False

    if not is_base_period_timestamp_exists_flag:
        if has_values_data and has_timestamps_data:
            ca_len = len(reporting_period_data['names'])
            time = reporting_period_data['timestamps'][0]
            real_timestamps_len = timestamps_data_not_equal_0(report['parameters']['timestamps'])
            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)] = name + ' ' + _('Detailed Data')

            current_row_number += 1

            chart_start_row_number = current_row_number

            # 1: Stand for blank line  2: Stand for title
            current_row_number += ca_len * 6 + real_timestamps_len * 6 + 1 + 2
            table_start_row_number = current_row_number

            ws.row_dimensions[current_row_number].height = 60
            ws['B' + str(current_row_number)].fill = table_fill
            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)].alignment = c_c_alignment
            ws['B' + str(current_row_number)].border = f_border
            ws['B' + str(current_row_number)] = _('Datetime')

            col = 'C'

            for i in range(0, ca_len):
                ws[col + str(current_row_number)].fill = table_fill
                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)].border = f_border
                ws[col + str(current_row_number)] = \
                    reporting_period_data['names'][i] + " (" + reporting_period_data['units'][i] + ")"
                col = chr(ord(col) + 1)

            current_row_number += 1

            for i in range(0, len(time)):
                ws['B' + str(current_row_number)].font = title_font
                ws['B' + str(current_row_number)].alignment = c_c_alignment
                ws['B' + str(current_row_number)].border = f_border
                ws['B' + str(current_row_number)] = time[i]

                col = 'C'
                for j in range(0, ca_len):
                    ws[col + str(current_row_number)].font = title_font
                    ws[col + str(current_row_number)].alignment = c_c_alignment
                    ws[col + str(current_row_number)].border = f_border
                    ws[col + str(current_row_number)] = round(reporting_period_data['values'][j][i], 2) \
                        if reporting_period_data['values'][j][i] is not None else 0.00
                    col = chr(ord(col) + 1)

                current_row_number += 1

            table_end_row_number = current_row_number - 1

            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)].alignment = c_c_alignment
            ws['B' + str(current_row_number)].border = f_border
            ws['B' + str(current_row_number)] = _('Subtotal')

            col = 'C'

            for i in range(0, ca_len):
                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)].border = f_border
                ws[col + str(current_row_number)] = round(reporting_period_data['subtotals'][i], 2)
                col = chr(ord(col) + 1)

            current_row_number += 2

            for i in range(0, ca_len):
                line = LineChart()
                line.title = _('Reporting Period Consumption') + ' - ' \
                    + reporting_period_data['names'][i] + " (" + reporting_period_data['units'][i] + ")"
                labels = Reference(ws, min_col=2,
                                   min_row=table_start_row_number + 1,
                                   max_row=table_end_row_number)
                reporting_line_data = Reference(ws, min_col=3 + i,
                                                min_row=table_start_row_number,
                                                max_row=table_end_row_number)
                line.add_data(reporting_line_data, titles_from_data=True)
                line.set_categories(labels)
                reporting_line_data = line.series[0]
                reporting_line_data.marker.symbol = "circle"
                reporting_line_data.smooth = True
                line.x_axis.crosses = 'min'
                line.height = 8.25
                line.width = 24
                line.dLbls = DataLabelList()
                line.dLbls.dLblPos = 't'
                line.dLbls.showVal = True
                line.dLbls.showPercent = False
                chart_col = 'B'
                chart_cell = chart_col + str(chart_start_row_number)
                chart_start_row_number += 6
                ws.add_chart(line, chart_cell)

    else:
        if has_values_data and has_timestamps_data:
            base_period_data = report['base_period']
            reporting_period_data = report['reporting_period']
            base_period_timestamps = base_period_data['timestamps']
            reporting_period_timestamps = reporting_period_data['timestamps']
            # Tip:
            #     base_period_data['names'] == reporting_period_data['names']
            #     base_period_data['units'] == reporting_period_data['units']
            base_period_data_ca_len = len(base_period_data['names'])
            reporting_period_data_ca_len = len(reporting_period_data['names'])
            real_timestamps_len = timestamps_data_not_equal_0(report['parameters']['timestamps'])
            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)] = name + ' ' + _('Detailed Data')
            current_row_number += 1

            chart_start_row_number = current_row_number

            # 1: Stand for blank line  2: Stand for title
            current_row_number += reporting_period_data_ca_len * 6 + real_timestamps_len * 6 + 1 + 2
            table_start_row_number = current_row_number

            ws.row_dimensions[current_row_number].height = 60
            current_col_number = 2
            col = format_cell.get_column_letter(current_col_number)
            ws[col + str(current_row_number)].fill = table_fill
            ws[col + str(current_row_number)].font = title_font
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)].border = f_border
            ws[col + str(current_row_number)] = _('Base Period') + " - " + _('Datetime')

            for i in range(0, base_period_data_ca_len):
                current_col_number += 1
                col = format_cell.get_column_letter(current_col_number)

                ws[col + str(current_row_number)].fill = table_fill
                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)] = _('Base Period') + " - " + base_period_data['names'][i] + \
                    " (" + base_period_data['units'][i] + ")"
                ws[col + str(current_row_number)].border = f_border

            current_col_number += 1
            col = format_cell.get_column_letter(current_col_number)

            ws[col + str(current_row_number)].fill = table_fill
            ws[col + str(current_row_number)].font = title_font
            ws[col + str(current_row_number)].border = f_border
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)] = _('Reporting Period') + " - " + _('Datetime')

            for i in range(0, reporting_period_data_ca_len):
                current_col_number += 1
                col = format_cell.get_column_letter(current_col_number)
                ws[col + str(current_row_number)].fill = table_fill
                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)] = _('Reporting Period') + " - " \
                    + reporting_period_data['names'][i] + " (" + \
                    reporting_period_data['units'][i] + ")"
                ws[col + str(current_row_number)].border = f_border

            current_row_number += 1

            max_timestamps_len = len(base_period_timestamps[0]) \
                if len(base_period_timestamps[0]) >= len(reporting_period_timestamps[0]) \
                else len(reporting_period_timestamps[0])

            for i in range(0, max_timestamps_len):
                current_col_number = 2
                col = format_cell.get_column_letter(current_col_number)
                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)] = base_period_timestamps[0][i] \
                    if i < len(base_period_timestamps[0]) else None
                ws[col + str(current_row_number)].border = f_border

                for j in range(0, base_period_data_ca_len):
                    current_col_number += 1
                    col = format_cell.get_column_letter(current_col_number)

                    ws[col + str(current_row_number)].font = title_font
                    ws[col + str(current_row_number)].alignment = c_c_alignment
                    ws[col + str(current_row_number)] = round(base_period_data['values'][j][i], 2) \
                        if i < len(base_period_data['values'][j]) else None
                    ws[col + str(current_row_number)].border = f_border
                current_col_number += 1
                col = format_cell.get_column_letter(current_col_number)

                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)] = reporting_period_timestamps[0][i] \
                    if i < len(reporting_period_timestamps[0]) else None
                ws[col + str(current_row_number)].border = f_border

                for j in range(0, reporting_period_data_ca_len):
                    current_col_number += 1
                    col = format_cell.get_column_letter(current_col_number)

                    ws[col + str(current_row_number)].font = title_font
                    ws[col + str(current_row_number)].alignment = c_c_alignment
                    ws[col + str(current_row_number)] = round(reporting_period_data['values'][j][i], 2) \
                        if i < len(reporting_period_data['values'][j]) else None
                    ws[col + str(current_row_number)].border = f_border

                current_row_number += 1

            current_col_number = 2
            col = format_cell.get_column_letter(current_col_number)

            ws[col + str(current_row_number)].font = title_font
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)] = _('Subtotal')
            ws[col + str(current_row_number)].border = f_border

            for i in range(0, base_period_data_ca_len):
                current_col_number += 1
                col = format_cell.get_column_letter(current_col_number)
                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)] = round(base_period_data['subtotals'][i], 2)
                ws[col + str(current_row_number)].border = f_border

            current_col_number += 1
            col = format_cell.get_column_letter(current_col_number)

            ws[col + str(current_row_number)].font = title_font
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)] = _('Subtotal')
            ws[col + str(current_row_number)].border = f_border

            for i in range(0, reporting_period_data_ca_len):
                current_col_number += 1
                col = format_cell.get_column_letter(current_col_number)
                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)] = round(reporting_period_data['subtotals'][i], 2)
                ws[col + str(current_row_number)].border = f_border

            current_row_number += 2

            for i in range(0, reporting_period_data_ca_len):
                line = LineChart()
                line.title = _('Base Period Consumption') + ' / ' \
                    + _('Reporting Period Consumption') + ' - ' \
                    + reporting_period_data['names'][i] + " (" + reporting_period_data['units'][i] + ")"
                labels = Reference(ws, min_col=2 + base_period_data_ca_len + 1,
                                   min_row=table_start_row_number + 1,
                                   max_row=table_start_row_number + len(reporting_period_timestamps[0]))
                base_line_data = Reference(ws, min_col=3 + i,
                                           min_row=table_start_row_number,
                                           max_row=table_start_row_number + len(reporting_period_timestamps[0]))
                reporting_line_data = Reference(ws, min_col=3 + base_period_data_ca_len + 1 + i,
                                                min_row=table_start_row_number,
                                                max_row=table_start_row_number + len(reporting_period_timestamps[0]))
                line.add_data(base_line_data, titles_from_data=True)
                line.add_data(reporting_line_data, titles_from_data=True)
                line.set_categories(labels)
                for j in range(len(line.series)):
                    line.series[j].marker.symbol = "circle"
                    line.series[j].smooth = True
                line.x_axis.crosses = 'min'
                line.height = 8.25
                line.width = 24
                line.dLbls = DataLabelList()
                line.dLbls.dLblPos = 't'
                line.dLbls.showVal = True
                line.dLbls.showPercent = False
                chart_col = 'B'
                chart_cell = chart_col + str(chart_start_row_number)
                chart_start_row_number += 6
                ws.add_chart(line, chart_cell)

    ####################################################################################################################
    if "associated_equipment" not in report.keys() or \
            "energy_item_names" not in report['associated_equipment'].keys() or \
            len(report['associated_equipment']["energy_item_names"]) == 0 \
            or 'associated_equipment_names_array' not in report['associated_equipment'].keys() \
            or report['associated_equipment']['associated_equipment_names_array'] is None \
            or len(report['associated_equipment']['associated_equipment_names_array']) == 0 \
            or len(report['associated_equipment']['associated_equipment_names_array'][0]) == 0:
        pass
    else:
        associated_equipment = report['associated_equipment']

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)] = name + ' ' + _('Associated Equipment Data')

        current_row_number += 1

        ws.row_dimensions[current_row_number].height = 60
        ws['B' + str(current_row_number)].fill = table_fill
        ws['B' + str(current_row_number)].font = name_font
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)].border = f_border
        ws['B' + str(current_row_number)] = _('Associated Equipment')
        ca_len = len(associated_equipment['energy_item_names'])

        for i in range(0, ca_len):
            row = chr(ord('C') + i)
            ws[row + str(current_row_number)].fill = table_fill
            ws[row + str(current_row_number)].font = name_font
            ws[row + str(current_row_number)].alignment = c_c_alignment
            ws[row + str(current_row_number)].border = f_border
            ws[row + str(current_row_number)] = \
                reporting_period_data['names'][i] + " (" + reporting_period_data['units'][i] + ")"

        associated_equipment_len = len(associated_equipment['associated_equipment_names_array'][0])

        for i in range(0, associated_equipment_len):
            current_row_number += 1
            row = str(current_row_number)

            ws['B' + row].font = title_font
            ws['B' + row].alignment = c_c_alignment
            ws['B' + row] = associated_equipment['associated_equipment_names_array'][0][i]
            ws['B' + row].border = f_border

            for j in range(0, ca_len):
                col = chr(ord('C') + j)
                ws[col + row].font = title_font
                ws[col + row].alignment = c_c_alignment
                ws[col + row] = round(associated_equipment['subtotals_array'][j][i], 2)
                ws[col + row].border = f_border
    ####################################################################################################################
    current_sheet_parameters_row_number = chart_start_row_number + 1
    has_parameters_names_and_timestamps_and_values_data = True
    if 'parameters' not in report.keys() or \
            report['parameters'] is None or \
            'names' not in report['parameters'].keys() or \
            report['parameters']['names'] is None or \
            len(report['parameters']['names']) == 0 or \
            'timestamps' not in report['parameters'].keys() or \
            report['parameters']['timestamps'] is None or \
            len(report['parameters']['timestamps']) == 0 or \
            'values' not in report['parameters'].keys() or \
            report['parameters']['values'] is None or \
            len(report['parameters']['values']) == 0 or \
            timestamps_data_all_equal_0(report['parameters']['timestamps']):

        has_parameters_names_and_timestamps_and_values_data = False
    if has_parameters_names_and_timestamps_and_values_data:

        ################################################################################################################
        # new worksheet
        ################################################################################################################

        parameters_data = report['parameters']
        parameters_names_len = len(parameters_data['names'])

        file_name = (re.sub(r'[^A-Z]', '', ws.title))+'_'
        parameters_ws = wb.create_sheet(file_name + _('Parameters'))

        parameters_timestamps_data_max_len = \
            get_parameters_timestamps_lists_max_len(list(parameters_data['timestamps']))

        # Row height
        parameters_ws.row_dimensions[1].height = 102
        for i in range(2, 7 + 1):
            parameters_ws.row_dimensions[i].height = 42

        for i in range(8, parameters_timestamps_data_max_len + 10):
            parameters_ws.row_dimensions[i].height = 60

        # Col width
        parameters_ws.column_dimensions['A'].width = 1.5

        parameters_ws.column_dimensions['B'].width = 25.0

        for i in range(3, 12+parameters_names_len*3):
            parameters_ws.column_dimensions[format_cell.get_column_letter(i)].width = 15.0

        # Img
        img = Image("excelexporters/myems.png")
        parameters_ws.add_image(img, 'A1')

        # Title
        parameters_ws['B3'].alignment = b_r_alignment
        parameters_ws['B3'] = _('Name') + ':'
        parameters_ws['C3'].border = b_border
        parameters_ws['C3'].alignment = b_c_alignment
        parameters_ws['C3'] = name

        parameters_ws['D3'].alignment = b_r_alignment
        parameters_ws['D3'] = _('Period Type') + ':'
        parameters_ws['E3'].border = b_border
        parameters_ws['E3'].alignment = b_c_alignment
        parameters_ws['E3'] = period_type

        parameters_ws['B4'].alignment = b_r_alignment
        parameters_ws['B4'] = _('Reporting Start Datetime') + ':'
        parameters_ws['C4'].border = b_border
        parameters_ws['C4'].alignment = b_c_alignment
        parameters_ws['C4'] = reporting_start_datetime_local

        parameters_ws['D4'].alignment = b_r_alignment
        parameters_ws['D4'] = _('Reporting End Datetime') + ':'
        parameters_ws['E4'].border = b_border
        parameters_ws['E4'].alignment = b_c_alignment
        parameters_ws['E4'] = reporting_end_datetime_local

        parameters_ws_current_row_number = 6

        parameters_ws['B' + str(parameters_ws_current_row_number)].font = title_font
        parameters_ws['B' + str(parameters_ws_current_row_number)] = name + ' ' + _('Parameters')

        parameters_ws_current_row_number += 1

        parameters_table_start_row_number = parameters_ws_current_row_number

        parameters_ws.row_dimensions[parameters_ws_current_row_number].height = 80

        parameters_ws_current_row_number += 1

        table_current_col_number = 2

        for i in range(0, parameters_names_len):

            if len(parameters_data['timestamps'][i]) == 0:
                continue

            col = format_cell.get_column_letter(table_current_col_number)

            parameters_ws[col + str(parameters_ws_current_row_number-1)].fill = table_fill
            parameters_ws[col + str(parameters_ws_current_row_number-1)].border = f_border

            col = format_cell.get_column_letter(table_current_col_number + 1)

            parameters_ws[col + str(parameters_ws_current_row_number-1)].fill = table_fill
            parameters_ws[col + str(parameters_ws_current_row_number-1)].border = f_border
            parameters_ws[col + str(parameters_ws_current_row_number-1)].font = name_font
            parameters_ws[col + str(parameters_ws_current_row_number-1)].alignment = c_c_alignment
            parameters_ws[col + str(parameters_ws_current_row_number-1)] = parameters_data['names'][i]

            table_current_row_number = parameters_ws_current_row_number

            for j, value in enumerate(list(parameters_data['timestamps'][i])):
                col = format_cell.get_column_letter(table_current_col_number)

                parameters_ws[col + str(table_current_row_number)].border = f_border
                parameters_ws[col + str(table_current_row_number)].font = title_font
                parameters_ws[col + str(table_current_row_number)].alignment = c_c_alignment
                parameters_ws[col + str(table_current_row_number)] = value

                col = format_cell.get_column_letter(table_current_col_number + 1)

                parameters_ws[col + str(table_current_row_number)].border = f_border
                parameters_ws[col + str(table_current_row_number)].font = title_font
                parameters_ws[col + str(table_current_row_number)].alignment = c_c_alignment
                parameters_ws[col + str(table_current_row_number)] = round(parameters_data['values'][i][j], 2)

                table_current_row_number += 1

            table_current_col_number = table_current_col_number + 3

        ################################################################################################################
        # parameters chart and parameters table
        ################################################################################################################

        ws['B' + str(current_sheet_parameters_row_number)].font = title_font
        ws['B' + str(current_sheet_parameters_row_number)] = name + ' ' + _('Parameters')

        current_sheet_parameters_row_number += 1

        chart_start_row_number = current_sheet_parameters_row_number

        col_index = 0

        for i in range(0, parameters_names_len):

            if len(parameters_data['timestamps'][i]) == 0:
                continue

            line = LineChart()
            data_col = 3+col_index*3
            labels_col = 2+col_index*3
            col_index += 1
            line.title = _('Parameters') + ' - ' + \
                parameters_ws.cell(row=parameters_table_start_row_number, column=data_col).value
            labels = Reference(parameters_ws, min_col=labels_col, min_row=parameters_table_start_row_number + 1,
                               max_row=(len(parameters_data['timestamps'][i])+parameters_table_start_row_number))
            line_data = Reference(parameters_ws, min_col=data_col, min_row=parameters_table_start_row_number,
                                  max_row=(len(parameters_data['timestamps'][i])+parameters_table_start_row_number))
            line.add_data(line_data, titles_from_data=True)
            line.set_categories(labels)
            line_data = line.series[0]
            line_data.marker.symbol = "circle"
            line_data.smooth = True
            line.x_axis.crosses = 'min'
            line.height = 8.25
            line.width = 24
            line.dLbls = DataLabelList()
            line.dLbls.dLblPos = 't'
            line.dLbls.showVal = False
            line.dLbls.showPercent = False
            chart_col = 'B'
            chart_cell = chart_col + str(chart_start_row_number)
            chart_start_row_number += 6
            ws.add_chart(line, chart_cell)

        current_sheet_parameters_row_number = chart_start_row_number

        current_sheet_parameters_row_number += 1
    ####################################################################################################################
    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename


def group_by_category(category_list):
    category_dict = dict()
    for i, value in enumerate(category_list):
        if value not in category_dict.keys():
            category_dict[value] = list()
        category_dict[value].append(i)
    return category_dict


def get_parameters_timestamps_lists_max_len(parameters_timestamps_lists):
    max_len = 0
    for i, value in enumerate(list(parameters_timestamps_lists)):
        if len(value) > max_len:
            max_len = len(value)

    return max_len


def timestamps_data_all_equal_0(lists):
    for i, value in enumerate(list(lists)):
        if len(value) > 0:
            return False

    return True


def timestamps_data_not_equal_0(lists):
    number = 0
    for i, value in enumerate(list(lists)):
        if len(value) > 0:
            number += 1
    return number


def is_base_period_timestamp_exists(base_period_data):
    timestamps = base_period_data['timestamps']

    if len(timestamps) == 0:
        return False

    for timestamp in timestamps:
        if len(timestamp) > 0:
            return True

    return False

