"""
Combined Equipment Batch Excel Exporter

This module provides functionality to export combined equipment batch data to Excel format.
It generates comprehensive reports showing energy consumption data for multiple combined equipment
within a specific space and time period.

Key Features:
- Multi-combined equipment energy consumption comparison
- Energy category breakdown with units
- Formatted Excel output with proper styling
- Multi-language support
- Base64 encoding for file transmission

The exported Excel file includes:
- Combined equipment names and associated spaces
- Energy consumption by category
- Proper formatting and borders
- Logo and header information
"""

import base64
from core.utilities import get_translation
import os
import uuid
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from core.utilities import round2


########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file from the report data
# Step 3: Encode the excel file to Base64
########################################################################################################################
def export(result, space_name, reporting_start_datetime_local, reporting_end_datetime_local, language):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if result is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(result,
                              space_name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              language)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        print(str(ex))

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        print(str(ex))
    return base64_message


def generate_excel(report, space_name, reporting_start_datetime_local, reporting_end_datetime_local, language):

    trans = get_translation(language)
    trans.install()
    _ = trans.gettext

    wb = Workbook()
    ws = wb.active
    ws.title = "CombinedEquipmentBatch"

    # Row height
    ws.row_dimensions[1].height = 102
    for i in range(2, 5 + 1):
        ws.row_dimensions[i].height = 42

    for i in range(6, len(report['combined_equipments']) + 15):
        ws.row_dimensions[i].height = 60

    # Col width
    ws.column_dimensions['A'].width = 1.5

    ws.column_dimensions['B'].width = 25.0

    for i in range(ord('C'), ord('L')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Arial', size=15, bold=True)
    title_font = Font(name='Arial', size=15, bold=True)
    data_font = Font(name='Franklin Gothic Book', size=11)

    table_fill = PatternFill(fill_type='solid', fgColor='90ee90')
    f_border = Border(left=Side(border_style='medium'),
                      right=Side(border_style='medium'),
                      bottom=Side(border_style='medium'),
                      top=Side(border_style='medium')
                      )
    b_border = Border(
        bottom=Side(border_style='medium'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)

    # Img
    img = Image("excelexporters/myems.png")
    ws.add_image(img, 'A1')

    # Title
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = _('Space') + ':'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'] = space_name

    ws['B4'].alignment = b_r_alignment
    ws['B4'] = _('Reporting Start Datetime') + ':'
    ws['C4'].border = b_border
    ws['C4'].alignment = b_c_alignment
    ws['C4'] = reporting_start_datetime_local

    ws['B5'].alignment = b_r_alignment
    ws['B5'] = _('Reporting End Datetime') + ':'
    ws['C5'].border = b_border
    ws['C5'].alignment = b_c_alignment
    ws['C5'] = reporting_end_datetime_local

    # Title
    ws['B6'].border = f_border
    ws['B6'].font = name_font
    ws['B6'].alignment = c_c_alignment
    ws['B6'].fill = table_fill
    ws['B6'] = _('Name')

    ws['C6'].border = f_border
    ws['C6'].alignment = c_c_alignment
    ws['C6'].font = name_font
    ws['C6'].fill = table_fill
    ws['C6'] = _('Space')

    ca_len = len(report['energycategories'])

    for i in range(0, ca_len):
        col = chr(ord('D') + i)
        ws[col + '6'].fill = table_fill
        ws[col + '6'].font = name_font
        ws[col + '6'].alignment = c_c_alignment
        ws[col + '6'] = report['energycategories'][i]['name'] + \
            " (" + report['energycategories'][i]['unit_of_measure'] + ")"
        ws[col + '6'].border = f_border

    current_row_number = 7
    for i in range(0, len(report['combined_equipments'])):

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)].border = f_border
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)] = report['combined_equipments'][i]['combined_equipment_name']

        ws['C' + str(current_row_number)].font = title_font
        ws['C' + str(current_row_number)].border = f_border
        ws['C' + str(current_row_number)].alignment = c_c_alignment
        ws['C' + str(current_row_number)] = report['combined_equipments'][i]['space_name']

        ca_len = len(report['combined_equipments'][i]['values'])
        for j in range(0, ca_len):
            col = chr(ord('D') + j)
            ws[col + str(current_row_number)].font = data_font
            ws[col + str(current_row_number)].border = f_border
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)] = round2(report['combined_equipments'][i]['values'][j], 2)

        current_row_number += 1

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename
