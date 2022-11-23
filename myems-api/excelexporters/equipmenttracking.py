import base64
import gettext
import os
import uuid

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Alignment, Font


########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file from the report data
# Step 3: Encode the excel file to Base64
########################################################################################################################

def export(result, space_name, language):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if result is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(result,
                              space_name,
                              language)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        print(str(ex))

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        print(str(ex))
    return base64_message


def generate_excel(report, space_name, language):
    locale_path = './i18n/'
    if language == 'zh_CN':
        trans = gettext.translation('myems', locale_path, languages=['zh_CN'])
    elif language == 'de':
        trans = gettext.translation('myems', locale_path, languages=['de'])
    elif language == 'en':
        trans = gettext.translation('myems', locale_path, languages=['en'])
    else:
        trans = gettext.translation('myems', locale_path, languages=['en'])
    trans.install()
    _ = trans.gettext
    wb = Workbook()
    ws = wb.active
    ws.title = "EquipmentTracking"

    # Row height
    ws.row_dimensions[1].height = 118
    for i in range(2, 5000 + 1):
        ws.row_dimensions[i].height = 30
    # Col width
    ws.column_dimensions['A'].width = 1

    # Font
    name_font = Font(name='Arial', size=15, bold=True)
    title_font = Font(name='Arial', size=15, bold=True)

    f_border = Border(left=Side(border_style='medium'),
                      right=Side(border_style='medium'),
                      bottom=Side(border_style='medium'),
                      top=Side(border_style='medium')
                      )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    for i in range(ord('B'), ord('F')):
        ws.column_dimensions[chr(i)].width = 30.0

    # Img
    img = Image("excelexporters/myems.png")
    ws.add_image(img, 'A1')

    # Title
    ws['B3'].border = f_border
    ws['B3'].font = name_font
    ws['B3'].alignment = b_c_alignment
    ws['B3'] = _('Name') + ':'

    ws['C3'].border = f_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'].font = name_font
    ws['C3'] = _('Space') + ':'

    ws['D3'].border = f_border
    ws['D3'].font = name_font
    ws['D3'].alignment = b_c_alignment
    ws['D3'] = _('Cost Center') + ':'

    ws['E3'].border = f_border
    ws['E3'].alignment = b_c_alignment
    ws['E3'].font = name_font
    ws['E3'] = _('Description') + ':'

    current_row_number = 4
    for i in range(0, len(report['equipments'])):

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)].border = f_border
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)] = report['equipments'][i]['equipment_name']

        ws['C' + str(current_row_number)].font = title_font
        ws['C' + str(current_row_number)].border = f_border
        ws['C' + str(current_row_number)].alignment = c_c_alignment
        ws['C' + str(current_row_number)] = report['equipments'][i]['space_name']

        ws['D' + str(current_row_number)].font = title_font
        ws['D' + str(current_row_number)].border = f_border
        ws['D' + str(current_row_number)].alignment = c_c_alignment
        ws['D' + str(current_row_number)] = report['equipments'][i]['cost_center_name']

        ws['E' + str(current_row_number)].font = title_font
        ws['E' + str(current_row_number)].border = f_border
        ws['E' + str(current_row_number)].alignment = c_c_alignment
        ws['E' + str(current_row_number)] = report['equipments'][i]['description']
        current_row_number += 1

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename
