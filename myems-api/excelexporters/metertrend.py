import base64
import gettext
import os
import re
import uuid

import openpyxl.utils.cell as format_cell
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font


########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file
# Step 3: Encode the excel file bytes to Base64
########################################################################################################################


def export(result,
           name,
           reporting_start_datetime_local,
           reporting_end_datetime_local,
           period_type,
           language):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if result is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(result,
                              name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              period_type,
                              language)

    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        print(str(ex))

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        print(str(ex))
    return base64_message


def generate_excel(report,
                   name,
                   reporting_start_datetime_local,
                   reporting_end_datetime_local,
                   period_type,
                   language):

    locale_path = './i18n/'
    if language == 'zh_CN':
        trans = gettext.translation('myems', locale_path, languages=['zh_CN'])
    elif language == 'de':
        trans = gettext.translation('myems', locale_path, languages=['de'])
    elif language == 'en':
        trans = gettext.translation('myems', locale_path, languages=['en'])
    else:
        trans = gettext.translation('myems', locale_path, languages=['en'])
    trans.install()
    _ = trans.gettext

    wb = Workbook()
    ws = wb.active
    ws.title = "MeterTrend"

    # Row height
    ws.row_dimensions[1].height = 102
    for i in range(2, 8):
        ws.row_dimensions[i].height = 42

    # Col width
    ws.column_dimensions['A'].width = 1.5

    ws.column_dimensions['B'].width = 25.0

    for i in range(ord('C'), ord('V')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Arial', size=15, bold=True)
    title_font = Font(name='Arial', size=15, bold=True)

    table_fill = PatternFill(fill_type='solid', fgColor='90ee90')
    f_border = Border(left=Side(border_style='medium'),
                      right=Side(border_style='medium'),
                      bottom=Side(border_style='medium'),
                      top=Side(border_style='medium')
                      )
    b_border = Border(
        bottom=Side(border_style='medium'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)

    # Img
    img = Image("excelexporters/myems.png")
    ws.add_image(img, 'A1')

    # Title
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = _('Name') + ':'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'] = name

    ws['B4'].alignment = b_r_alignment
    ws['B4'] = _('Reporting Start Datetime') + ':'
    ws['C4'].border = b_border
    ws['C4'].alignment = b_c_alignment
    ws['C4'] = reporting_start_datetime_local

    ws['B5'].alignment = b_r_alignment
    ws['B5'] = _('Reporting End Datetime') + ':'
    ws['C5'].border = b_border
    ws['C5'].alignment = b_c_alignment
    ws['C5'] = reporting_end_datetime_local

    if "reporting_period" not in report.keys() or \
            "names" not in report['reporting_period'].keys() or len(report['reporting_period']['names']) == 0:
        filename = str(uuid.uuid4()) + '.xlsx'
        wb.save(filename)

        return filename
    ####################################################################################################################
    # First: Trend
    # 6: title
    # 7: table title
    # 8~ table_data
    ####################################################################################################################
    has_data_flag = True
    report['reporting_period'] = report['reporting_period']
    if "names" not in report['reporting_period'].keys() or \
            report['reporting_period']['names'] is None or \
            len(report['reporting_period']['names']) == 0:
        has_data_flag = False

    if "timestamps" not in report['reporting_period'].keys() or \
            report['reporting_period']['timestamps'] is None or \
            len(report['reporting_period']['timestamps']) == 0:
        has_data_flag = False

    if "values" not in report['reporting_period'].keys() or \
            report['reporting_period']['values'] is None or \
            len(report['reporting_period']['values']) == 0:
        has_data_flag = False

    ca_len = len(report['reporting_period']['names'])
    times = report['reporting_period']['timestamps']
    category = report['meter']['energy_category_name']
    parameters_names_len = len(report['parameters']['names'])
    parameters_data = report['parameters']
    parameters_parameters_datas_len = 0
    for i in range(0, parameters_names_len):
        if len(parameters_data['timestamps'][i]) == 0:
            continue
        parameters_parameters_datas_len += 1
    start_detail_data_row_num = 9 + (parameters_parameters_datas_len + ca_len) * 6
    if has_data_flag:
        time = times[0]
        for time in times:
            if len(time) > 0:
                break
        has_data = False
        current_sheet_parameters_row_number = 7
        for i in range(8, len(time) + 6 + ca_len * 6 + len(category) * 6 + 2):
            ws.row_dimensions[i].height = 42
        if len(time) > 0:
            has_data = True
            current_sheet_parameters_row_number = 7 + ca_len * 6
        if has_data:

            max_row = start_detail_data_row_num + len(time)
            ws['B6'].font = title_font
            ws['B6'] = name + ' ' + _('Trend')

            ws.row_dimensions[start_detail_data_row_num - 1].height = 60
            ws['B' + str(start_detail_data_row_num - 1)].fill = table_fill
            ws['B' + str(start_detail_data_row_num - 1)].font = title_font
            ws['B' + str(start_detail_data_row_num - 1)].border = f_border
            ws['B' + str(start_detail_data_row_num - 1)].alignment = c_c_alignment
            ws['B' + str(start_detail_data_row_num - 1)] = _('Datetime')

            for i in range(0, len(time)):
                col = 'B'
                row = str(start_detail_data_row_num + i)
                ws[col + row].font = title_font
                ws[col + row].alignment = c_c_alignment
                ws[col + row] = time[i]
                ws[col + row].border = f_border

            for i in range(0, ca_len):
                # 38 title
                col = format_cell.get_column_letter(3 + i)

                ws[col + str(start_detail_data_row_num - 1)].fill = table_fill
                ws[col + str(start_detail_data_row_num - 1)].font = title_font
                ws[col + str(start_detail_data_row_num - 1)].alignment = c_c_alignment
                ws[col + str(start_detail_data_row_num - 1)] = report['reporting_period']['names'][i]
                ws[col + str(start_detail_data_row_num - 1)].border = f_border

                for j in range(0, len(time)):

                    row = str(start_detail_data_row_num + j)
                    ws[col + row].font = title_font
                    ws[col + row].alignment = c_c_alignment
                    try:
                        ws[col + row] = round(report['reporting_period']['values'][i][j], 3) if \
                            len(report['reporting_period']['values'][i]) > 0 and \
                            len(report['reporting_period']['values'][i]) > j and \
                            report['reporting_period']['values'][i][j] is not None else ''
                    except Exception as e:
                        print('error 1 in excelexporters\\metertrend: ' + str(e))

                    ws[col + row].border = f_border
            # line
            # 39~: line
                line = LineChart()
                line.title = report['reporting_period']['names'][i]
                labels = Reference(ws, min_col=2, min_row=start_detail_data_row_num, max_row=max_row-1)
                line_data = Reference(ws, min_col=3 + i, min_row=start_detail_data_row_num+1, max_row=max_row-1)
                line.add_data(line_data, titles_from_data=True)
                line.set_categories(labels)
                line_data.smooth = True
                line.x_axis.crosses = 'min'
                line.height = 8.25
                line.width = 36
                line.dLbls = DataLabelList()
                line.dLbls.dLblPos = 't'
                line.dLbls.showVal = False  # val show
                line.dLbls.showPercent = False  # percent show
                chart_col = chr(ord('B'))
                chart_cell = chart_col + str(7 + 6*i)

                ws.add_chart(line, chart_cell)

    ####################################################################################################################
    # 12 is the starting line number of the last line chart in the report period
    if 'parameters' not in report.keys() or \
            report['parameters'] is None or \
            'names' not in report['parameters'].keys() or \
            report['parameters']['names'] is None or \
            len(report['parameters']['names']) == 0 or \
            'timestamps' not in report['parameters'].keys() or \
            report['parameters']['timestamps'] is None or \
            len(report['parameters']['timestamps']) == 0 or \
            'values' not in report['parameters'].keys() or \
            report['parameters']['values'] is None or \
            len(report['parameters']['values']) == 0 or \
            timestamps_data_all_equal_0(report['parameters']['timestamps']):
        pass
    else:

        ################################################################################################################
        # new worksheet
        ################################################################################################################

        parameters_data = report['parameters']

        parameters_names_len = len(parameters_data['names'])

        file_name = (re.sub(r'[^A-Z]', '', ws.title))+'_'
        parameters_ws = wb.create_sheet(file_name + _('Parameters'))

        parameters_timestamps_data_max_len = \
            get_parameters_timestamps_lists_max_len(list(parameters_data['timestamps']))

        # Row height
        parameters_ws.row_dimensions[1].height = 102
        for i in range(2, 7 + 1):
            parameters_ws.row_dimensions[i].height = 42

        for i in range(8, parameters_timestamps_data_max_len + 10):
            parameters_ws.row_dimensions[i].height = 60

        # Col width
        parameters_ws.column_dimensions['A'].width = 1.5

        parameters_ws.column_dimensions['B'].width = 25.0

        for i in range(3, 12 + parameters_names_len * 3):
            parameters_ws.column_dimensions[format_cell.get_column_letter(i)].width = 15.0

        # Img
        img = Image("excelexporters/myems.png")
        parameters_ws.add_image(img, 'A1')

        # Title
        parameters_ws['B3'].alignment = b_r_alignment
        parameters_ws['B3'] = _('Name') + ':'
        parameters_ws['C3'].border = b_border
        parameters_ws['C3'].alignment = b_c_alignment
        parameters_ws['C3'] = name

        parameters_ws['B4'].alignment = b_r_alignment
        parameters_ws['B4'] = _('Reporting Start Datetime') + ':'
        parameters_ws['C4'].border = b_border
        parameters_ws['C4'].alignment = b_c_alignment
        parameters_ws['C4'] = reporting_start_datetime_local

        parameters_ws['B5'].alignment = b_r_alignment
        parameters_ws['B5'] = _('Reporting End Datetime') + ':'
        parameters_ws['C5'].border = b_border
        parameters_ws['C5'].alignment = b_c_alignment
        parameters_ws['C5'] = reporting_end_datetime_local

        parameters_ws_current_row_number = 6

        parameters_ws['B' + str(parameters_ws_current_row_number)].font = title_font
        parameters_ws['B' + str(parameters_ws_current_row_number)] = name + ' ' + _('Parameters')

        parameters_ws_current_row_number += 1

        parameters_table_start_row_number = parameters_ws_current_row_number

        parameters_ws.row_dimensions[parameters_ws_current_row_number].height = 80

        parameters_ws_current_row_number += 1

        table_current_col_number = 2

        for i in range(0, parameters_names_len):

            if len(parameters_data['timestamps'][i]) == 0:
                continue

            col = format_cell.get_column_letter(table_current_col_number)

            parameters_ws[col + str(parameters_ws_current_row_number - 1)].fill = table_fill
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].border = f_border

            col = format_cell.get_column_letter(table_current_col_number + 1)

            parameters_ws[col + str(parameters_ws_current_row_number - 1)].fill = table_fill
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].border = f_border
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].font = name_font
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].alignment = c_c_alignment
            parameters_ws[col + str(parameters_ws_current_row_number - 1)] = parameters_data['names'][i]

            table_current_row_number = parameters_ws_current_row_number

            for j, value in enumerate(list(parameters_data['timestamps'][i])):
                col = format_cell.get_column_letter(table_current_col_number)

                parameters_ws[col + str(table_current_row_number)].border = f_border
                parameters_ws[col + str(table_current_row_number)].font = title_font
                parameters_ws[col + str(table_current_row_number)].alignment = c_c_alignment
                parameters_ws[col + str(table_current_row_number)] = value

                col = format_cell.get_column_letter(table_current_col_number + 1)

                parameters_ws[col + str(table_current_row_number)].border = f_border
                parameters_ws[col + str(table_current_row_number)].font = title_font
                parameters_ws[col + str(table_current_row_number)].alignment = c_c_alignment
                try:
                    parameters_ws[col + str(table_current_row_number)] = round(parameters_data['values'][i][j], 2)
                except Exception as e:
                    print('error 2 in excelexporters\\metertrend: ' + str(e))
                
                table_current_row_number += 1

            table_current_col_number = table_current_col_number + 3

        ################################################################################################################
        # parameters chart and parameters table
        ################################################################################################################

        ws['B' + str(current_sheet_parameters_row_number)].font = title_font
        ws['B' + str(current_sheet_parameters_row_number)] = name + ' ' + _('Parameters')

        current_sheet_parameters_row_number += 1

        chart_start_row_number = current_sheet_parameters_row_number

        col_index = 0

        for i in range(0, parameters_names_len):

            if len(parameters_data['timestamps'][i]) == 0:
                continue

            line = LineChart()
            data_col = 3 + col_index * 3
            labels_col = 2 + col_index * 3
            col_index += 1
            line.title = _('Parameters') + ' - ' + \
                parameters_ws.cell(row=parameters_table_start_row_number, column=data_col).value
            labels = Reference(parameters_ws, min_col=labels_col, min_row=parameters_table_start_row_number + 1,
                               max_row=(len(parameters_data['timestamps'][i]) + parameters_table_start_row_number))
            line_data = Reference(parameters_ws, min_col=data_col, min_row=parameters_table_start_row_number,
                                  max_row=(len(parameters_data['timestamps'][i]) + parameters_table_start_row_number))
            line.add_data(line_data, titles_from_data=True)
            line.set_categories(labels)
            line_data = line.series[0]
            line_data.marker.symbol = "circle"
            line_data.smooth = True
            line.x_axis.crosses = 'min'
            line.height = 8.25
            line.width = 24
            line.dLbls = DataLabelList()
            line.dLbls.dLblPos = 't'
            line.dLbls.showVal = False
            line.dLbls.showPercent = False
            chart_col = 'B'
            chart_cell = chart_col + str(chart_start_row_number)
            chart_start_row_number += 6
            ws.add_chart(line, chart_cell)

        current_sheet_parameters_row_number = chart_start_row_number

        current_sheet_parameters_row_number += 1

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename


def timestamps_data_all_equal_0(lists):
    for i, value in enumerate(list(lists)):
        if len(value) > 0:
            return False

    return True


def get_parameters_timestamps_lists_max_len(parameters_timestamps_lists):
    max_len = 0
    for i, value in enumerate(list(parameters_timestamps_lists)):
        if len(value) > max_len:
            max_len = len(value)

    return max_len


def timestamps_data_not_equal_0(lists):
    number = 0
    for i, value in enumerate(list(lists)):
        if len(value) > 0:
            number += 1
    return number
