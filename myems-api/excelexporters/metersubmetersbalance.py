import base64
import uuid
import os
from openpyxl.chart import (
    BarChart,
    LineChart,
    Reference,
)
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl.chart.label import DataLabelList


####################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excelexporters file
# Step 3: Encode the excelexporters file to Base64
####################################################################################################################

def export(result, name, reporting_start_datetime_local, reporting_end_datetime_local, period_type):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if result is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(result,
                              name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              period_type)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        pass

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        pass
    return base64_message


def generate_excel(report, name, reporting_start_datetime_local, reporting_end_datetime_local, period_type):
    wb = Workbook()
    ws = wb.active

    # Row height
    ws.row_dimensions[1].height = 102
    for i in range(2, 2000 + 1):
        ws.row_dimensions[i].height = 42

    # Col width
    ws.column_dimensions['A'].width = 1.5

    ws.column_dimensions['B'].width = 25.0

    for i in range(ord('C'), ord('L')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Constantia', size=15, bold=True)
    title_font = Font(name='宋体', size=15, bold=True)
    data_font = Font(name='Franklin Gothic Book', size=11)

    table_fill = PatternFill(fill_type='solid', fgColor='1F497D')
    f_border = Border(left=Side(border_style='medium', color='00000000'),
                      right=Side(border_style='medium', color='00000000'),
                      bottom=Side(border_style='medium', color='00000000'),
                      top=Side(border_style='medium', color='00000000')
                      )
    b_border = Border(
        bottom=Side(border_style='medium', color='00000000'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_r_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)

    # Img
    img = Image("excelexporters/myems.png")
    img.width = img.width * 0.85
    img.height = img.height * 0.85
    # img = Image("myems.png")
    ws.add_image(img, 'B1')

    # Title
    ws.row_dimensions[3].height = 60

    ws['B3'].font = name_font
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = 'Name:'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'].font = name_font
    ws['C3'] = name

    ws['D3'].font = name_font
    ws['D3'].alignment = b_r_alignment
    ws['D3'] = 'Period:'
    ws['E3'].border = b_border
    ws['E3'].alignment = b_c_alignment
    ws['E3'].font = name_font
    ws['E3'] = period_type

    ws['F3'].font = name_font
    ws['F3'].alignment = b_r_alignment
    ws['F3'] = 'Date:'
    ws['G3'].alignment = b_c_alignment
    ws['G3'].font = name_font
    ws['G3'] = reporting_start_datetime_local + "__" + reporting_end_datetime_local
    ws.merge_cells("G3:H3")

    if "reporting_period" not in report.keys() or \
            "difference_values" not in report['reporting_period'].keys() or \
            len(report['reporting_period']['difference_values']) == 0:
        filename = str(uuid.uuid4()) + '.xlsx'
        wb.save(filename)

        return filename
    #################################################

    has_difference_values_data_flag = True
    if 'difference_values' not in report['reporting_period'].keys() or len(
            report['reporting_period']['difference_values']) == 0:
        has_difference_values_data_flag = False

    current_row_number = 6

    if has_difference_values_data_flag:
        reporting_period_data = report['reporting_period']
        category = report['meter']['energy_category_name']

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)] = name + ' 报告期'

        current_row_number += 1

        ws.row_dimensions[current_row_number].height = 60

        ws['B' + str(current_row_number)].fill = table_fill
        ws['B' + str(current_row_number)].border = f_border
        if not isinstance(category, list):
            ws['C' + str(current_row_number)].fill = table_fill
            ws['C' + str(current_row_number)].font = name_font
            ws['C' + str(current_row_number)].alignment = c_c_alignment
            ws['C' + str(current_row_number)].border = f_border
            ws['C' + str(current_row_number)] = report['meter']['energy_category_name'] + " (" + report['meter'][
                'unit_of_measure'] + ")"

            current_row_number += 1

            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)].alignment = c_c_alignment
            ws['B' + str(current_row_number)].border = f_border
            ws['B' + str(current_row_number)] = '总表消耗'

            ws['C' + str(current_row_number)].font = name_font
            ws['C' + str(current_row_number)].alignment = c_c_alignment
            ws['C' + str(current_row_number)].border = f_border
            ws['C' + str(current_row_number)] = round(reporting_period_data['master_meter_consumption_in_category'], 2)

            current_row_number += 1

            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)].alignment = c_c_alignment
            ws['B' + str(current_row_number)].border = f_border
            ws['B' + str(current_row_number)] = '分表消耗'

            ws['C' + str(current_row_number)].font = name_font
            ws['C' + str(current_row_number)].alignment = c_c_alignment
            ws['C' + str(current_row_number)].border = f_border
            ws['C' + str(current_row_number)] = round(reporting_period_data['submeters_consumption_in_category'], 2)

            current_row_number += 1

            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)].alignment = c_c_alignment
            ws['B' + str(current_row_number)].border = f_border
            ws['B' + str(current_row_number)] = '差值'

            ws['C' + str(current_row_number)].font = name_font
            ws['C' + str(current_row_number)].alignment = c_c_alignment
            ws['C' + str(current_row_number)].border = f_border
            ws['C' + str(current_row_number)] = round(reporting_period_data['difference_in_category'], 2)

            current_row_number += 1

            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)].alignment = c_c_alignment
            ws['B' + str(current_row_number)].border = f_border
            ws['B' + str(current_row_number)] = '差值百分比'

            ws['C' + str(current_row_number)].font = name_font
            ws['C' + str(current_row_number)].alignment = c_c_alignment
            ws['C' + str(current_row_number)].border = f_border
            ws['C' + str(current_row_number)] = str(
                round(reporting_period_data['percentage_difference'] * 100, 2)) + '%'

            current_row_number += 2

            time = reporting_period_data['timestamps']
            has_time_data_flag = False
            if time is not None and len(time) > 0:
                has_time_data_flag = True

            if has_time_data_flag:

                ws['B' + str(current_row_number)].font = title_font
                ws['B' + str(current_row_number)] = name + ' 详细数据'

                current_row_number += 1
                chart_start_number = current_row_number
                current_row_number = current_row_number + 6
                table_start_number = current_row_number

                ws.row_dimensions[current_row_number].height = 60

                ws['B' + str(current_row_number)].fill = table_fill
                ws['B' + str(current_row_number)].font = title_font
                ws['B' + str(current_row_number)].border = f_border
                ws['B' + str(current_row_number)].alignment = c_c_alignment
                ws['B' + str(current_row_number)] = '日期时间'

                ws['C' + str(current_row_number)].fill = table_fill
                ws['C' + str(current_row_number)].font = title_font
                ws['C' + str(current_row_number)].border = f_border
                ws['C' + str(current_row_number)].alignment = c_c_alignment
                ws['C' + str(current_row_number)] = report['meter']['energy_category_name'] + " (" + report['meter'][
                    'unit_of_measure'] + ")"

                current_row_number += 1

                for i in range(0, len(time)):
                    ws['B' + str(current_row_number)].font = title_font
                    ws['B' + str(current_row_number)].border = f_border
                    ws['B' + str(current_row_number)].alignment = c_c_alignment
                    ws['B' + str(current_row_number)] = time[i]

                    ws['C' + str(current_row_number)].font = title_font
                    ws['C' + str(current_row_number)].border = f_border
                    ws['C' + str(current_row_number)].alignment = c_c_alignment
                    ws['C' + str(current_row_number)] = round(reporting_period_data['difference_values'][i], 2)

                    current_row_number += 1

                table_end_number = current_row_number - 1

                line = LineChart()
                line.title = '报告期差值 - ' + report['meter']['energy_category_name'] + " (" + report['meter'][
                    'unit_of_measure'] + ")"
                labels = Reference(ws, min_col=2, min_row=table_start_number + 1, max_row=table_end_number)
                line_data = Reference(ws, min_col=3, min_row=table_start_number, max_row=table_end_number)
                line.add_data(line_data, titles_from_data=True)
                line.set_categories(labels)
                line_data = line.series[0]
                line_data.marker.symbol = "circle"
                line_data.smooth = True
                line.x_axis.crosses = 'min'
                line.height = 8.25
                line.width = 24
                line.dLbls = DataLabelList()
                line.dLbls.dLblPos = 't'
                line.dLbls.showVal = True  # 数量显示
                ws.add_chart(line, "B" + str(chart_start_number))

                ws['B' + str(current_row_number)].font = title_font
                ws['B' + str(current_row_number)].border = f_border
                ws['B' + str(current_row_number)].alignment = c_c_alignment
                ws['B' + str(current_row_number)] = '总计'

                ws['C' + str(current_row_number)].font = title_font
                ws['C' + str(current_row_number)].border = f_border
                ws['C' + str(current_row_number)].alignment = c_c_alignment
                ws['C' + str(current_row_number)] = round(reporting_period_data['master_meter_consumption_in_category']
                                                          , 2)

        else:
            pass

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename
