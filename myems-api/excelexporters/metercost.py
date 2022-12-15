import base64
import gettext
import os
import re
import uuid

import openpyxl.utils.cell as format_cell
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font


########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excelexporters file
# Step 3: Encode the excelexporters file to Base64
########################################################################################################################

def export(report,
           name,
           base_period_start_datetime_local,
           base_period_end_datetime_local,
           reporting_start_datetime_local,
           reporting_end_datetime_local,
           period_type,
           language):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if report is None:
        return None

    if "reporting_period" not in report.keys() or \
            "values" not in report['reporting_period'].keys() or len(report['reporting_period']['values']) == 0:
        return None
    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(report,
                              name,
                              base_period_start_datetime_local,
                              base_period_end_datetime_local,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              period_type,
                              language)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        print(str(ex))

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        print(str(ex))
    return base64_message


def generate_excel(report,
                   name,
                   base_period_start_datetime_local,
                   base_period_end_datetime_local,
                   reporting_start_datetime_local,
                   reporting_end_datetime_local,
                   period_type,
                   language):

    locale_path = './i18n/'
    if language == 'zh_CN':
        trans = gettext.translation('myems', locale_path, languages=['zh_CN'])
    elif language == 'de':
        trans = gettext.translation('myems', locale_path, languages=['de'])
    elif language == 'en':
        trans = gettext.translation('myems', locale_path, languages=['en'])
    else:
        trans = gettext.translation('myems', locale_path, languages=['en'])
    trans.install()
    _ = trans.gettext

    wb = Workbook()

    ws = wb.active
    ws.title = "MeterCost"
    # Row height
    ws.row_dimensions[1].height = 102
    for i in range(2, 2000 + 1):
        ws.row_dimensions[i].height = 42

    # Col width
    ws.column_dimensions['A'].width = 1.5

    ws.column_dimensions['B'].width = 25.0

    for i in range(ord('C'), ord('Z')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Arial', size=15, bold=True)
    title_font = Font(name='Arial', size=15, bold=True)

    table_fill = PatternFill(fill_type='solid', fgColor='90ee90')
    f_border = Border(left=Side(border_style='medium'),
                      right=Side(border_style='medium'),
                      bottom=Side(border_style='medium'),
                      top=Side(border_style='medium')
                      )
    b_border = Border(
        bottom=Side(border_style='medium'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)

    # Img
    img = Image("excelexporters/myems.png")
    ws.add_image(img, 'A1')

    # Title
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = _('Name') + ':'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'] = name

    ws['D3'].alignment = b_r_alignment
    ws['D3'] = _('Period Type') + ':'
    ws['E3'].border = b_border
    ws['E3'].alignment = b_c_alignment
    ws['E3'] = period_type

    ws['B4'].alignment = b_r_alignment
    ws['B4'] = _('Reporting Start Datetime') + ':'
    ws['C4'].border = b_border
    ws['C4'].alignment = b_c_alignment
    ws['C4'] = reporting_start_datetime_local

    ws['D4'].alignment = b_r_alignment
    ws['D4'] = _('Reporting End Datetime') + ':'
    ws['E4'].border = b_border
    ws['E4'].alignment = b_c_alignment
    ws['E4'] = reporting_end_datetime_local

    is_base_period_timestamp_exists_flag = is_base_period_timestamp_exists(report['base_period'])

    if is_base_period_timestamp_exists_flag:
        ws['B5'].alignment = b_r_alignment
        ws['B5'] = _('Base Period Start Datetime') + ':'
        ws['C5'].border = b_border
        ws['C5'].alignment = b_c_alignment
        ws['C5'] = base_period_start_datetime_local

        ws['D5'].alignment = b_r_alignment
        ws['D5'] = _('Base Period End Datetime') + ':'
        ws['E5'].border = b_border
        ws['E5'].alignment = b_c_alignment
        ws['E5'] = base_period_end_datetime_local

    if "reporting_period" not in report.keys() or \
            "values" not in report['reporting_period'].keys() or len(report['reporting_period']['values']) == 0:
        filename = str(uuid.uuid4()) + '.xlsx'
        wb.save(filename)

        return filename

    ####################################################################################################################
    if "values" not in report['reporting_period'].keys() or len(report['reporting_period']['values']) == 0:
        for i in range(7, 10 + 1):
            ws.row_dimensions[i].height = 0.1
    else:
        ws['B7'].font = title_font
        ws['B7'] = name + _('Reporting Period Costs')

        reporting_period_data = report['reporting_period']

        ws.row_dimensions[8].height = 60
        ws['B8'].fill = table_fill
        ws['B8'].border = f_border

        ws['B9'].font = title_font
        ws['B9'].alignment = c_c_alignment
        ws['B9'] = _('Cost')
        ws['B9'].border = f_border

        ws['B10'].font = title_font
        ws['B10'].alignment = c_c_alignment
        ws['B10'] = _('Increment Rate')
        ws['B10'].border = f_border

        ws['C8'].fill = table_fill
        ws['C8'].font = name_font
        ws['C8'].alignment = c_c_alignment
        ws['C8'] = report['meter']['energy_category_name'] + " (" + report['meter']['unit_of_measure'] + ")"
        ws['C8'].border = f_border

        ws['C9'].font = name_font
        ws['C9'].alignment = c_c_alignment
        ws['C9'] = round(reporting_period_data['total_in_category'], 2)
        ws['C9'].border = f_border

        ws['C10'].font = name_font
        ws['C10'].alignment = c_c_alignment
        ws['C10'] = str(round(reporting_period_data['increment_rate'] * 100, 2)) + "%" \
            if reporting_period_data['increment_rate'] is not None else "-"
        ws['C10'].border = f_border

        # TCE
        ws['D8'].fill = table_fill
        ws['D8'].font = name_font
        ws['D8'].alignment = c_c_alignment
        ws['D8'] = _('Ton of Standard Coal') + '(TCE)'
        ws['D8'].border = f_border

        ws['D9'].font = name_font
        ws['D9'].alignment = c_c_alignment
        ws['D9'] = round(reporting_period_data['total_in_kgce'] / 1000, 2)
        ws['D9'].border = f_border

        ws['D10'].font = name_font
        ws['D10'].alignment = c_c_alignment
        ws['D10'] = str(round(reporting_period_data['increment_rate'] * 100, 2)) + "%" \
            if reporting_period_data['increment_rate'] is not None else "-"
        ws['D10'].border = f_border

        # TCO2E
        ws['E8'].fill = table_fill
        ws['E8'].font = name_font
        ws['E8'].alignment = c_c_alignment
        ws['E8'] = _('Ton of Carbon Dioxide Emissions') + '(TCO2E)'
        ws['E8'].border = f_border

        ws['E9'].font = name_font
        ws['E9'].alignment = c_c_alignment
        ws['E9'] = round(reporting_period_data['total_in_kgco2e'] / 1000, 2)
        ws['E9'].border = f_border

        ws['E10'].font = name_font
        ws['E10'].alignment = c_c_alignment
        ws['E10'] = str(round(reporting_period_data['increment_rate'] * 100, 2)) + "%" \
            if reporting_period_data['increment_rate'] is not None else "-"
        ws['E10'].border = f_border

    ####################################################################################################################
    current_row_number = 12

    table_start_draw_flag = current_row_number + 1
    reporting_period_data = report['reporting_period']

    if "timestamps" not in reporting_period_data.keys() or \
            reporting_period_data['timestamps'] is None or \
            len(reporting_period_data['timestamps']) == 0:
        pass
    else:
        if not is_base_period_timestamp_exists_flag:
            reporting_period_data = report['reporting_period']
            time = reporting_period_data['timestamps']
            real_timestamps_len = timestamps_data_not_equal_0(report['parameters']['timestamps'])
            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)] = name + ' ' + _('Detailed Data')

            current_row_number += 1
            # 1: Stand for blank line  2: Stand for title
            current_row_number += 1 * 6 + real_timestamps_len * 6 + 1 + 2
            table_start_row_number = current_row_number

            has_data = False

            if len(time) > 0:
                has_data = True

            if has_data:

                ws.row_dimensions[current_row_number].height = 60
                current_col_number = 2
                col = format_cell.get_column_letter(current_col_number)
                ws[col + str(current_row_number)].fill = table_fill
                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].border = f_border
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)] = _('Datetime')

                current_col_number += 1
                col = format_cell.get_column_letter(current_col_number)

                ws[col + str(current_row_number)].fill = table_fill
                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)] = report['meter']['energy_category_name'] + \
                    " (" + report['meter']['unit_of_measure'] + ")"
                ws[col + str(current_row_number)].border = f_border

                current_row_number += 1

                for i in range(0, len(time)):
                    current_col_number = 2
                    col = format_cell.get_column_letter(current_col_number)

                    ws[col + str(current_row_number)].font = title_font
                    ws[col + str(current_row_number)].alignment = c_c_alignment
                    ws[col + str(current_row_number)] = time[i]
                    ws[col + str(current_row_number)].border = f_border

                    current_col_number += 1
                    col = format_cell.get_column_letter(current_col_number)

                    ws[col + str(current_row_number)].font = title_font
                    ws[col + str(current_row_number)].alignment = c_c_alignment
                    ws[col + str(current_row_number)] = round(reporting_period_data['values'][i], 2)
                    ws[col + str(current_row_number)].border = f_border

                    current_row_number += 1

                table_end_row_number = current_row_number - 1

                current_col_number = 2
                col = format_cell.get_column_letter(current_col_number)

                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)] = _('Total')
                ws[col + str(current_row_number)].border = f_border

                current_col_number += 1
                col = format_cell.get_column_letter(current_col_number)
                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)] = round(reporting_period_data['total_in_category'], 2)
                ws[col + str(current_row_number)].border = f_border

                # line
                line = LineChart()
                line.title = _('Reporting Period Costs') + ' - ' \
                    + report['meter']['energy_category_name'] + \
                    " (" + report['meter']['unit_of_measure'] + ")"
                labels = Reference(ws, min_col=2, min_row=table_start_row_number + 1, max_row=table_end_row_number)
                line_data = Reference(ws, min_col=3, min_row=table_start_row_number,
                                      max_row=table_end_row_number)
                line.add_data(line_data, titles_from_data=True)
                line.set_categories(labels)
                line_data = line.series[0]
                line_data.marker.symbol = "circle"
                line_data.smooth = True
                line.x_axis.crosses = 'min'
                line.height = 8.25
                line.width = 24
                line.dLbls = DataLabelList()
                line.dLbls.dLblPos = 't'
                line.dLbls.showVal = True
                line.dLbls.showPercent = False
                chart_col = 'B'
                chart_cell = chart_col + str(table_start_draw_flag)
                ws.add_chart(line, chart_cell)

                current_row_number += 2
        else:
            base_period_data = report['base_period']
            reporting_period_data = report['reporting_period']
            base_period_timestamps = base_period_data['timestamps']
            reporting_period_timestamps = reporting_period_data['timestamps']
            real_timestamps_len = timestamps_data_not_equal_0(report['parameters']['timestamps'])
            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)] = name + ' ' + _('Detailed Data')

            current_row_number += 1
            # 1: Stand for blank line  2: Stand for title
            current_row_number += 1 * 6 + real_timestamps_len * 6 + 1 + 2
            table_start_row_number = current_row_number

            has_data = False

            if len(base_period_timestamps[0]) or len(reporting_period_timestamps[0]) > 0:
                has_data = True

            if has_data:
                ws.row_dimensions[current_row_number].height = 60
                current_col_number = 2
                col = format_cell.get_column_letter(current_col_number)
                ws[col + str(current_row_number)].fill = table_fill
                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].border = f_border
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)] = _('Base Period') + " - " + _('Datetime')

                current_col_number += 1
                col = format_cell.get_column_letter(current_col_number)

                ws[col + str(current_row_number)].fill = table_fill
                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)] = _('Base Period') + " - " + \
                    report['meter']['energy_category_name'] + \
                    " (" + report['meter']['unit_of_measure'] + ")"
                ws[col + str(current_row_number)].border = f_border

                current_col_number += 1
                col = format_cell.get_column_letter(current_col_number)

                ws[col + str(current_row_number)].fill = table_fill
                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].border = f_border
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)] = _('Reporting Period') + " - " + _('Datetime')

                current_col_number += 1
                col = format_cell.get_column_letter(current_col_number)
                ws[col + str(current_row_number)].fill = table_fill
                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)] = _('Reporting Period') + " - " \
                    + report['meter']['energy_category_name'] + \
                    " (" + report['meter']['unit_of_measure'] + ")"
                ws[col + str(current_row_number)].border = f_border

                current_row_number += 1

                max_timestamps_len = len(base_period_timestamps) \
                    if len(base_period_timestamps) >= len(reporting_period_timestamps) \
                    else len(reporting_period_timestamps)

                for i in range(0, max_timestamps_len):
                    current_col_number = 2
                    col = format_cell.get_column_letter(current_col_number)
                    ws[col + str(current_row_number)].font = title_font
                    ws[col + str(current_row_number)].alignment = c_c_alignment
                    ws[col + str(current_row_number)] = base_period_timestamps[i] \
                        if i < len(base_period_timestamps) else None
                    ws[col + str(current_row_number)].border = f_border

                    current_col_number += 1
                    col = format_cell.get_column_letter(current_col_number)

                    ws[col + str(current_row_number)].font = title_font
                    ws[col + str(current_row_number)].alignment = c_c_alignment
                    ws[col + str(current_row_number)] = round(base_period_data['values'][i], 2) \
                        if i < len(base_period_data['values']) else None
                    ws[col + str(current_row_number)].border = f_border

                    current_col_number += 1
                    col = format_cell.get_column_letter(current_col_number)

                    ws[col + str(current_row_number)].font = title_font
                    ws[col + str(current_row_number)].alignment = c_c_alignment
                    ws[col + str(current_row_number)] = reporting_period_timestamps[i] \
                        if i < len(reporting_period_timestamps) else None
                    ws[col + str(current_row_number)].border = f_border

                    current_col_number += 1
                    col = format_cell.get_column_letter(current_col_number)

                    ws[col + str(current_row_number)].font = title_font
                    ws[col + str(current_row_number)].alignment = c_c_alignment
                    ws[col + str(current_row_number)] = round(reporting_period_data['values'][i], 2) \
                        if i < len(reporting_period_data['values']) else None
                    ws[col + str(current_row_number)].border = f_border

                    current_row_number += 1

                current_col_number = 2
                col = format_cell.get_column_letter(current_col_number)
                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)] = _('Total')
                ws[col + str(current_row_number)].border = f_border

                current_col_number += 1
                col = format_cell.get_column_letter(current_col_number)
                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)] = round(base_period_data['total_in_category'], 2)
                ws[col + str(current_row_number)].border = f_border

                current_col_number += 1
                col = format_cell.get_column_letter(current_col_number)

                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)] = _('Total')
                ws[col + str(current_row_number)].border = f_border

                current_col_number += 1
                col = format_cell.get_column_letter(current_col_number)
                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)] = round(reporting_period_data['total_in_category'], 2)
                ws[col + str(current_row_number)].border = f_border

                # line
                line = LineChart()
                line.title = _('Base Period Costs') + " / " \
                    + _('Reporting Period Costs') + ' - ' \
                    + report['meter']['energy_category_name'] + \
                    " (" + report['meter']['unit_of_measure'] + ")"
                labels = Reference(ws, min_col=4,
                                   min_row=table_start_row_number + 1,
                                   max_row=table_start_row_number + len(reporting_period_timestamps))
                base_line_data = Reference(ws, min_col=3, min_row=table_start_row_number,
                                           max_row=table_start_row_number + len(reporting_period_timestamps))
                reporting_line_data = Reference(ws, min_col=5,
                                                min_row=table_start_row_number,
                                                max_row=table_start_row_number
                                                + len(reporting_period_timestamps))
                line.add_data(base_line_data, titles_from_data=True)
                line.add_data(reporting_line_data, titles_from_data=True)
                line.set_categories(labels)
                for j in range(len(line.series)):
                    line.series[j].marker.symbol = "circle"
                    line.series[j].smooth = True
                line.x_axis.crosses = 'min'
                line.height = 8.25
                line.width = 24
                line.dLbls = DataLabelList()
                line.dLbls.dLblPos = 't'
                line.dLbls.showVal = True
                line.dLbls.showPercent = False
                chart_col = 'B'
                chart_cell = chart_col + str(table_start_draw_flag)
                ws.add_chart(line, chart_cell)

                current_row_number += 2

    ####################################################################################################################
    # table_start_draw_flag is the starting line number of the last line chart in the report period
    current_sheet_parameters_row_number = table_start_draw_flag + 1 * 6 + 1
    if 'parameters' not in report.keys() or \
            report['parameters'] is None or \
            'names' not in report['parameters'].keys() or \
            report['parameters']['names'] is None or \
            len(report['parameters']['names']) == 0 or \
            'timestamps' not in report['parameters'].keys() or \
            report['parameters']['timestamps'] is None or \
            len(report['parameters']['timestamps']) == 0 or \
            'values' not in report['parameters'].keys() or \
            report['parameters']['values'] is None or \
            len(report['parameters']['values']) == 0 or \
            timestamps_data_all_equal_0(report['parameters']['timestamps']):
        pass
    else:

        ################################################################################################################
        # new worksheet
        ################################################################################################################

        parameters_data = report['parameters']

        parameters_names_len = len(parameters_data['names'])

        file_name = (re.sub(r'[^A-Z]', '', ws.title))+'_'
        parameters_ws = wb.create_sheet(file_name + _('Parameters'))

        parameters_timestamps_data_max_len = \
            get_parameters_timestamps_lists_max_len(list(parameters_data['timestamps']))

        # Row height
        parameters_ws.row_dimensions[1].height = 102
        for i in range(2, 7 + 1):
            parameters_ws.row_dimensions[i].height = 42

        for i in range(8, parameters_timestamps_data_max_len + 10):
            parameters_ws.row_dimensions[i].height = 60

        # Col width
        parameters_ws.column_dimensions['A'].width = 1.5

        parameters_ws.column_dimensions['B'].width = 25.0

        for i in range(3, 12 + parameters_names_len * 3):
            parameters_ws.column_dimensions[format_cell.get_column_letter(i)].width = 15.0

        # Img
        img = Image("excelexporters/myems.png")
        parameters_ws.add_image(img, 'A1')

        # Title
        parameters_ws['B3'].alignment = b_r_alignment
        parameters_ws['B3'] = _('Name') + ':'
        parameters_ws['C3'].border = b_border
        parameters_ws['C3'].alignment = b_c_alignment
        parameters_ws['C3'] = name

        parameters_ws['D3'].alignment = b_r_alignment
        parameters_ws['D3'] = _('Period Type') + ':'
        parameters_ws['E3'].border = b_border
        parameters_ws['E3'].alignment = b_c_alignment
        parameters_ws['E3'] = period_type

        parameters_ws['B4'].alignment = b_r_alignment
        parameters_ws['B4'] = _('Reporting Start Datetime') + ':'
        parameters_ws['C4'].border = b_border
        parameters_ws['C4'].alignment = b_c_alignment
        parameters_ws['C4'] = reporting_start_datetime_local

        parameters_ws['D4'].alignment = b_r_alignment
        parameters_ws['D4'] = _('Reporting End Datetime') + ':'
        parameters_ws['E4'].border = b_border
        parameters_ws['E4'].alignment = b_c_alignment
        parameters_ws['E4'] = reporting_end_datetime_local

        parameters_ws_current_row_number = 6

        parameters_ws['B' + str(parameters_ws_current_row_number)].font = title_font
        parameters_ws['B' + str(parameters_ws_current_row_number)] = name + ' ' + 'Parameters'

        parameters_ws_current_row_number += 1

        parameters_table_start_row_number = parameters_ws_current_row_number

        parameters_ws.row_dimensions[parameters_ws_current_row_number].height = 80

        parameters_ws_current_row_number += 1

        table_current_col_number = 2

        for i in range(0, parameters_names_len):

            if len(parameters_data['timestamps'][i]) == 0:
                continue

            col = format_cell.get_column_letter(table_current_col_number)

            parameters_ws[col + str(parameters_ws_current_row_number - 1)].fill = table_fill
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].border = f_border

            col = format_cell.get_column_letter(table_current_col_number + 1)

            parameters_ws[col + str(parameters_ws_current_row_number - 1)].fill = table_fill
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].border = f_border
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].font = name_font
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].alignment = c_c_alignment
            parameters_ws[col + str(parameters_ws_current_row_number - 1)] = parameters_data['names'][i]

            table_current_row_number = parameters_ws_current_row_number

            for j, value in enumerate(list(parameters_data['timestamps'][i])):
                col = format_cell.get_column_letter(table_current_col_number)

                parameters_ws[col + str(table_current_row_number)].border = f_border
                parameters_ws[col + str(table_current_row_number)].font = title_font
                parameters_ws[col + str(table_current_row_number)].alignment = c_c_alignment
                parameters_ws[col + str(table_current_row_number)] = value

                col = format_cell.get_column_letter(table_current_col_number + 1)

                parameters_ws[col + str(table_current_row_number)].border = f_border
                parameters_ws[col + str(table_current_row_number)].font = title_font
                parameters_ws[col + str(table_current_row_number)].alignment = c_c_alignment
                parameters_ws[col + str(table_current_row_number)] = round(parameters_data['values'][i][j], 2)

                table_current_row_number += 1

            table_current_col_number = table_current_col_number + 3

        ################################################################################################################
        # parameters chart and parameters table
        ################################################################################################################

        ws['B' + str(current_sheet_parameters_row_number)].font = title_font
        ws['B' + str(current_sheet_parameters_row_number)] = name + ' ' + _('Parameters')

        current_sheet_parameters_row_number += 1

        chart_start_row_number = current_sheet_parameters_row_number

        col_index = 0

        for i in range(0, parameters_names_len):

            if len(parameters_data['timestamps'][i]) == 0:
                continue

            line = LineChart()
            data_col = 3 + col_index * 3
            labels_col = 2 + col_index * 3
            col_index += 1
            line.title = _('Parameters') + ' - ' + \
                parameters_ws.cell(row=parameters_table_start_row_number, column=data_col).value
            labels = Reference(parameters_ws, min_col=labels_col, min_row=parameters_table_start_row_number + 1,
                               max_row=(len(parameters_data['timestamps'][i]) + parameters_table_start_row_number))
            line_data = Reference(parameters_ws, min_col=data_col, min_row=parameters_table_start_row_number,
                                  max_row=(len(parameters_data['timestamps'][i]) + parameters_table_start_row_number))
            line.add_data(line_data, titles_from_data=True)
            line.set_categories(labels)
            line_data = line.series[0]
            line_data.marker.symbol = "circle"
            line_data.smooth = True
            line.x_axis.crosses = 'min'
            line.height = 8.25
            line.width = 24
            line.dLbls = DataLabelList()
            line.dLbls.dLblPos = 't'
            line.dLbls.showVal = False
            line.dLbls.showPercent = False
            chart_col = 'B'
            chart_cell = chart_col + str(chart_start_row_number)
            chart_start_row_number += 6
            ws.add_chart(line, chart_cell)

        current_sheet_parameters_row_number = chart_start_row_number

        current_sheet_parameters_row_number += 1

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename


def timestamps_data_all_equal_0(lists):
    for i, value in enumerate(list(lists)):
        if len(value) > 0:
            return False

    return True


def get_parameters_timestamps_lists_max_len(parameters_timestamps_lists):
    max_len = 0
    for i, value in enumerate(list(parameters_timestamps_lists)):
        if len(value) > max_len:
            max_len = len(value)

    return max_len


def timestamps_data_not_equal_0(lists):
    number = 0
    for i, value in enumerate(list(lists)):
        if len(value) > 0:
            number += 1
    return number


def is_base_period_timestamp_exists(base_period_data):
    timestamps = base_period_data['timestamps']

    if len(timestamps) == 0:
        return False

    for timestamp in timestamps:
        if len(timestamp) > 0:
            return True

    return False
