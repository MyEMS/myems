import base64
import gettext
import os
import uuid

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font


########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file from the report data
# Step 3: Encode the excel file to Base64
########################################################################################################################

def export(report, name, reporting_start_datetime_local, reporting_end_datetime_local, period_type, language):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if report is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(report,
                              name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              period_type,
                              language)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        print(str(ex))

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        print(str(ex))
    return base64_message


def generate_excel(report, name, reporting_start_datetime_local, reporting_end_datetime_local, period_type, language):

    locale_path = './i18n/'
    if language == 'zh_CN':
        trans = gettext.translation('myems', locale_path, languages=['zh_CN'])
    elif language == 'de':
        trans = gettext.translation('myems', locale_path, languages=['de'])
    elif language == 'en':
        trans = gettext.translation('myems', locale_path, languages=['en'])
    else:
        trans = gettext.translation('myems', locale_path, languages=['en'])
    trans.install()
    _ = trans.gettext

    wb = Workbook()
    ws = wb.active
    ws.title = "VirtualMeterSaving"
    # Row height
    ws.row_dimensions[1].height = 102
    for i in range(2, 2000 + 1):
        ws.row_dimensions[i].height = 42

    # Col width
    ws.column_dimensions['A'].width = 1.5

    ws.column_dimensions['B'].width = 25.0

    for i in range(ord('C'), ord('L')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Arial', size=15, bold=True)
    title_font = Font(name='Arial', size=15, bold=True)

    table_fill = PatternFill(fill_type='solid', fgColor='90ee90')
    f_border = Border(left=Side(border_style='medium'),
                      right=Side(border_style='medium'),
                      bottom=Side(border_style='medium'),
                      top=Side(border_style='medium')
                      )
    b_border = Border(
        bottom=Side(border_style='medium'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)

    # Img
    img = Image("excelexporters/myems.png")
    ws.add_image(img, 'A1')

    # Title
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = _('Name') + ':'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'] = name

    ws['D3'].alignment = b_r_alignment
    ws['D3'] = _('Period Type') + ':'
    ws['E3'].border = b_border
    ws['E3'].alignment = b_c_alignment
    ws['E3'] = period_type

    ws['B4'].alignment = b_r_alignment
    ws['B4'] = _('Reporting Start Datetime') + ':'
    ws['C4'].border = b_border
    ws['C4'].alignment = b_c_alignment
    ws['C4'] = reporting_start_datetime_local

    ws['D4'].alignment = b_r_alignment
    ws['D4'] = _('Reporting End Datetime') + ':'
    ws['E4'].border = b_border
    ws['E4'].alignment = b_c_alignment
    ws['E4'] = reporting_end_datetime_local

    if "reporting_period" not in report.keys() or \
            "values_saving" not in report['reporting_period'].keys() or \
            len(report['reporting_period']['values_saving']) == 0:
        filename = str(uuid.uuid4()) + '.xlsx'
        wb.save(filename)

        return filename

    ####################################################################################################################
    if "values_saving" not in report['reporting_period'].keys() or \
            len(report['reporting_period']['values_saving']) == 0:
        for i in range(6, 9 + 1):
            ws.row_dimensions[i].height = 0.1
    else:
        ws['B6'].font = title_font
        ws['B6'] = name + _('Reporting Period Saving')

        reporting_period_data = report['reporting_period']

        ws.row_dimensions[7].height = 60
        ws['B7'].fill = table_fill
        ws['B7'].border = f_border

        ws['B8'].font = title_font
        ws['B8'].alignment = c_c_alignment
        ws['B8'] = _('Saving')
        ws['B8'].border = f_border

        ws['B9'].font = title_font
        ws['B9'].alignment = c_c_alignment
        ws['B9'] = _('Increment Rate')
        ws['B9'].border = f_border

        ws['C7'].fill = table_fill
        ws['C7'].font = name_font
        ws['C7'].alignment = c_c_alignment
        ws['C7'] = report['virtual_meter']['energy_category_name'] + " (" \
            + report['virtual_meter']['unit_of_measure'] + ")"
        ws['C7'].border = f_border

        ws['C8'].font = name_font
        ws['C8'].alignment = c_c_alignment
        ws['C8'] = round(reporting_period_data['total_in_category_saving'], 2)
        ws['C8'].border = f_border

        ws['C9'].font = name_font
        ws['C9'].alignment = c_c_alignment
        ws['C9'] = str(round(reporting_period_data['increment_rate_saving'] * 100, 2)) + "%" \
            if reporting_period_data['increment_rate_saving'] is not None else "-"
        ws['C9'].border = f_border

        # TCE
        ws['D7'].fill = table_fill
        ws['D7'].font = name_font
        ws['D7'].alignment = c_c_alignment
        ws['D7'] = _('Ton of Standard Coal') + '(TCE)' + _('Saving')
        ws['D7'].border = f_border

        ws['D8'].font = name_font
        ws['D8'].alignment = c_c_alignment
        ws['D8'] = round(reporting_period_data['total_in_kgce_saving'] / 1000, 2)
        ws['D8'].border = f_border

        ws['D9'].font = name_font
        ws['D9'].alignment = c_c_alignment
        ws['D9'] = str(round(reporting_period_data['increment_rate_saving'] * 100, 2)) + "%" \
            if reporting_period_data['increment_rate_saving'] is not None else "-"
        ws['D9'].border = f_border

        # TCO2E
        ws['E7'].fill = table_fill
        ws['E7'].font = name_font
        ws['E7'].alignment = c_c_alignment
        ws['E7'] = _('Ton of Carbon Dioxide Emissions') + '(TCO2E)' + _('Decreased')
        ws['E7'].border = f_border

        ws['E8'].font = name_font
        ws['E8'].alignment = c_c_alignment
        ws['E8'] = round(reporting_period_data['total_in_kgco2e_saving'] / 1000, 2)
        ws['E8'].border = f_border

        ws['E9'].font = name_font
        ws['E9'].alignment = c_c_alignment
        ws['E9'] = str(round(reporting_period_data['increment_rate_saving'] * 100, 2)) + "%" \
            if reporting_period_data['increment_rate_saving'] is not None else "-"
        ws['E9'].border = f_border

    ####################################################################################################################
    reporting_period_data = report['reporting_period']

    times = reporting_period_data['timestamps']

    if "values_saving" not in reporting_period_data.keys() or len(reporting_period_data['values_saving']) == 0:
        for i in range(11, 43 + 1):
            ws.row_dimensions[i].height = 0.0
    else:
        ws['B11'].font = title_font
        ws['B11'] = name + _('Detailed Data')

        ws.row_dimensions[18].height = 60
        ws['B18'].fill = table_fill
        ws['B18'].font = title_font
        ws['B18'].border = f_border
        ws['B18'].alignment = c_c_alignment
        ws['B18'] = _('Datetime')
        time = times
        has_data = False
        max_row = 0
        if len(time) > 0:
            has_data = True
            max_row = 18 + len(time)

        if has_data:

            end_data_row_number = 19

            for i in range(0, len(time)):
                col = 'B'
                end_data_row_number = 19 + i
                row = str(end_data_row_number)

                ws[col + row].font = title_font
                ws[col + row].alignment = c_c_alignment
                ws[col + row] = time[i]
                ws[col + row].border = f_border

            ws['B' + str(end_data_row_number + 1)].font = title_font
            ws['B' + str(end_data_row_number + 1)].alignment = c_c_alignment
            ws['B' + str(end_data_row_number + 1)] = _('Total')
            ws['B' + str(end_data_row_number + 1)].border = f_border

            ws['C18'].fill = table_fill
            ws['C18'].font = title_font
            ws['C18'].alignment = c_c_alignment
            ws['C18'] = report['virtual_meter']['energy_category_name'] + " (" \
                + report['virtual_meter']['unit_of_measure'] + ")"
            ws['C18'].border = f_border

            time = times
            time_len = len(time)

            for j in range(0, time_len):
                row = str(19 + j)

                ws['C' + row].font = title_font
                ws['C' + row].alignment = c_c_alignment
                ws['C' + row] = round(reporting_period_data['values_saving'][j], 2)
                ws['C' + row].border = f_border

            ws['C' + str(end_data_row_number + 1)].font = title_font
            ws['C' + str(end_data_row_number + 1)].alignment = c_c_alignment
            ws['C' + str(end_data_row_number + 1)] = round(reporting_period_data['total_in_category_saving'], 2)
            ws['C' + str(end_data_row_number + 1)].border = f_border

            line = LineChart()
            labels = Reference(ws, min_col=2, min_row=19, max_row=max_row)
            line_data = Reference(ws, min_col=3, min_row=18, max_row=max_row)
            line.series.append(Series(line_data, title_from_data=True))
            line.set_categories(labels)
            line_data = line.series[0]
            line_data.marker.symbol = "circle"
            line_data.smooth = True
            line.x_axis.crosses = 'min'
            line.title = _('Reporting Period Saving') + ' - ' + report['virtual_meter']['energy_category_name'] + \
                " (" + report['virtual_meter']['unit_of_measure'] + ")"
            line.dLbls = DataLabelList()
            line.dLbls.dLblPos = 't'
            line.dLbls.showVal = True
            line.height = 8.25
            line.width = 24
            ws.add_chart(line, "B12")

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename
