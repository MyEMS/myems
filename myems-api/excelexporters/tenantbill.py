import base64
import datetime
import gettext
import os
import uuid
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font


########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file
# Step 3: Encode the excel file to Base64
########################################################################################################################


def export(report,
           name,
           reporting_start_datetime_local,
           reporting_end_datetime_local,
           period_type,
           language):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if report is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(report,
                              name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              period_type,
                              language)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        print(str(ex))

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        print(str(ex))
    return base64_message


def generate_excel(report,
                   name,
                   reporting_start_datetime_local,
                   reporting_end_datetime_local,
                   period_type,
                   language):

    locale_path = './i18n/'
    if language == 'zh_CN':
        trans = gettext.translation('myems', locale_path, languages=['zh_CN'])
    elif language == 'de':
        trans = gettext.translation('myems', locale_path, languages=['de'])
    elif language == 'en':
        trans = gettext.translation('myems', locale_path, languages=['en'])
    else:
        trans = gettext.translation('myems', locale_path, languages=['en'])
    trans.install()
    _ = trans.gettext

    wb = Workbook()
    ws = wb.active
    ws.title = "TenantBill"

    # Row height
    for i in range(1, 11 + 1):
        ws.row_dimensions[i].height = 0.1
    ws.row_dimensions[12].height = 30.0
    ws.row_dimensions[13].height = 10.0
    ws.merge_cells('B13:I13')
    for i in range(14, 23 + 1):
        ws.row_dimensions[i].height = 0.1
    ws.row_dimensions[24].height = 20.0
    ws.row_dimensions[25].height = 10.0
    ws.merge_cells('B25:I25')
    for i in range(26, 35 + 1):
        ws.row_dimensions[i].height = 0.1
    for i in range(36, 41 + 1):
        ws.row_dimensions[i].height = 20.0
    ws.row_dimensions[42].height = 10.0
    ws.merge_cells('B42:I42')
    for i in range(43, 52 + 1):
        ws.row_dimensions[i].height = 0.1

    # Col width
    ws.column_dimensions['A'].width = 1.5
    for i in range(ord('B'), ord('J')):
        ws.column_dimensions[chr(i)].width = 16
    ws.column_dimensions['J'].width = 1.5

    # merge cell
    ws.merge_cells('C12:H12')

    ws.merge_cells('C24:I24')

    # Font
    notice_font = Font(name='Arial', size=20, bold=True)
    name_font = Font(name='Arial', size=12, bold=True)
    title_font = Font(name='Arial', size=11, bold=True)

    table_fill = PatternFill(fill_type='solid', fgColor='90ee90')
    f_border = Border(left=Side(border_style='medium'),
                      right=Side(border_style='medium'),
                      bottom=Side(border_style='medium'),
                      top=Side(border_style='medium')
                      )

    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_l_alignment = Alignment(vertical='bottom',
                              horizontal='left',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)

    ws['C12'].font = notice_font
    ws['C12'].alignment = c_c_alignment
    ws['C12'] = _('Payment Notice')

    # img
    img = Image("excelexporters/myemslogo.png")
    img.width = 117
    img.height = 117
    ws.add_image(img, 'I12')

    if "tenant" not in report.keys() or \
            report['tenant'] is None or \
            'lease_number' not in report['tenant'].keys() or \
            report['tenant']['lease_number'] is None:
        ws.row_dimensions[24].height = 0.1
    else:
        ws['B24'].font = name_font
        ws['B24'].alignment = b_r_alignment
        ws['B24'] = _('Lease Number') + ':'
        ws['C24'].alignment = b_l_alignment
        ws['C24'].font = name_font
        ws['C24'] = report['tenant']['lease_number']

    if "tenant" not in report.keys() or \
            report['tenant'] is None:
        for i in range(36, 41 + 1):
            ws.row_dimensions[i].height = 0.1
    else:
        report_tenant_data = report['tenant']
        for i in range(36, 41 + 1):
            ws.merge_cells('C{}:D{}'.format(i, i))
            ws['C' + str(i)].alignment = b_l_alignment
            ws['C' + str(i)].font = name_font

        ws['C36'] = report_tenant_data['name']
        ws.merge_cells('E36:I36')

        ws['C37'] = report_tenant_data['rooms']

        ws['C38'] = report_tenant_data['floors']

        ws['C39'] = report_tenant_data['buildings']

        ws['C40'] = report_tenant_data['email']

        ws['C41'] = report_tenant_data['phone']

        for i in range(37, 41 + 1):
            ws.merge_cells('E{}:G{}'.format(i, i))
            ws.merge_cells('H{}:I{}'.format(i, i))
            ws['E' + str(i)].alignment = b_r_alignment
            ws['E' + str(i)].font = name_font
            ws['H' + str(i)].alignment = b_l_alignment
            ws['H' + str(i)].font = name_font

        ws['E37'] = _('Bill Number') + ':'
        ws['E38'] = _('Lease Contract Number') + ':'
        ws['E39'] = _('Bill Date') + ':'
        ws['E40'] = _('Payment Due Date') + ':'
        ws['E41'] = _('Amount Payable') + ':'

        # Simulated data
        ws['H37'] = ''
        ws['H38'] = report_tenant_data['lease_number']
        ws['H39'] = datetime.datetime.strptime(reporting_start_datetime_local, '%Y-%m-%dT%H:%M:%S').strftime('%Y-%m-%d')
        ws['H40'] = datetime.datetime.strptime(reporting_end_datetime_local, '%Y-%m-%dT%H:%M:%S').strftime('%Y-%m-%d')
        ws['H41'] = report['reporting_period']['currency_unit'] + \
            str(round(report['reporting_period']['total_cost']
                      if 'reporting_period' in report.keys()
                         and 'total_cost' in report['reporting_period'].keys()
                         and report['reporting_period']['total_cost'] is not None
                      else 0, 2))

    if 'reporting_period' not in report.keys() \
            or report['reporting_period'] is None:
        pass
    else:
        ws.row_dimensions[53].height = 25.0
        for i in range(ord('B'), ord('J')):
            ws[chr(i) + '53'].fill = table_fill
            ws[chr(i) + '53'].font = title_font
            ws[chr(i) + '53'].alignment = c_c_alignment
            ws[chr(i) + '53'].border = f_border

        ws['B53'] = _('Energy Category')
        ws['C53'] = _('Billing Period Start')
        ws['D53'] = _('Billing Period End')
        ws['E53'] = _('Quantity')
        ws['F53'] = _('Unit')
        ws['G53'] = _('Amount')
        ws['H53'] = _('Tax Rate')
        ws['I53'] = _('VAT Output Tax')

        reporting_period_data = report['reporting_period']
        names = reporting_period_data['names']
        ca_len = len(names) if names is not None else 0

        for i in range(54, 54 + ca_len):
            ws.row_dimensions[i].height = 20.0
            for j in range(ord('B'), ord('J')):
                ws[chr(j) + str(i)].font = title_font
                ws[chr(j) + str(i)].alignment = c_c_alignment
                ws[chr(j) + str(i)].border = f_border

                if chr(j) == 'B':
                    ws[chr(j) + str(i)] = reporting_period_data['names'][i - 54]
                elif chr(j) == 'C':
                    ws[chr(j) + str(i)] = datetime.datetime.strptime(reporting_start_datetime_local,
                                                                     '%Y-%m-%dT%H:%M:%S').strftime('%Y-%m-%d')
                elif chr(j) == 'D':
                    ws[chr(j) + str(i)] = datetime.datetime.strptime(reporting_end_datetime_local,
                                                                     '%Y-%m-%dT%H:%M:%S').strftime('%Y-%m-%d')
                elif chr(j) == 'E':
                    ws[chr(j) + str(i)] = round(reporting_period_data['subtotals_input'][i - 54], 3)
                elif chr(j) == 'F':
                    ws[chr(j) + str(i)] = reporting_period_data['units'][i - 54]
                elif chr(j) == 'G':
                    ws[chr(j) + str(i)] = round(reporting_period_data['subtotals_cost'][i - 54], 2)
                elif chr(j) == 'H':
                    # Simulated data
                    ws[chr(j) + str(i)] = 0
                elif chr(j) == 'I':
                    # Simulated data
                    ws[chr(j) + str(i)] = 0

        ws.row_dimensions[54 + ca_len].height = 10.0
        ws.merge_cells('B{}:H{}'.format((54 + ca_len), (54 + ca_len)))

        current_row_number = 54 + ca_len + 1
        for i in range(current_row_number, current_row_number + 3):
            ws.row_dimensions[i].height = 20.0
            ws['B' + str(i)].alignment = b_r_alignment
            ws['B' + str(i)].font = name_font
            ws['H' + str(i)].alignment = b_l_alignment
            ws['H' + str(i)].font = name_font
            ws.merge_cells('B{}:G{}'.format(i, i))
            ws.merge_cells('H{}:I{}'.format(i, i))

        ws['B' + str(current_row_number)] = _('Subtotal') + ':'
        ws['H' + str(current_row_number)] = report['reporting_period']['currency_unit'] + str(
            round(report['reporting_period']['total_cost']
                  if 'reporting_period' in report.keys()
                     and 'total_cost' in report['reporting_period'].keys()
                     and report['reporting_period']['total_cost'] is not None
                  else 0, 2))

        current_row_number += 1

        # Simulated data
        taxes = Decimal(0.00)

        ws['B' + str(current_row_number)] = _('VAT Output Tax') + ':'
        ws['H' + str(current_row_number)] = report['reporting_period']['currency_unit'] + str(round(taxes, 2))

        current_row_number += 1

        ws['B' + str(current_row_number)] = _('Total Amount Payable') + ':'
        ws['H' + str(current_row_number)] = report['reporting_period']['currency_unit'] + str(
            round(report['reporting_period']['total_cost'] + taxes
                  if 'reporting_period' in report.keys()
                     and 'total_cost' in report['reporting_period'].keys()
                     and report['reporting_period']['total_cost'] is not None
                  else 0 + taxes, 2))

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename
