import base64
import uuid
import os
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.drawing.image import Image
from openpyxl import Workbook


########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excelexporters file
# Step 3: Encode the excelexporters file to Base64
########################################################################################################################

def export(result, space_name, reporting_start_datetime_local, reporting_end_datetime_local):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if result is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(result,
                              space_name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        pass

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        pass
    return base64_message


def generate_excel(report, space_name, reporting_start_datetime_local, reporting_end_datetime_local):

    wb = Workbook()
    ws = wb.active
    ws.title = "StoreBatch"

    # Row height
    ws.row_dimensions[1].height = 102
    for i in range(2, 5 + 1):
        ws.row_dimensions[i].height = 42

    for i in range(6, len(report['stores']) + 15):
        ws.row_dimensions[i].height = 60

    # Col width
    ws.column_dimensions['A'].width = 1.5

    ws.column_dimensions['B'].width = 25.0

    for i in range(ord('C'), ord('L')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Constantia', size=15, bold=True)
    title_font = Font(name='宋体', size=15, bold=True)
    data_font = Font(name='Franklin Gothic Book', size=11)

    table_fill = PatternFill(fill_type='solid', fgColor='1F497D')
    f_border = Border(left=Side(border_style='medium', color='00000000'),
                      right=Side(border_style='medium', color='00000000'),
                      bottom=Side(border_style='medium', color='00000000'),
                      top=Side(border_style='medium', color='00000000')
                      )
    b_border = Border(
        bottom=Side(border_style='medium', color='00000000'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)

    # Img
    img = Image("excelexporters/myems.png")
    img.width = img.width * 0.85
    img.height = img.height * 0.85
    ws.add_image(img, 'B1')

    # Title
    ws.row_dimensions[3].height = 60

    ws['B3'].font = name_font
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = '空间:'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'].font = name_font
    ws['C3'] = space_name

    ws['F3'].font = name_font
    ws['F3'].alignment = b_r_alignment
    ws['F3'] = '日期:'
    ws['G3'].border = b_border
    ws['G3'].alignment = b_c_alignment
    ws['G3'].font = name_font
    ws['G3'] = reporting_start_datetime_local + "~" + reporting_end_datetime_local
    ws.merge_cells("G3:H3")

    # Title
    ws['B6'].border = f_border
    ws['B6'].font = name_font
    ws['B6'].alignment = c_c_alignment
    ws['B6'].fill = table_fill
    ws['B6'] = '名称'

    ws['C6'].border = f_border
    ws['C6'].alignment = c_c_alignment
    ws['C6'].font = name_font
    ws['C6'].fill = table_fill
    ws['C6'] = '空间'

    ca_len = len(report['energycategories'])

    for i in range(0, ca_len):
        col = chr(ord('D') + i)
        ws[col + '6'].fill = table_fill
        ws[col + '6'].font = name_font
        ws[col + '6'].alignment = c_c_alignment
        ws[col + '6'] = report['energycategories'][i]['name'] + \
            " (" + report['energycategories'][i]['unit_of_measure'] + ")"
        ws[col + '6'].border = f_border

    current_row_number = 7
    for i in range(0, len(report['stores'])):

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)].border = f_border
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)] = report['stores'][i]['store_name']

        ws['C' + str(current_row_number)].font = title_font
        ws['C' + str(current_row_number)].border = f_border
        ws['C' + str(current_row_number)].alignment = c_c_alignment
        ws['C' + str(current_row_number)] = report['stores'][i]['space_name']

        ca_len = len(report['stores'][i]['values'])
        for j in range(0, ca_len):
            col = chr(ord('D') + j)
            ws[col + str(current_row_number)].font = data_font
            ws[col + str(current_row_number)].border = f_border
            ws[col + str(current_row_number)].alignment = c_c_alignment
            ws[col + str(current_row_number)] = round(report['stores'][i]['values'][j], 2)

        current_row_number += 1

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename
