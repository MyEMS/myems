"""
Energy Storage Power Station Reporting Parameters Excel Exporter

This module provides functionality to export energy storage power station parameters reporting data to Excel format.
It generates comprehensive reports showing energy storage power station parameters performance
with detailed analysis and visualizations.

Key Features:
- Energy storage power station parameters performance analysis
- Base period vs reporting period comparison
- Formatted Excel output with proper styling
- Multi-language support
- Base64 encoding for file transmission

The exported Excel file includes:
- Energy storage power station parameters performance summary
- Base period comparison data
- Proper formatting and borders
- Logo and header information
"""

import base64
from core.utilities import get_translation
import os
import uuid
import openpyxl.utils.cell as format_cell
from openpyxl import Workbook
from core.utilities import round2


########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file from the report data
# Step 3: Encode the excel file to Base64
########################################################################################################################

def export(report, name, reporting_start_datetime_local, reporting_end_datetime_local, language):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if report is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(report,
                              name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              language)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        print(str(ex))

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        print(str(ex))
    return base64_message


def generate_excel(report,
                   name,
                   reporting_start_datetime_local,
                   reporting_end_datetime_local,
                   language):
    trans = get_translation(language)
    trans.install()
    _ = trans.gettext
    wb = Workbook()
    ws = wb.active
    ws.title = "ESPSReportingParameters"

    ####################################################################################################################
    if 'parameters' not in report.keys() or \
            report['parameters'] is None or \
            'names' not in report['parameters'].keys() or \
            report['parameters']['names'] is None or \
            len(report['parameters']['names']) == 0 or \
            'timestamps' not in report['parameters'].keys() or \
            report['parameters']['timestamps'] is None or \
            len(report['parameters']['timestamps']) == 0 or \
            'values' not in report['parameters'].keys() or \
            report['parameters']['values'] is None or \
            len(report['parameters']['values']) == 0 or \
            timestamps_data_all_equal_0(report['parameters']['timestamps']):
        pass
    else:
        # the first column is for timestamps
        for i, t in enumerate(list(report['parameters']['timestamps'][0])):
            ws['A' + str(i+2)] = t

        for i in range(0, len(report['parameters']['names'])):
            if len(report['parameters']['timestamps'][i]) == 0:
                continue
            column_letter = format_cell.get_column_letter(i + 2)
            # the first row is for name
            current_row_number = 1
            ws[column_letter + str(current_row_number)] = report['parameters']['names'][i]
            # store value from the second row
            current_row_number = 2
            for j, value in enumerate(list(report['parameters']['values'][i])):
                ws[column_letter + str(current_row_number)] = round2(report['parameters']['values'][i][j], 2)
                current_row_number += 1

    ####################################################################################################################
    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename


def timestamps_data_all_equal_0(lists):
    for i, value in enumerate(list(lists)):
        if len(value) > 0:
            return False

    return True


def get_parameters_timestamps_lists_max_len(parameters_timestamps_lists):
    max_len = 0
    for i, value in enumerate(list(parameters_timestamps_lists)):
        if len(value) > max_len:
            max_len = len(value)

    return max_len


def timestamps_data_not_equal_0(lists):
    number = 0
    for i, value in enumerate(list(lists)):
        if len(value) > 0:
            number += 1
    return number
