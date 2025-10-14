import base64
from core.utilities import get_translation
import os
import re
import uuid
import openpyxl.utils.cell as format_cell
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from core.utilities import round2


########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file from the report data
# Step 3: Encode the excel file to Base64
########################################################################################################################
def export(
    result,
    equipment1_name,
    equipment2_name,
    energy_category_name,
    reporting_start_datetime_local,
    reporting_end_datetime_local,
    period_type,
    language,
):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if result is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(
        result,
        equipment1_name,
        equipment2_name,
        energy_category_name,
        reporting_start_datetime_local,
        reporting_end_datetime_local,
        period_type,
        language,
    )
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    binary_file_data = b""
    try:
        with open(filename, "rb") as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        print(str(ex))

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode("utf-8")
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        print(str(ex))
    return base64_message


def generate_excel(
    report,
    equipment1_name,
    equipment2_name,
    energy_category_name,
    reporting_start_datetime_local,
    reporting_end_datetime_local,
    period_type,
    language,
):
    trans = get_translation(language)
    trans.install()
    _ = trans.gettext

    wb = Workbook()
    ws = wb.active
    ws.title = "EquipmentComparison"
    # Row height
    ws.row_dimensions[1].height = 102
    for i in range(2, 2000 + 1):
        ws.row_dimensions[i].height = 42

    # Col width
    ws.column_dimensions["A"].width = 1.5

    ws.column_dimensions["B"].width = 25.0

    for i in range(ord("C"), ord("L")):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name="Arial", size=15, bold=True)
    title_font = Font(name="Arial", size=15, bold=True)

    table_fill = PatternFill(fill_type="solid", fgColor="90ee90")
    f_border = Border(
        left=Side(border_style="medium"),
        right=Side(border_style="medium"),
        bottom=Side(border_style="medium"),
        top=Side(border_style="medium"),
    )
    b_border = Border(
        bottom=Side(border_style="medium"),
    )

    b_c_alignment = Alignment(
        vertical="bottom",
        horizontal="center",
        text_rotation=0,
        wrap_text=True,
        shrink_to_fit=False,
        indent=0,
    )
    c_c_alignment = Alignment(
        vertical="center",
        horizontal="center",
        text_rotation=0,
        wrap_text=True,
        shrink_to_fit=False,
        indent=0,
    )
    b_r_alignment = Alignment(
        vertical="bottom",
        horizontal="right",
        text_rotation=0,
        wrap_text=True,
        shrink_to_fit=False,
        indent=0,
    )

    # Img
    img = Image("excelexporters/myems.png")
    ws.add_image(img, "A1")

    # Title
    ws["B3"].alignment = b_r_alignment
    ws["B3"] = _("Equipment") + "1:"
    ws["C3"].border = b_border
    ws["C3"].alignment = b_c_alignment
    ws["C3"] = equipment1_name

    ws["D3"].alignment = b_r_alignment
    ws["D3"] = _("Equipment") + "2:"
    ws["E3"].border = b_border
    ws["E3"].alignment = b_c_alignment
    ws["E3"] = equipment2_name

    ws["F3"].alignment = b_r_alignment
    ws["F3"] = _("Energy Category") + ":"
    ws["G3"].border = b_border
    ws["G3"].alignment = b_c_alignment
    ws["G3"] = energy_category_name

    ws["B4"].alignment = b_r_alignment
    ws["B4"] = _("Period Type") + ":"
    ws["C4"].border = b_border
    ws["C4"].alignment = b_c_alignment
    ws["C4"] = period_type

    ws["D4"].alignment = b_r_alignment
    ws["D4"] = _("Reporting Start Datetime") + ":"
    ws["E4"].border = b_border
    ws["E4"].alignment = b_c_alignment
    ws["E4"] = reporting_start_datetime_local

    ws["F4"].alignment = b_r_alignment
    ws["F4"] = _("Reporting End Datetime") + ":"
    ws["G4"].border = b_border
    ws["G4"].alignment = b_c_alignment
    ws["G4"] = reporting_end_datetime_local

    if (
        "reporting_period1" not in report.keys()
        or "values" not in report["reporting_period1"].keys()
        or len(report["reporting_period1"]["values"]) == 0
    ):
        filename = str(uuid.uuid4()) + ".xlsx"
        wb.save(filename)

        return filename
    ####################################################################################################################
    # First: Consumption
    # 6: equipment1 title
    # 7: equipment1 table title
    # 8~9 equipment1 table_data
    # 10: equipment2 title
    # 11: equipment2 table title
    # 12~13: equipment2 table_data
    ####################################################################################################################
    if (
        "values" not in report["reporting_period1"].keys()
        or len(report["reporting_period1"]["values"]) == 0
    ):
        for i in range(6, 9 + 1):
            ws.row_dimensions[i].height = 0.1
    else:
        reporting_period_data1 = report["reporting_period1"]

        ws.row_dimensions[7].height = 60
        ws["B7"].font = title_font
        ws["B7"].alignment = c_c_alignment
        ws["B7"] = equipment1_name
        ws["B7"].fill = table_fill
        ws["B7"].border = f_border

        ws["B8"].font = title_font
        ws["B8"].alignment = c_c_alignment
        ws["B8"] = _("Consumption")
        ws["B8"].border = f_border

        ws["C7"].fill = table_fill
        ws["C7"].font = name_font
        ws["C7"].alignment = c_c_alignment
        ws["C7"] = (
            energy_category_name
            + " ("
            + report["energy_category"]["unit_of_measure"]
            + ")"
        )
        ws["C7"].border = f_border

        ws["C8"].font = name_font
        ws["C8"].alignment = c_c_alignment
        ws["C8"] = round2(reporting_period_data1["total_in_category"], 2)
        ws["C8"].border = f_border

    if (
        "values" not in report["reporting_period2"].keys()
        or len(report["reporting_period2"]["values"]) == 0
    ):
        for i in range(11, 14 + 1):
            ws.row_dimensions[i].height = 0.1
    else:
        reporting_period_data2 = report["reporting_period2"]

        ws.row_dimensions[12].height = 60
        ws["B11"].font = title_font
        ws["B11"].alignment = c_c_alignment
        ws["B11"].fill = table_fill
        ws["B11"].border = f_border
        ws["B11"] = equipment2_name

        ws["B12"].font = title_font
        ws["B12"].alignment = c_c_alignment
        ws["B12"] = _("Consumption")
        ws["B12"].border = f_border

        ws["C11"].fill = table_fill
        ws["C11"].font = name_font
        ws["C11"].alignment = c_c_alignment
        ws["C11"] = (
            energy_category_name
            + " ("
            + report["energy_category"]["unit_of_measure"]
            + ")"
        )
        ws["C11"].border = f_border

        ws["C12"].font = name_font
        ws["C12"].alignment = c_c_alignment
        ws["C12"] = round2(reporting_period_data2["total_in_category"], 2)
        ws["C12"].border = f_border

    ####################################################################################################################
    # Second: Detailed Data
    # 15: title
    # 12 ~ 16: chart
    # 18 + 6 * parameterlen + : table title
    # 19 + 6 * parameterlen~18 + 6 * parameterlen + timestamps_len: table_data
    # parameter_len: len(report['parameters1']['names']) + len(report['parameters1']['names'])
    # timestamps_len: reporting_period_data1['timestamps']
    ####################################################################################################################
    times = report["reporting_period1"]["timestamps"]

    if (
        "values" not in report["reporting_period1"].keys()
        or len(report["reporting_period1"]["values"]) == 0
        or "values" not in report["reporting_period2"].keys()
        or len(report["reporting_period2"]["values"]) == 0
    ):
        for i in range(11, 43 + 1):
            ws.row_dimensions[i].height = 0.0
    else:
        reporting_period_data1 = report["reporting_period1"]
        reporting_period_data2 = report["reporting_period2"]
        diff_data = report["diff"]
        parameters_names_len = len(report["parameters1"]["names"])
        parameters_data = report["parameters1"]
        parameters_parameters_datas_len = 0
        for i in range(0, parameters_names_len):
            if len(parameters_data["timestamps"][i]) == 0:
                continue
            parameters_parameters_datas_len += 1
        parameters_names_len = len(report["parameters2"]["names"])
        parameters_data = report["parameters2"]
        for i in range(0, parameters_names_len):
            if len(parameters_data["timestamps"][i]) == 0:
                continue
            parameters_parameters_datas_len += 1
        start_detail_data_row_num = 15 + (parameters_parameters_datas_len + 1 + 1) * 6
        ws["B14"].font = title_font
        ws["B14"] = equipment1_name + " and " + equipment2_name + _("Detailed Data")

        ws.row_dimensions[start_detail_data_row_num].height = 60

        ws["B" + str(start_detail_data_row_num)].fill = table_fill
        ws["B" + str(start_detail_data_row_num)].font = title_font
        ws["B" + str(start_detail_data_row_num)].border = f_border
        ws["B" + str(start_detail_data_row_num)].alignment = c_c_alignment
        ws["B" + str(start_detail_data_row_num)] = _("Datetime")
        time = times
        has_data = False
        max_row = 0
        if len(time) > 0:
            has_data = True
            max_row = start_detail_data_row_num + len(time)

        if has_data:
            for i in range(0, len(time)):
                col = "B"
                row = str(start_detail_data_row_num + 1 + i)
                # col = chr(ord('B') + i)
                ws[col + row].font = title_font
                ws[col + row].alignment = c_c_alignment
                ws[col + row] = time[i]
                ws[col + row].border = f_border

            # table_title
            col = chr(ord(col) + 1)

            ws[col + str(start_detail_data_row_num)].fill = table_fill
            ws[col + str(start_detail_data_row_num)].font = title_font
            ws[col + str(start_detail_data_row_num)].alignment = c_c_alignment
            ws[col + str(start_detail_data_row_num)] = (
                equipment1_name
                + " "
                + energy_category_name
                + " ("
                + report["energy_category"]["unit_of_measure"]
                + ")"
            )
            ws[col + str(start_detail_data_row_num)].border = f_border

            # table_data
            time = times
            time_len = len(time)

            for j in range(0, time_len):
                row = str(start_detail_data_row_num + 1 + j)
                # col = chr(ord('B') + i)
                ws[col + row].font = title_font
                ws[col + row].alignment = c_c_alignment
                ws[col + row] = round2(reporting_period_data1["values"][j], 2)
                ws[col + row].border = f_border

            # table_title
            col = chr(ord(col) + 1)

            ws[col + str(start_detail_data_row_num)].fill = table_fill
            ws[col + str(start_detail_data_row_num)].font = title_font
            ws[col + str(start_detail_data_row_num)].alignment = c_c_alignment
            ws[col + str(start_detail_data_row_num)] = (
                equipment2_name
                + " "
                + energy_category_name
                + " ("
                + report["energy_category"]["unit_of_measure"]
                + ")"
            )
            ws[col + str(start_detail_data_row_num)].border = f_border

            # table_data
            time = times
            time_len = len(time)

            for j in range(0, time_len):
                row = str(start_detail_data_row_num + 1 + j)
                # col = chr(ord('B') + i)
                ws[col + row].font = title_font
                ws[col + row].alignment = c_c_alignment
                ws[col + row] = round2(reporting_period_data2["values"][j], 2)
                ws[col + row].border = f_border

            # table_title
            col = chr(ord(col) + 1)

            ws[col + str(start_detail_data_row_num)].fill = table_fill
            ws[col + str(start_detail_data_row_num)].font = title_font
            ws[col + str(start_detail_data_row_num)].alignment = c_c_alignment
            ws[col + str(start_detail_data_row_num)] = _("Difference")
            ws[col + str(start_detail_data_row_num)].border = f_border

            # table_data
            time = times
            time_len = len(time)

            for j in range(0, time_len):
                row = str(start_detail_data_row_num + 1 + j)
                # col = chr(ord('B') + i)
                ws[col + row].font = title_font
                ws[col + row].alignment = c_c_alignment
                ws[col + row] = round2(diff_data["values"][j], 2)
                ws[col + row].border = f_border
            # line
            # 15~: line
            line = LineChart()
            line.title = _("Reporting Period Consumption")
            labels = Reference(
                ws, min_col=2, min_row=start_detail_data_row_num + 1, max_row=max_row
            )
            line_data = Reference(
                ws,
                min_col=3,
                max_col=2 + 1 + 1,
                min_row=start_detail_data_row_num,
                max_row=max_row,
            )
            line.add_data(line_data, titles_from_data=True)
            line.set_categories(labels)
            for j in range(0, len(line.series)):
                line.series[j].marker.symbol = "auto"
                line.series[j].smooth = True
            line.x_axis.crosses = "min"
            line.height = 8.25
            line.width = 24
            ws.add_chart(line, "B15")

            col = "B"
            row = str(start_detail_data_row_num + 1 + len(time))

            ws[col + row].font = title_font
            ws[col + row].alignment = c_c_alignment
            ws[col + row] = _("Total")
            ws[col + row].border = f_border

            col = chr(ord(col) + 1)
            ws[col + row].font = title_font
            ws[col + row].alignment = c_c_alignment
            ws[col + row] = round2(reporting_period_data1["total_in_category"], 2)
            ws[col + row].border = f_border

            col = chr(ord(col) + 1)
            ws[col + row].font = title_font
            ws[col + row].alignment = c_c_alignment
            ws[col + row] = round2(reporting_period_data2["total_in_category"], 2)
            ws[col + row].border = f_border

            col = chr(ord(col) + 1)
            ws[col + row].font = title_font
            ws[col + row].alignment = c_c_alignment
            ws[col + row] = round2(diff_data["total_in_category"], 2)
            ws[col + row].border = f_border

    ####################################################################################################################
    has_parameters_names_and_timestamps_and_values_data = True
    # 12 is the starting line number of the last line chart in the report period
    current_sheet_parameters_row_number = 10 + (1 + 1) * 6
    if (
        "parameters1" not in report.keys()
        or report["parameters1"] is None
        or "names" not in report["parameters1"].keys()
        or report["parameters1"]["names"] is None
        or len(report["parameters1"]["names"]) == 0
        or "timestamps" not in report["parameters1"].keys()
        or report["parameters1"]["timestamps"] is None
        or len(report["parameters1"]["timestamps"]) == 0
        or "values" not in report["parameters1"].keys()
        or report["parameters1"]["values"] is None
        or len(report["parameters1"]["values"]) == 0
        or timestamps_data_all_equal_0(report["parameters1"]["timestamps"])
    ):
        has_parameters_names_and_timestamps_and_values_data = False

    if (
        "parameters2" not in report.keys()
        or report["parameters2"] is None
        or "names" not in report["parameters2"].keys()
        or report["parameters2"]["names"] is None
        or len(report["parameters2"]["names"]) == 0
        or "timestamps" not in report["parameters2"].keys()
        or report["parameters2"]["timestamps"] is None
        or len(report["parameters2"]["timestamps"]) == 0
        or "values" not in report["parameters2"].keys()
        or report["parameters2"]["values"] is None
        or len(report["parameters2"]["values"]) == 0
        or timestamps_data_all_equal_0(report["parameters2"]["timestamps"])
    ):
        has_parameters_names_and_timestamps_and_values_data = False

    if has_parameters_names_and_timestamps_and_values_data:

        parameters_data1 = report["parameters1"]

        parameters_names_len = len(parameters_data1["names"])

        file_name = (re.sub(r"[^A-Z]", "", ws.title)) + "_"
        parameters_ws = wb.create_sheet(file_name + "Parameters1")

        parameters_timestamps_data_max_len = get_parameters_timestamps_lists_max_len(
            list(parameters_data1["timestamps"])
        )

        # Row height
        parameters_ws.row_dimensions[1].height = 102
        for i in range(2, 7 + 1):
            parameters_ws.row_dimensions[i].height = 42

        for i in range(8, parameters_timestamps_data_max_len + 10):
            parameters_ws.row_dimensions[i].height = 60

        # Col width
        parameters_ws.column_dimensions["A"].width = 1.5

        parameters_ws.column_dimensions["B"].width = 25.0

        for i in range(3, 12 + parameters_names_len * 3):
            parameters_ws.column_dimensions[format_cell.get_column_letter(i)].width = (
                15.0
            )

        # Img
        img = Image("excelexporters/myems.png")
        parameters_ws.add_image(img, "A1")

        # Title
        parameters_ws["B3"].alignment = b_r_alignment
        parameters_ws["B3"] = _("Equipment") + "1:"
        parameters_ws["C3"].border = b_border
        parameters_ws["C3"].alignment = b_c_alignment
        parameters_ws["C3"] = equipment1_name

        parameters_ws["D3"].alignment = b_r_alignment
        parameters_ws["D3"] = _("Energy Category") + ":"
        parameters_ws["E3"].border = b_border
        parameters_ws["E3"].alignment = b_c_alignment
        parameters_ws["E3"] = energy_category_name

        parameters_ws["B4"].alignment = b_r_alignment
        parameters_ws["B4"] = _("Period Type") + ":"
        parameters_ws["C4"].border = b_border
        parameters_ws["C4"].alignment = b_c_alignment
        parameters_ws["C4"] = period_type

        parameters_ws["D4"].alignment = b_r_alignment
        parameters_ws["D4"] = _("Reporting Start Datetime") + ":"
        parameters_ws["E4"].border = b_border
        parameters_ws["E4"].alignment = b_c_alignment
        parameters_ws["E4"] = reporting_start_datetime_local

        parameters_ws["F4"].alignment = b_r_alignment
        parameters_ws["F4"] = _("Reporting End Datetime") + ":"
        parameters_ws["G4"].border = b_border
        parameters_ws["G4"].alignment = b_c_alignment
        parameters_ws["G4"] = reporting_end_datetime_local

        parameters_ws_current_row_number = 6

        parameters_ws["B" + str(parameters_ws_current_row_number)].font = title_font
        parameters_ws["B" + str(parameters_ws_current_row_number)] = (
            equipment1_name + " " + _("Parameters")
        )

        parameters_ws_current_row_number += 1

        parameters_ws.row_dimensions[parameters_ws_current_row_number].height = 80

        parameters_ws_current_row_number += 1

        table_current_col_number = 2

        for i in range(0, parameters_names_len):

            if len(parameters_data1["timestamps"][i]) == 0:
                continue

            col = format_cell.get_column_letter(table_current_col_number)

            parameters_ws[col + str(parameters_ws_current_row_number - 1)].fill = (
                table_fill
            )
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].border = (
                f_border
            )

            col = format_cell.get_column_letter(table_current_col_number + 1)

            parameters_ws[col + str(parameters_ws_current_row_number - 1)].fill = (
                table_fill
            )
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].border = (
                f_border
            )
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].font = (
                name_font
            )
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].alignment = (
                c_c_alignment
            )
            parameters_ws[col + str(parameters_ws_current_row_number - 1)] = (
                parameters_data1["names"][i]
            )

            table_current_row_number = parameters_ws_current_row_number

            for j, value in enumerate(list(parameters_data1["timestamps"][i])):
                col = format_cell.get_column_letter(table_current_col_number)

                parameters_ws[col + str(table_current_row_number)].border = f_border
                parameters_ws[col + str(table_current_row_number)].font = title_font
                parameters_ws[col + str(table_current_row_number)].alignment = (
                    c_c_alignment
                )
                parameters_ws[col + str(table_current_row_number)] = value

                col = format_cell.get_column_letter(table_current_col_number + 1)

                parameters_ws[col + str(table_current_row_number)].border = f_border
                parameters_ws[col + str(table_current_row_number)].font = title_font
                parameters_ws[col + str(table_current_row_number)].alignment = (
                    c_c_alignment
                )
                try:
                    parameters_ws[col + str(table_current_row_number)] = round2(
                        parameters_data1["values"][i][j], 2
                    )
                except Exception as e:
                    print("error 1 in excelexporters\\equipmentcomparison: " + str(e))

                table_current_row_number += 1

            table_current_col_number = table_current_col_number + 3

        parameters_data2 = report["parameters2"]

        parameters_names_len = len(parameters_data2["names"])

        file_name = (re.sub(r"[^A-Z]", "", ws.title)) + "_"
        parameters_ws = wb.create_sheet(file_name + "Parameters2")

        parameters_timestamps_data_max_len = get_parameters_timestamps_lists_max_len(
            list(parameters_data2["timestamps"])
        )

        # Row height
        parameters_ws.row_dimensions[1].height = 102
        for i in range(2, 7 + 1):
            parameters_ws.row_dimensions[i].height = 42

        for i in range(8, parameters_timestamps_data_max_len + 10):
            parameters_ws.row_dimensions[i].height = 60

        # Col width
        parameters_ws.column_dimensions["A"].width = 1.5

        parameters_ws.column_dimensions["B"].width = 25.0

        for i in range(3, 12 + parameters_names_len * 3):
            parameters_ws.column_dimensions[format_cell.get_column_letter(i)].width = (
                15.0
            )

        # Img
        img = Image("excelexporters/myems.png")
        parameters_ws.add_image(img, "A1")

        # Title
        parameters_ws["B3"].alignment = b_r_alignment
        parameters_ws["B3"] = _("Equipment") + "2:"
        parameters_ws["C3"].border = b_border
        parameters_ws["C3"].alignment = b_c_alignment
        parameters_ws["C3"] = equipment2_name

        parameters_ws["D3"].alignment = b_r_alignment
        parameters_ws["D3"] = _("Energy Category") + ":"
        parameters_ws["E3"].border = b_border
        parameters_ws["E3"].alignment = b_c_alignment
        parameters_ws["E3"] = energy_category_name

        parameters_ws["B4"].alignment = b_r_alignment
        parameters_ws["B4"] = _("Period Type") + ":"
        parameters_ws["C4"].border = b_border
        parameters_ws["C4"].alignment = b_c_alignment
        parameters_ws["C4"] = period_type

        parameters_ws["D4"].alignment = b_r_alignment
        parameters_ws["D4"] = _("Reporting Start Datetime") + ":"
        parameters_ws["E4"].border = b_border
        parameters_ws["E4"].alignment = b_c_alignment
        parameters_ws["E4"] = reporting_start_datetime_local

        parameters_ws["F4"].alignment = b_r_alignment
        parameters_ws["F4"] = _("Reporting End Datetime") + ":"
        parameters_ws["G4"].border = b_border
        parameters_ws["G4"].alignment = b_c_alignment
        parameters_ws["G4"] = reporting_end_datetime_local

        parameters_ws_current_row_number = 6

        parameters_ws["B" + str(parameters_ws_current_row_number)].font = title_font
        parameters_ws["B" + str(parameters_ws_current_row_number)] = (
            equipment2_name + " " + _("Parameters")
        )

        parameters_ws_current_row_number += 1

        parameters_table_start_row_number = parameters_ws_current_row_number

        parameters_ws.row_dimensions[parameters_ws_current_row_number].height = 80

        parameters_ws_current_row_number += 1

        table_current_col_number = 2

        for i in range(0, parameters_names_len):

            if len(parameters_data2["timestamps"][i]) == 0:
                continue

            col = format_cell.get_column_letter(table_current_col_number)

            parameters_ws[col + str(parameters_ws_current_row_number - 1)].fill = (
                table_fill
            )
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].border = (
                f_border
            )

            col = format_cell.get_column_letter(table_current_col_number + 1)

            parameters_ws[col + str(parameters_ws_current_row_number - 1)].fill = (
                table_fill
            )
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].border = (
                f_border
            )
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].font = (
                name_font
            )
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].alignment = (
                c_c_alignment
            )
            parameters_ws[col + str(parameters_ws_current_row_number - 1)] = (
                parameters_data2["names"][i]
            )

            table_current_row_number = parameters_ws_current_row_number

            for j, value in enumerate(list(parameters_data2["timestamps"][i])):
                col = format_cell.get_column_letter(table_current_col_number)

                parameters_ws[col + str(table_current_row_number)].border = f_border
                parameters_ws[col + str(table_current_row_number)].font = title_font
                parameters_ws[col + str(table_current_row_number)].alignment = (
                    c_c_alignment
                )
                parameters_ws[col + str(table_current_row_number)] = value

                col = format_cell.get_column_letter(table_current_col_number + 1)

                parameters_ws[col + str(table_current_row_number)].border = f_border
                parameters_ws[col + str(table_current_row_number)].font = title_font
                parameters_ws[col + str(table_current_row_number)].alignment = (
                    c_c_alignment
                )
                try:
                    parameters_ws[col + str(table_current_row_number)] = round2(
                        parameters_data2["values"][i][j], 2
                    )
                except Exception as e:
                    print("error 1 in excelexporters\\equipmentcomparison: " + str(e))

                table_current_row_number += 1

            table_current_col_number = table_current_col_number + 3

        ################################################################################################################
        # parameters chart and parameters table
        ################################################################################################################

        ws["B" + str(current_sheet_parameters_row_number)].font = title_font
        ws["B" + str(current_sheet_parameters_row_number)] = (
            equipment1_name + " " + _("Parameters")
        )
        parameters_names_len = len(report["parameters1"]["names"])
        parameters_ws = wb[file_name + "Parameters1"]

        current_sheet_parameters_row_number += 1

        chart_start_row_number = current_sheet_parameters_row_number

        col_index = 0

        for i in range(0, parameters_names_len):

            if len(parameters_data1["timestamps"][i]) == 0:
                continue

            line = LineChart()
            data_col = 3 + col_index * 3
            labels_col = 2 + col_index * 3
            col_index += 1
            line.title = (
                _("Parameters")
                + " - "
                + parameters_ws.cell(
                    row=parameters_table_start_row_number, column=data_col
                ).value
            )
            labels = Reference(
                parameters_ws,
                min_col=labels_col,
                min_row=parameters_table_start_row_number + 1,
                max_row=(
                    len(parameters_data1["timestamps"][i])
                    + parameters_table_start_row_number
                ),
            )
            line_data = Reference(
                parameters_ws,
                min_col=data_col,
                min_row=parameters_table_start_row_number,
                max_row=(
                    len(parameters_data1["timestamps"][i])
                    + parameters_table_start_row_number
                ),
            )
            line.add_data(line_data, titles_from_data=True)
            line.set_categories(labels)
            line_data = line.series[0]
            line_data.marker.symbol = "auto"
            line_data.smooth = True
            line.x_axis.crosses = "min"
            line.height = 8.25
            line.width = 24
            chart_col = "B"
            chart_cell = chart_col + str(chart_start_row_number)
            chart_start_row_number += 6
            ws.add_chart(line, chart_cell)

        current_sheet_parameters_row_number = chart_start_row_number

        current_sheet_parameters_row_number += 1

        parameters_ws = wb[file_name + "Parameters2"]
        ws["B" + str(current_sheet_parameters_row_number)].font = title_font
        ws["B" + str(current_sheet_parameters_row_number)] = (
            equipment2_name + " " + _("Parameters")
        )

        current_sheet_parameters_row_number += 1

        chart_start_row_number = current_sheet_parameters_row_number

        col_index = 0

        parameters_names_len = len(report["parameters2"]["names"])

        for i in range(0, parameters_names_len):

            if len(parameters_data2["timestamps"][i]) == 0:
                continue

            line = LineChart()
            data_col = 3 + col_index * 3
            labels_col = 2 + col_index * 3
            col_index += 1
            line.title = (
                _("Parameters")
                + " - "
                + parameters_ws.cell(
                    row=parameters_table_start_row_number, column=data_col
                ).value
            )
            labels = Reference(
                parameters_ws,
                min_col=labels_col,
                min_row=parameters_table_start_row_number + 1,
                max_row=(
                    len(parameters_data2["timestamps"][i])
                    + parameters_table_start_row_number
                ),
            )
            line_data = Reference(
                parameters_ws,
                min_col=data_col,
                min_row=parameters_table_start_row_number,
                max_row=(
                    len(parameters_data2["timestamps"][i])
                    + parameters_table_start_row_number
                ),
            )
            line.add_data(line_data, titles_from_data=True)
            line.set_categories(labels)
            line_data = line.series[0]
            line_data.marker.symbol = "auto"
            line_data.smooth = True
            line.x_axis.crosses = "min"
            line.height = 8.25
            line.width = 24
            chart_col = "B"
            chart_cell = chart_col + str(chart_start_row_number)
            chart_start_row_number += 6
            ws.add_chart(line, chart_cell)

        current_sheet_parameters_row_number = chart_start_row_number

        current_sheet_parameters_row_number += 1

    filename = str(uuid.uuid4()) + ".xlsx"
    wb.save(filename)

    return filename


def timestamps_data_all_equal_0(lists):
    for i, value in enumerate(list(lists)):
        if len(value) > 0:
            return False

    return True


def get_parameters_timestamps_lists_max_len(parameters_timestamps_lists):
    max_len = 0
    for i, value in enumerate(list(parameters_timestamps_lists)):
        if len(value) > max_len:
            max_len = len(value)

    return max_len


def timestamps_data_not_equal_0(lists):
    number = 0
    for i, value in enumerate(list(lists)):
        if len(value) > 0:
            number += 1
    return number
