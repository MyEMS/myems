"""
Meter Tracking Excel Exporter

This module provides functionality to export meter tracking data to Excel format.
It generates comprehensive reports showing meter tracking information
with detailed data and proper formatting.

Key Features:
- Meter tracking data export
- Formatted Excel output with proper styling
- Multi-language support
- Base64 encoding for file transmission

The exported Excel file includes:
- Meter tracking information
- Proper formatting and borders
- Logo and header information
"""

import base64
from core.utilities import get_translation
import os
import uuid
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
from core.utilities import round2


########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excelexporters file
# Step 3: Encode the excelexporters file to Base64
########################################################################################################################
def export(result, space_name, energy_category_name, reporting_start_datetime_local, reporting_end_datetime_local,
           language):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if result is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(result,
                              space_name,
                              energy_category_name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              language)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        print(str(ex))

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        print(str(ex))
    return base64_message


def generate_excel(report, space_name, energy_category_name, reporting_start_datetime_local,
                   reporting_end_datetime_local, language):

    trans = get_translation(language)
    trans.install()
    _ = trans.gettext

    wb = Workbook()
    ws = wb.active
    ws.title = "MeterTracking"

    # Column width
    for i in range(ord('A'), ord('I')):
        ws.column_dimensions[chr(i)].width = 25.0

    # Head image
    ws.row_dimensions[1].height = 105
    img = Image("excelexporters/myems.png")
    ws.add_image(img, 'A1')

    # Query Parameters
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    ws['A3'].alignment = b_r_alignment
    ws['A3'] = _('Space') + ':'
    ws['B3'] = space_name
    ws['A4'].alignment = b_r_alignment
    ws['A4'] = _('Start Datetime') + ':'
    ws['B4'] = reporting_start_datetime_local
    ws['A5'].alignment = b_r_alignment
    ws['A5'] = _('End Datetime') + ':'
    ws['B5'] = reporting_end_datetime_local
    ws['A6'].alignment = b_r_alignment
    ws['A6'] = _('Start Integrity Rate') + ':'
    ws['B6'] = (str(round2(report['start_integrity_rate'] * Decimal(100.0), 2)) + '%') \
        if report['start_integrity_rate'] is not None else None
    ws['A7'].alignment = b_r_alignment
    ws['A7'] = _('End Integrity Rate') + ':'
    ws['B7'] = (str(round2(report['end_integrity_rate'] * Decimal(100.0), 2)) + '%') \
        if report['end_integrity_rate'] is not None else None
    ws['A8'].alignment = b_r_alignment
    ws['A8'] = _('Full Integrity Rate') + ':'
    ws['B8'] = (str(round2(report['full_integrity_rate'] * Decimal(100.0), 2)) + '%') \
        if report['full_integrity_rate'] is not None else None
    ws['A9'].alignment = b_r_alignment
    ws['A9'] = _('Energy Category') + ':'
    ws['B9'] = energy_category_name if energy_category_name is not None else _('All')

    # Title
    title_font = Font(size=12, bold=True)
    ws['A10'].font = title_font
    ws['A10'] = _('ID')
    ws['B10'].font = title_font
    ws['B10'] = _('Name')
    ws['C10'].font = title_font
    ws['C10'] = _('Space')
    ws['D10'].font = title_font
    ws['D10'] = _('Cost Center')
    ws['E10'].font = title_font
    ws['E10'] = _('Energy Category')
    ws['F10'].font = title_font
    ws['F10'] = _('Description')
    ws['G10'].font = title_font
    ws['G10'] = _('Start Value')
    ws['H10'].font = title_font
    ws['H10'] = _('End Value')
    ws['I10'].font = title_font
    ws['I10'] = _('Difference Value')

    current_row_number = 11
    for i in range(0, len(report['meters'])):
        ws['A' + str(current_row_number)] = report['meters'][i]['id']
        ws['B' + str(current_row_number)] = report['meters'][i]['meter_name']
        ws['C' + str(current_row_number)] = report['meters'][i]['space_name']
        ws['D' + str(current_row_number)] = report['meters'][i]['cost_center_name']
        ws['E' + str(current_row_number)] = report['meters'][i]['energy_category_name']
        ws['F' + str(current_row_number)] = report['meters'][i]['description']
        ws['G' + str(current_row_number)] = report['meters'][i]['start_value']
        ws['H' + str(current_row_number)] = report['meters'][i]['end_value']
        ws['I' + str(current_row_number)] = report['meters'][i]['difference_value']
        current_row_number += 1

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename
