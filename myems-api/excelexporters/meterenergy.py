import base64
import gettext
import os
import re
import uuid

import openpyxl.utils.cell as format_cell
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font


########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file from the report data
# Step 3: Encode the excel file to Base64
########################################################################################################################

def export(report, name, reporting_start_datetime_local, reporting_end_datetime_local, base_period_start_datetime,
           base_period_end_datetime, period_type, language):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if report is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(report,
                              name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              base_period_start_datetime,
                              base_period_end_datetime,
                              period_type,
                              language)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        print(str(ex))

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        print(str(ex))
    return base64_message


def generate_excel(report, name, reporting_start_datetime_local, reporting_end_datetime_local,
                   base_period_start_datetime, base_period_end_datetime, period_type, language):
    locale_path = './i18n/'
    if language == 'zh_CN':
        trans = gettext.translation('myems', locale_path, languages=['zh_CN'])
    elif language == 'de':
        trans = gettext.translation('myems', locale_path, languages=['de'])
    elif language == 'en':
        trans = gettext.translation('myems', locale_path, languages=['en'])
    else:
        trans = gettext.translation('myems', locale_path, languages=['en'])
    trans.install()
    _ = trans.gettext

    wb = Workbook()
    ws = wb.active
    ws.title = "MeterEnergy"

    # Row height
    ws.row_dimensions[1].height = 102
    for i in range(2, 2000 + 1):
        ws.row_dimensions[i].height = 42

    # Col width
    ws.column_dimensions['A'].width = 1.5

    ws.column_dimensions['B'].width = 25.0

    for i in range(ord('C'), ord('L')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Arial', size=15, bold=True)
    title_font = Font(name='Arial', size=15, bold=True)

    table_fill = PatternFill(fill_type='solid', fgColor='90ee90')
    f_border = Border(left=Side(border_style='medium'),
                      right=Side(border_style='medium'),
                      bottom=Side(border_style='medium'),
                      top=Side(border_style='medium')
                      )
    b_border = Border(
        bottom=Side(border_style='medium'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)

    # Img
    img = Image("excelexporters/myems.png")
    ws.add_image(img, 'A1')

    # Title
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = _('Name') + ':'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'] = name

    ws['D3'].alignment = b_r_alignment
    ws['D3'] = _('Period Type') + ':'
    ws['E3'].border = b_border
    ws['E3'].alignment = b_c_alignment
    ws['E3'] = period_type

    ws['B4'].alignment = b_r_alignment
    ws['B4'] = _('Reporting Start Datetime') + ':'
    ws['C4'].border = b_border
    ws['C4'].alignment = b_c_alignment
    ws['C4'] = reporting_start_datetime_local

    ws['D4'].alignment = b_r_alignment
    ws['D4'] = _('Reporting End Datetime') + ':'
    ws['E4'].border = b_border
    ws['E4'].alignment = b_c_alignment
    ws['E4'] = reporting_end_datetime_local

    is_base_period_timestamp_exists_flag = is_base_period_timestamp_exists(report['base_period'])

    if is_base_period_timestamp_exists_flag:
        ws['B5'].alignment = b_r_alignment
        ws['B5'] = _('Base Period Start Datetime') + ':'
        ws['C5'].border = b_border
        ws['C5'].alignment = b_c_alignment
        ws['C5'] = base_period_start_datetime

        ws['D5'].alignment = b_r_alignment
        ws['D5'] = _('Base Period End Datetime') + ':'
        ws['E5'].border = b_border
        ws['E5'].alignment = b_c_alignment
        ws['E5'] = base_period_end_datetime

    if "reporting_period" not in report.keys() or \
            "values" not in report['reporting_period'].keys() or len(report['reporting_period']['values']) == 0:
        filename = str(uuid.uuid4()) + '.xlsx'
        wb.save(filename)

        return filename
    ####################################################################################################################
    # First: Consumption
    # 6: title
    # 7: table title
    # 8~9 table_data
    ####################################################################################################################
    if "values" not in report['reporting_period'].keys() or len(report['reporting_period']['values']) == 0:
        for i in range(6, 9 + 1):
            ws.row_dimensions[i].height = 0.1
    else:
        reporting_period_data = report['reporting_period']

        ws.row_dimensions[7].height = 60

        ws['B7'].font = title_font
        ws['B7'].alignment = c_c_alignment
        ws['B7'].fill = table_fill
        ws['B7'].border = f_border
        ws['B7'] = name

        ws['B8'].font = title_font
        ws['B8'].alignment = c_c_alignment
        ws['B8'] = _('Consumption')
        ws['B8'].border = f_border

        ws['B9'].font = title_font
        ws['B9'].alignment = c_c_alignment
        ws['B9'] = _('Increment Rate')
        ws['B9'].border = f_border

        ws['C7'].fill = table_fill
        ws['C7'].font = name_font
        ws['C7'].alignment = c_c_alignment
        ws['C7'] = report['meter']['energy_category_name'] + " (" + report['meter']['unit_of_measure'] + ")"
        ws['C7'].border = f_border

        ws['C8'].font = name_font
        ws['C8'].alignment = c_c_alignment
        ws['C8'] = round(reporting_period_data['total_in_category'], 2)
        ws['C8'].border = f_border

        ws['C9'].font = name_font
        ws['C9'].alignment = c_c_alignment
        ws['C9'] = str(round(reporting_period_data['increment_rate'] * 100, 2)) + "%" \
            if reporting_period_data['increment_rate'] is not None else "-"
        ws['C9'].border = f_border

        # TCE
        ws['D7'].fill = table_fill
        ws['D7'].font = name_font
        ws['D7'].alignment = c_c_alignment
        ws['D7'] = _('Ton of Standard Coal') + '(TCE)'
        ws['D7'].border = f_border

        ws['D8'].font = name_font
        ws['D8'].alignment = c_c_alignment
        ws['D8'] = round(reporting_period_data['total_in_kgce'] / 1000, 2)
        ws['D8'].border = f_border

        ws['D9'].font = name_font
        ws['D9'].alignment = c_c_alignment
        ws['D9'] = str(round(reporting_period_data['increment_rate'] * 100, 2)) + "%" \
            if reporting_period_data['increment_rate'] is not None else "-"
        ws['D9'].border = f_border

        # TCO2E
        ws['E7'].fill = table_fill
        ws['E7'].font = name_font
        ws['E7'].alignment = c_c_alignment
        ws['E7'] = _('Ton of Carbon Dioxide Emissions') + '(TCO2E)'
        ws['E7'].border = f_border

        ws['E8'].font = name_font
        ws['E8'].alignment = c_c_alignment
        ws['E8'] = round(reporting_period_data['total_in_kgco2e'] / 1000, 2)
        ws['E8'].border = f_border

        ws['E9'].font = name_font
        ws['E9'].alignment = c_c_alignment
        ws['E9'] = str(round(reporting_period_data['increment_rate'] * 100, 2)) + "%" \
            if reporting_period_data['increment_rate'] is not None else "-"
        ws['E9'].border = f_border

    ####################################################################################################################
    # Second: Detailed Data
    # 11: title
    # 12 ~ 16: chart
    # 18: table title
    # 19~43: table_data
    ####################################################################################################################
    if "values" not in report['reporting_period'].keys() or len(report['reporting_period']['values']) == 0:
        for i in range(11, 43 + 1):
            ws.row_dimensions[i].height = 0.0
    else:
        reporting_period_data = report['reporting_period']
        parameters_names_len = len(report['parameters']['names'])
        parameters_parameters_datas_len = parameters_names_len
        start_detail_data_row_num = 13 + (parameters_parameters_datas_len + 1) * 6
        ws['B11'].font = title_font
        ws['B11'] = name + _('Detailed Data')

        ws.row_dimensions[start_detail_data_row_num].height = 60

        if not is_base_period_timestamp_exists_flag:
            ws['B' + str(start_detail_data_row_num)].fill = table_fill
            ws['B' + str(start_detail_data_row_num)].font = title_font
            ws['B' + str(start_detail_data_row_num)].border = f_border
            ws['B' + str(start_detail_data_row_num)].alignment = c_c_alignment
            ws['B' + str(start_detail_data_row_num)] = _('Datetime')
        else:
            ws['B' + str(start_detail_data_row_num)].fill = table_fill
            ws['B' + str(start_detail_data_row_num)].font = title_font
            ws['B' + str(start_detail_data_row_num)].border = f_border
            ws['B' + str(start_detail_data_row_num)].alignment = c_c_alignment
            ws['B' + str(start_detail_data_row_num)] = _('Base Period') + ' - ' + _('Datetime')

            ws['D' + str(start_detail_data_row_num)].fill = table_fill
            ws['D' + str(start_detail_data_row_num)].font = title_font
            ws['D' + str(start_detail_data_row_num)].border = f_border
            ws['D' + str(start_detail_data_row_num)].alignment = c_c_alignment
            ws['D' + str(start_detail_data_row_num)] = _('Reporting Period') + ' - ' + _('Datetime')

        has_data = False
        max_row = 0
        if len(reporting_period_data['timestamps']) > 0:
            has_data = True
            max_row = start_detail_data_row_num + len(reporting_period_data['timestamps'])

        if has_data:
            if not is_base_period_timestamp_exists_flag:
                for i in range(0, len(reporting_period_data['timestamps'])):
                    col = 'B'
                    row = str(start_detail_data_row_num + 1 + i)
                    # col = chr(ord('B') + i)
                    ws[col + row].font = title_font
                    ws[col + row].alignment = c_c_alignment
                    ws[col + row] = reporting_period_data['timestamps'][i]
                    ws[col + row].border = f_border
            else:
                for i in range(0, len(reporting_period_data['timestamps'])):
                    col = 'B'
                    row = str(start_detail_data_row_num + 1 + i)
                    # col = chr(ord('B') + i)
                    ws[col + row].font = title_font
                    ws[col + row].alignment = c_c_alignment
                    ws[col + row] = report['base_period']['timestamps'][i] \
                        if i < len(report['base_period']['timestamps']) else None
                    ws[col + row].border = f_border

                    col = 'D'
                    row = str(start_detail_data_row_num + 1 + i)
                    # col = chr(ord('B') + i)
                    ws[col + row].font = title_font
                    ws[col + row].alignment = c_c_alignment
                    ws[col + row] = reporting_period_data['timestamps'][i]
                    ws[col + row].border = f_border

            if not is_base_period_timestamp_exists_flag:
                ws['C' + str(start_detail_data_row_num)].fill = table_fill
                ws['C' + str(start_detail_data_row_num)].font = title_font
                ws['C' + str(start_detail_data_row_num)].alignment = c_c_alignment
                ws['C' + str(start_detail_data_row_num)] = report['meter']['energy_category_name'] + \
                    " (" + report['meter']['unit_of_measure'] + ")"
                ws['C' + str(start_detail_data_row_num)].border = f_border
            else:
                ws['C' + str(start_detail_data_row_num)].fill = table_fill
                ws['C' + str(start_detail_data_row_num)].font = title_font
                ws['C' + str(start_detail_data_row_num)].alignment = c_c_alignment
                ws['C' + str(start_detail_data_row_num)] = _('Base Period') + ' - ' + \
                    report['meter']['energy_category_name'] + " (" + report['meter']['unit_of_measure'] + ")"
                ws['C' + str(start_detail_data_row_num)].border = f_border

                ws['E' + str(start_detail_data_row_num)].fill = table_fill
                ws['E' + str(start_detail_data_row_num)].font = title_font
                ws['E' + str(start_detail_data_row_num)].alignment = c_c_alignment
                ws['E' + str(start_detail_data_row_num)] = _('Reporting Period') + ' - ' + \
                    report['meter']['energy_category_name'] + " (" + report['meter']['unit_of_measure'] + ")"
                ws['E' + str(start_detail_data_row_num)].border = f_border

            # 13 data
            if not is_base_period_timestamp_exists_flag:
                for j in range(0, len(reporting_period_data['timestamps'])):
                    row = str(start_detail_data_row_num + 1 + j)
                    ws['C' + row].font = title_font
                    ws['C' + row].alignment = c_c_alignment
                    ws['C' + row] = round(reporting_period_data['values'][j], 2)
                    ws['C' + row].border = f_border
                # line
                # 13~: line
                line = LineChart()
                line.title = _('Reporting Period Consumption') + ' - ' + report['meter']['energy_category_name'] + \
                    " (" + report['meter']['unit_of_measure'] + ")"
                labels = Reference(ws, min_col=2, min_row=start_detail_data_row_num + 1, max_row=max_row)
                line_data = Reference(ws, min_col=3, min_row=start_detail_data_row_num, max_row=max_row)
                line.add_data(line_data, titles_from_data=True)
                line.set_categories(labels)
                line_data = line.series[0]
                line_data.marker.symbol = "circle"
                line_data.smooth = True
                line.x_axis.crosses = 'min'
                line.height = 8.25
                line.width = 24
                line.dLbls = DataLabelList()
                line.dLbls.dLblPos = 't'
                line.dLbls.showVal = True
                line.dLbls.showPercent = False
                ws.add_chart(line, "B12")
            else:
                for j in range(0, len(reporting_period_data['timestamps'])):
                    row = str(start_detail_data_row_num + 1 + j)
                    ws['C' + row].font = title_font
                    ws['C' + row].alignment = c_c_alignment
                    ws['C' + row] = round(report['base_period']['values'][j], 2) \
                        if j < len(report['base_period']['values']) else None
                    ws['C' + row].border = f_border

                    ws['E' + row].font = title_font
                    ws['E' + row].alignment = c_c_alignment
                    ws['E' + row] = round(reporting_period_data['values'][j], 2)
                    ws['E' + row].border = f_border
                # line
                # 13~: line
                line = LineChart()
                line.title = _('Reporting Period Consumption') + ' - ' + report['meter']['energy_category_name'] + \
                    " (" + report['meter']['unit_of_measure'] + ")"
                labels = Reference(ws, min_col=4, min_row=start_detail_data_row_num + 1, max_row=max_row)
                base_line_data = Reference(ws, min_col=3, min_row=start_detail_data_row_num, max_row=max_row)
                reporting_data = Reference(ws, min_col=5, min_row=start_detail_data_row_num, max_row=max_row)
                line.add_data(base_line_data, titles_from_data=True)
                line.add_data(reporting_data, titles_from_data=True)
                line.set_categories(labels)
                line_data = line.series[0]
                line_data.marker.symbol = "circle"
                line_data.smooth = True
                line.x_axis.crosses = 'min'
                line.height = 8.25
                line.width = 24
                line.dLbls = DataLabelList()
                line.dLbls.dLblPos = 't'
                line.dLbls.showVal = True
                line.dLbls.showPercent = False
                ws.add_chart(line, "B12")

            row = str(start_detail_data_row_num + 1 + len(reporting_period_data['timestamps']))

            if not is_base_period_timestamp_exists_flag:
                ws['B' + row].font = title_font
                ws['B' + row].alignment = c_c_alignment
                ws['B' + row] = _('Total')
                ws['B' + row].border = f_border

                ws['C' + row].font = title_font
                ws['C' + row].alignment = c_c_alignment
                ws['C' + row] = round(reporting_period_data['total_in_category'], 2)
                ws['C' + row].border = f_border
            else:
                ws['B' + row].font = title_font
                ws['B' + row].alignment = c_c_alignment
                ws['B' + row] = _('Total')
                ws['B' + row].border = f_border

                ws['C' + row].font = title_font
                ws['C' + row].alignment = c_c_alignment
                ws['C' + row] = round(report['base_period']['total_in_category'], 2)
                ws['C' + row].border = f_border

                ws['D' + row].font = title_font
                ws['D' + row].alignment = c_c_alignment
                ws['D' + row] = _('Total')
                ws['D' + row].border = f_border

                ws['E' + row].font = title_font
                ws['E' + row].alignment = c_c_alignment
                ws['E' + row] = round(reporting_period_data['total_in_category'], 2)
                ws['E' + row].border = f_border

    ##########################################
    # 12 is the starting line number of the last line chart in the report period
    current_sheet_parameters_row_number = 12 + 1 * 6
    if 'parameters' not in report.keys() or \
            report['parameters'] is None or \
            'names' not in report['parameters'].keys() or \
            report['parameters']['names'] is None or \
            len(report['parameters']['names']) == 0 or \
            'timestamps' not in report['parameters'].keys() or \
            report['parameters']['timestamps'] is None or \
            len(report['parameters']['timestamps']) == 0 or \
            'values' not in report['parameters'].keys() or \
            report['parameters']['values'] is None or \
            len(report['parameters']['values']) == 0 or \
            timestamps_data_all_equal_0(report['parameters']['timestamps']):
        pass
    else:
        ################################################################################################################
        # new worksheet
        ################################################################################################################

        parameters_data = report['parameters']

        parameters_names_len = len(parameters_data['names'])

        file_name = (re.sub(r'[^A-Z]', '', ws.title)) + '_'
        parameters_ws = wb.create_sheet(file_name + _('Parameters'))

        parameters_timestamps_data_max_len = \
            get_parameters_timestamps_lists_max_len(list(parameters_data['timestamps']))

        # Row height
        parameters_ws.row_dimensions[1].height = 102
        for i in range(2, 7 + 1):
            parameters_ws.row_dimensions[i].height = 42

        for i in range(8, parameters_timestamps_data_max_len + 10):
            parameters_ws.row_dimensions[i].height = 60

        # Col width
        parameters_ws.column_dimensions['A'].width = 1.5

        parameters_ws.column_dimensions['B'].width = 25.0

        for i in range(3, 12 + parameters_names_len * 3):
            parameters_ws.column_dimensions[format_cell.get_column_letter(i)].width = 15.0

        # Img
        img = Image("excelexporters/myems.png")
        parameters_ws.add_image(img, 'A1')

        # Title
        parameters_ws['B3'].alignment = b_r_alignment
        parameters_ws['B3'] = _('Name') + ':'
        parameters_ws['C3'].border = b_border
        parameters_ws['C3'].alignment = b_c_alignment
        parameters_ws['C3'] = name

        parameters_ws['D3'].alignment = b_r_alignment
        parameters_ws['D3'] = _('Period Type') + ':'
        parameters_ws['E3'].border = b_border
        parameters_ws['E3'].alignment = b_c_alignment
        parameters_ws['E3'] = period_type

        parameters_ws['B4'].alignment = b_r_alignment
        parameters_ws['B4'] = _('Reporting Start Datetime') + ':'
        parameters_ws['C4'].border = b_border
        parameters_ws['C4'].alignment = b_c_alignment
        parameters_ws['C4'] = reporting_start_datetime_local

        parameters_ws['D4'].alignment = b_r_alignment
        parameters_ws['D4'] = _('Reporting End Datetime') + ':'
        parameters_ws['E4'].border = b_border
        parameters_ws['E4'].alignment = b_c_alignment
        parameters_ws['E4'] = reporting_end_datetime_local

        parameters_ws_current_row_number = 6

        parameters_ws['B' + str(parameters_ws_current_row_number)].font = title_font
        parameters_ws['B' + str(parameters_ws_current_row_number)] = name + ' ' + _('Parameters')

        parameters_ws_current_row_number += 1

        parameters_table_start_row_number = parameters_ws_current_row_number

        parameters_ws.row_dimensions[parameters_ws_current_row_number].height = 80

        parameters_ws_current_row_number += 1

        table_current_col_number = 2

        for i in range(0, parameters_names_len):
            col = format_cell.get_column_letter(table_current_col_number)

            parameters_ws[col + str(parameters_ws_current_row_number - 1)].fill = table_fill
            col = format_cell.get_column_letter(table_current_col_number + 1)

            parameters_ws[col + str(parameters_ws_current_row_number - 1)].fill = table_fill
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].font = name_font
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].alignment = c_c_alignment
            parameters_ws[col + str(parameters_ws_current_row_number - 1)] = parameters_data['names'][i]

            table_current_col_number = table_current_col_number + 3

        ################################################################################################################
        # Optimized insert parameter data
        ################################################################################################################

        timestamps_list = parameters_data['timestamps']
        values_list = parameters_data['values']

        timestamps_data_temp_save_start_row = parameters_timestamps_data_max_len + 10

        values_data_temp_save_start_row = parameters_timestamps_data_max_len * 2 + 10 + 1

        parameters_ws["A" + str(timestamps_data_temp_save_start_row)] = ""
        for i in range(parameters_timestamps_data_max_len):
            temp_list = []
            for j in range(len(timestamps_list)):
                try:
                    temp_list.append(timestamps_list[j][i])
                except IndexError:
                    temp_list.append("")
            parameters_ws.append(temp_list)

        parameters_ws["A" + str(values_data_temp_save_start_row)] = ""
        for i in range(parameters_timestamps_data_max_len):
            temp_list = []
            for j in range(len(values_list)):
                try:
                    temp_list.append(values_list[j][i])
                except IndexError:
                    temp_list.append("")
            parameters_ws.append(temp_list)

        parameter_current_col_number = 1

        for i in range(len(timestamps_list)):
            col = format_cell.get_column_letter(parameter_current_col_number)
            parameters_ws.move_range(
                "{}{}:{}{}".format(col, timestamps_data_temp_save_start_row + 1, col,
                                   timestamps_data_temp_save_start_row + parameters_timestamps_data_max_len),
                (- timestamps_data_temp_save_start_row + (parameters_ws_current_row_number - 1)), (i * 2) + 1)
            parameters_ws.move_range(
                "{}{}:{}{}".format(col, values_data_temp_save_start_row + 1, col,
                                   values_data_temp_save_start_row + parameters_timestamps_data_max_len),
                (- values_data_temp_save_start_row + (parameters_ws_current_row_number - 1)), (i * 2) + 2)

            parameter_current_col_number += 1

        ################################################################################################################
        # parameters chart and parameters table
        ################################################################################################################

        ws['B' + str(current_sheet_parameters_row_number)].font = title_font
        ws['B' + str(current_sheet_parameters_row_number)] = name + ' ' + _('Parameters')

        current_sheet_parameters_row_number += 1

        chart_start_row_number = current_sheet_parameters_row_number

        col_index = 0

        for i in range(0, parameters_names_len):
            line = LineChart()
            data_col = 3 + col_index * 3
            labels_col = 2 + col_index * 3
            col_index += 1
            line.title = _('Parameters') + ' - ' + \
                parameters_ws.cell(row=parameters_table_start_row_number, column=data_col).value
            labels = Reference(parameters_ws, min_col=labels_col, min_row=parameters_table_start_row_number + 1,
                               max_row=(len(parameters_data['timestamps'][i]) + parameters_table_start_row_number))
            line_data = Reference(parameters_ws, min_col=data_col, min_row=parameters_table_start_row_number,
                                  max_row=(len(parameters_data['timestamps'][i]) + parameters_table_start_row_number))
            line.add_data(line_data, titles_from_data=True)
            line.set_categories(labels)
            line_data = line.series[0]
            line_data.marker.symbol = "circle"
            line_data.smooth = True
            line.x_axis.crosses = 'min'
            line.height = 8.25
            line.width = 24
            line.dLbls = DataLabelList()
            line.dLbls.dLblPos = 't'
            line.dLbls.showVal = False
            line.dLbls.showPercent = False
            chart_col = 'B'
            chart_cell = chart_col + str(chart_start_row_number)
            chart_start_row_number += 6
            ws.add_chart(line, chart_cell)

        current_sheet_parameters_row_number = chart_start_row_number

        current_sheet_parameters_row_number += 1

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename


def timestamps_data_all_equal_0(lists):
    for i, value in enumerate(list(lists)):
        if len(value) > 0:
            return False

    return True


def get_parameters_timestamps_lists_max_len(parameters_timestamps_lists):
    max_len = 0
    for i, value in enumerate(list(parameters_timestamps_lists)):
        if len(value) > max_len:
            max_len = len(value)

    return max_len


def timestamps_data_not_equal_0(lists):
    number = 0
    for i, value in enumerate(list(lists)):
        if len(value) > 0:
            number += 1
    return number


def is_base_period_timestamp_exists(base_period_data):
    timestamps = base_period_data['timestamps']

    if len(timestamps) == 0:
        return False

    for timestamp in timestamps:
        if len(timestamp) > 0:
            return True

    return False
