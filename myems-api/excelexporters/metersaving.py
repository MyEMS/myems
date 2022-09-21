import base64
import uuid
import os
import re
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl.chart.label import DataLabelList
import openpyxl.utils.cell as format_cell
import gettext


########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file from the report data
# Step 3: Encode the excel file to Base64
########################################################################################################################

def export(report, name, reporting_start_datetime_local, reporting_end_datetime_local, period_type, language):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if report is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(report,
                              name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              period_type,
                              language)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        pass

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        pass
    return base64_message


def generate_excel(report, name, reporting_start_datetime_local, reporting_end_datetime_local, period_type, language):

    locale_path = './i18n/'
    if language == 'zh_CN':
        trans = gettext.translation('myems', locale_path, languages=['zh_CN'])
    elif language == 'de':
        trans = gettext.translation('myems', locale_path, languages=['de'])
    elif language == 'en':
        trans = gettext.translation('myems', locale_path, languages=['en'])
    else:
        trans = gettext.translation('myems', locale_path, languages=['en'])
    trans.install()
    _ = trans.gettext

    wb = Workbook()
    ws = wb.active
    ws.title = "MeterSaving"

    # Row height
    ws.row_dimensions[1].height = 102
    for i in range(2, 2000 + 1):
        ws.row_dimensions[i].height = 42

    # Col width
    ws.column_dimensions['A'].width = 1.5

    ws.column_dimensions['B'].width = 25.0

    for i in range(ord('C'), ord('L')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Arial', size=15, bold=True)
    title_font = Font(name='Arial', size=15, bold=True)

    table_fill = PatternFill(fill_type='solid', fgColor='1F497D')
    f_border = Border(left=Side(border_style='medium', color='00000000'),
                      right=Side(border_style='medium', color='00000000'),
                      bottom=Side(border_style='medium', color='00000000'),
                      top=Side(border_style='medium', color='00000000')
                      )
    b_border = Border(
        bottom=Side(border_style='medium', color='00000000'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)

    # Img
    img = Image("excelexporters/myems.png")
    ws.add_image(img, 'A1')

    # Title
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = _('Name') + ':'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'] = name

    ws['D3'].alignment = b_r_alignment
    ws['D3'] = _('Period') + ':'
    ws['E3'].border = b_border
    ws['E3'].alignment = b_c_alignment
    ws['E3'] = period_type

    ws['B4'].alignment = b_r_alignment
    ws['B4'] = _('Reporting Start Datetime') + ':'
    ws['C4'].border = b_border
    ws['C4'].alignment = b_c_alignment
    ws['C4'] = reporting_start_datetime_local

    ws['D4'].alignment = b_r_alignment
    ws['D4'] = _('Reporting End Datetime') + ':'
    ws['E4'].border = b_border
    ws['E4'].alignment = b_c_alignment
    ws['E4'] = reporting_end_datetime_local

    if "reporting_period" not in report.keys() or \
            "values_saving" not in report['reporting_period'].keys() or \
            len(report['reporting_period']['values_saving']) == 0:
        filename = str(uuid.uuid4()) + '.xlsx'
        wb.save(filename)

        return filename
    ####################################################################################################################
    # First: Saving
    # 6: title
    # 7: table title
    # 8~9 table_data
    ####################################################################################################################
    if "values_saving" not in report['reporting_period'].keys() or \
            len(report['reporting_period']['values_saving']) == 0:
        for i in range(6, 9 + 1):
            ws.row_dimensions[i].height = 0.1
    else:
        ws['B6'].font = title_font
        ws['B6'] = name + ' ' + _('Saving')

        reporting_period_data = report['reporting_period']

        category = report['meter']['energy_category_name']
        ca_len = len(category)

        ws.row_dimensions[7].height = 60

        ws['B7'].fill = table_fill
        ws['B7'].border = f_border

        ws['B8'].font = title_font
        ws['B8'].alignment = c_c_alignment
        ws['B8'] = _('Saving')
        ws['B8'].border = f_border

        ws['B9'].font = title_font
        ws['B9'].alignment = c_c_alignment
        ws['B9'] = _('Increment Rate')
        ws['B9'].border = f_border

        col = ''

        for i in range(0, ca_len):
            col = chr(ord('C') + i)
            row = '7'
            cell = col + row
            ws[col + '7'].fill = table_fill
            ws[col + '7'].font = name_font
            ws[col + '7'].alignment = c_c_alignment
            ws[col + '7'] = report['meter']['energy_category_name'] + " (" + report['meter']['unit_of_measure'] + ")"
            ws[col + '7'].border = f_border

            ws[col + '8'].font = name_font
            ws[col + '8'].alignment = c_c_alignment
            ws[col + '8'] = round(reporting_period_data['total_in_category_saving'], 2)
            ws[col + '8'].border = f_border

            ws[col + '9'].font = name_font
            ws[col + '9'].alignment = c_c_alignment
            ws[col + '9'] = str(round(reporting_period_data['increment_rate_saving'] * 100, 2)) + "%" \
                if reporting_period_data['increment_rate_saving'] is not None else "-"
            ws[col + '9'].border = f_border

        # TCE TCO2E
        end_col = col
        # TCE
        tce_col = chr(ord(end_col) + 1)
        ws[tce_col + '7'].fill = table_fill
        ws[tce_col + '7'].font = name_font
        ws[tce_col + '7'].alignment = c_c_alignment
        ws[tce_col + '7'] = _('Ton of Standard Coal') + '(TCE)'
        ws[tce_col + '7'].border = f_border

        ws[tce_col + '8'].font = name_font
        ws[tce_col + '8'].alignment = c_c_alignment
        ws[tce_col + '8'] = round(reporting_period_data['total_in_kgce_saving'] / 1000, 2)
        ws[tce_col + '8'].border = f_border

        ws[tce_col + '9'].font = name_font
        ws[tce_col + '9'].alignment = c_c_alignment
        ws[tce_col + '9'] = str(round(reporting_period_data['increment_rate_saving'] * 100, 2)) + "%" \
            if reporting_period_data['increment_rate_saving'] is not None else "-"
        ws[tce_col + '9'].border = f_border

        # TCO2E
        tco2e_col = chr(ord(end_col) + 2)
        ws[tco2e_col + '7'].fill = table_fill
        ws[tco2e_col + '7'].font = name_font
        ws[tco2e_col + '7'].alignment = c_c_alignment
        ws[tco2e_col + '7'] = _('Ton of Carbon Dioxide Emissions') + '(TCO2E)' + _('Decreased')
        ws[tco2e_col + '7'].border = f_border

        ws[tco2e_col + '8'].font = name_font
        ws[tco2e_col + '8'].alignment = c_c_alignment
        ws[tco2e_col + '8'] = round(reporting_period_data['total_in_kgco2e_saving'] / 1000, 2)
        ws[tco2e_col + '8'].border = f_border

        ws[tco2e_col + '9'].font = name_font
        ws[tco2e_col + '9'].alignment = c_c_alignment
        ws[tco2e_col + '9'] = str(round(reporting_period_data['increment_rate_saving'] * 100, 2)) + "%" \
            if reporting_period_data['increment_rate_saving'] is not None else "-"
        ws[tco2e_col + '9'].border = f_border

    ####################################################################################################################
    # Second: Detailed Data
    # 11: title
    # 12 ~ 16: chart
    # 18: table title
    # 19~43: table_data
    ####################################################################################################################
    reporting_period_data = report['reporting_period']
    times = reporting_period_data['timestamps']

    if "values_saving" not in report['reporting_period'].keys() or \
            len(report['reporting_period']['values_saving']) == 0:
        for i in range(11, 43 + 1):
            ws.row_dimensions[i].height = 0.0
    else:
        reporting_period_data = report['reporting_period']
        category = report['meter']['energy_category_name']
        ca_len = len(category)
        parameters_names_len = len(report['parameters']['names'])
        parameters_data = report['parameters']
        parameters_parameters_datas_len = parameters_names_len
        start_detail_data_row_num = 13 + (parameters_parameters_datas_len + ca_len) * 6
        ws['B11'].font = title_font
        ws['B11'] = name + _('Detailed Data')

        ws.row_dimensions[start_detail_data_row_num].height = 60

        ws['B' + str(start_detail_data_row_num)].fill = table_fill
        ws['B' + str(start_detail_data_row_num)].font = title_font
        ws['B' + str(start_detail_data_row_num)].border = f_border
        ws['B' + str(start_detail_data_row_num)].alignment = c_c_alignment
        ws['B' + str(start_detail_data_row_num)] = _('Datetime')
        time = times
        has_data = False
        max_row = 0
        if len(time) > 0:
            has_data = True
            max_row = start_detail_data_row_num + len(time)

        if has_data:
            for i in range(0, len(time)):
                col = 'B'
                row = str(start_detail_data_row_num + 1 + i)
                # col = chr(ord('B') + i)
                ws[col + row].font = title_font
                ws[col + row].alignment = c_c_alignment
                ws[col + row] = time[i]
                ws[col + row].border = f_border

            for i in range(0, ca_len):
                # 12 title
                col = chr(ord('C') + i)

                ws[col + str(start_detail_data_row_num)].fill = table_fill
                ws[col + str(start_detail_data_row_num)].font = title_font
                ws[col + str(start_detail_data_row_num)].alignment = c_c_alignment
                ws[col + str(start_detail_data_row_num)] = report['meter']['energy_category_name'] + \
                    " (" + report['meter']['unit_of_measure'] + ")"
                ws[col + str(start_detail_data_row_num)].border = f_border

                # 13 data
                time = times
                time_len = len(time)

                for j in range(0, time_len):
                    row = str(start_detail_data_row_num + 1 + j)
                    # col = chr(ord('B') + i)
                    ws[col + row].font = title_font
                    ws[col + row].alignment = c_c_alignment
                    ws[col + row] = round(reporting_period_data['values_saving'][j], 2)
                    ws[col + row].border = f_border
            # line
            # 13~: line
            line = LineChart()
            line.title = _('Reporting Period Saving') + ' - ' + report['meter']['energy_category_name'] + \
                " (" + report['meter']['unit_of_measure'] + ")"
            labels = Reference(ws, min_col=2, min_row=start_detail_data_row_num + 1, max_row=max_row)
            bar_data = Reference(ws, min_col=3, min_row=start_detail_data_row_num, max_row=max_row)
            line.add_data(bar_data, titles_from_data=True)
            line.set_categories(labels)
            line_data = line.series[0]
            line_data.marker.symbol = "circle"
            line_data.smooth = True
            line.x_axis.crosses = 'min'
            line.height = 8.25
            line.width = 24
            line.dLbls = DataLabelList()
            line.dLbls = DataLabelList()
            line.dLbls.dLblPos = 't'
            line.dLbls.showVal = True
            line.dLbls.showPercent = False
            ws.add_chart(line, "B12")

            col = 'B'
            row = str(start_detail_data_row_num + 1 + len(time))

            ws[col + row].font = title_font
            ws[col + row].alignment = c_c_alignment
            ws[col + row] = _('Total')
            ws[col + row].border = f_border

            for i in range(0, ca_len):
                col = chr(ord(col) + 1)
                ws[col + row].font = title_font
                ws[col + row].alignment = c_c_alignment
                ws[col + row] = round(reporting_period_data['total_in_category_saving'], 2)
                ws[col + row].border = f_border

    ##########################################
    # 12 is the starting line number of the last line chart in the report period
    category = report['meter']['energy_category_name']
    time_len = len(reporting_period_data['timestamps'])
    ca_len = len(category)
    current_sheet_parameters_row_number = 12 + ca_len * 6
    if 'parameters' not in report.keys() or \
            report['parameters'] is None or \
            'names' not in report['parameters'].keys() or \
            report['parameters']['names'] is None or \
            len(report['parameters']['names']) == 0 or \
            'timestamps' not in report['parameters'].keys() or \
            report['parameters']['timestamps'] is None or \
            len(report['parameters']['timestamps']) == 0 or \
            'values' not in report['parameters'].keys() or \
            report['parameters']['values'] is None or \
            len(report['parameters']['values']) == 0 or \
            timestamps_data_all_equal_0(report['parameters']['timestamps']):
        pass
    else:

        ################################################################################################################
        # new worksheet
        ################################################################################################################

        parameters_data = report['parameters']

        parameters_names_len = len(parameters_data['names'])

        file_name = (re.sub(r'[^A-Z]', '', ws.title))+'_'
        parameters_ws = wb.create_sheet(file_name + _('Parameters'))

        parameters_timestamps_data_max_len = \
            get_parameters_timestamps_lists_max_len(list(parameters_data['timestamps']))

        # Row height
        parameters_ws.row_dimensions[1].height = 102
        for i in range(2, 7 + 1):
            parameters_ws.row_dimensions[i].height = 42

        for i in range(8, parameters_timestamps_data_max_len + 10):
            parameters_ws.row_dimensions[i].height = 60

        # Col width
        parameters_ws.column_dimensions['A'].width = 1.5

        parameters_ws.column_dimensions['B'].width = 25.0

        for i in range(3, 12 + parameters_names_len * 3):
            parameters_ws.column_dimensions[format_cell.get_column_letter(i)].width = 15.0

        # Img
        img = Image("excelexporters/myems.png")
        parameters_ws.add_image(img, 'A1')

        # Title
        parameters_ws['B3'].alignment = b_r_alignment
        parameters_ws['B3'] = _('Name') + ':'
        parameters_ws['C3'].border = b_border
        parameters_ws['C3'].alignment = b_c_alignment
        parameters_ws['C3'] = name

        parameters_ws['D3'].alignment = b_r_alignment
        parameters_ws['D3'] = _('Period') + ':'
        parameters_ws['E3'].border = b_border
        parameters_ws['E3'].alignment = b_c_alignment
        parameters_ws['E3'] = period_type

        parameters_ws['B4'].alignment = b_r_alignment
        parameters_ws['B4'] = _('Reporting Start Datetime') + ':'
        parameters_ws['C4'].border = b_border
        parameters_ws['C4'].alignment = b_c_alignment
        parameters_ws['C4'] = reporting_start_datetime_local

        parameters_ws['D4'].alignment = b_r_alignment
        parameters_ws['D4'] = _('Reporting End Datetime') + ':'
        parameters_ws['E4'].border = b_border
        parameters_ws['E4'].alignment = b_c_alignment
        parameters_ws['E4'] = reporting_end_datetime_local

        parameters_ws_current_row_number = 6

        parameters_ws['B' + str(parameters_ws_current_row_number)].font = title_font
        parameters_ws['B' + str(parameters_ws_current_row_number)] = name + ' ' + _('Parameters')

        parameters_ws_current_row_number += 1

        parameters_table_start_row_number = parameters_ws_current_row_number

        parameters_ws.row_dimensions[parameters_ws_current_row_number].height = 80

        parameters_ws_current_row_number += 1

        table_current_col_number = 2

        for i in range(0, parameters_names_len):
            col = format_cell.get_column_letter(table_current_col_number)

            parameters_ws[col + str(parameters_ws_current_row_number - 1)].fill = table_fill
            col = format_cell.get_column_letter(table_current_col_number + 1)

            parameters_ws[col + str(parameters_ws_current_row_number - 1)].fill = table_fill
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].font = name_font
            parameters_ws[col + str(parameters_ws_current_row_number - 1)].alignment = c_c_alignment
            parameters_ws[col + str(parameters_ws_current_row_number - 1)] = parameters_data['names'][i]

            table_current_col_number = table_current_col_number + 3

        ################################################################################################################
        # Optimized insert parameter data
        ################################################################################################################

        timestamps_list = parameters_data['timestamps']
        values_list = parameters_data['values']

        timestamps_data_temp_save_start_row = parameters_timestamps_data_max_len + 10

        values_data_temp_save_start_row = parameters_timestamps_data_max_len * 2 + 10 + 1

        parameters_ws["A" + str(timestamps_data_temp_save_start_row)] = ""
        for i in range(parameters_timestamps_data_max_len):
            temp_list = []
            for j in range(len(timestamps_list)):
                try:
                    temp_list.append(timestamps_list[j][i])
                except IndexError:
                    temp_list.append("")
            parameters_ws.append(temp_list)

        parameters_ws["A" + str(values_data_temp_save_start_row)] = ""
        for i in range(parameters_timestamps_data_max_len):
            temp_list = []
            for j in range(len(values_list)):
                try:
                    temp_list.append(values_list[j][i])
                except IndexError:
                    temp_list.append("")
            parameters_ws.append(temp_list)

        parameter_current_col_number = 1

        for i in range(len(timestamps_list)):
            col = format_cell.get_column_letter(parameter_current_col_number)
            parameters_ws.move_range(
                "{}{}:{}{}".format(col, timestamps_data_temp_save_start_row + 1, col,
                                   timestamps_data_temp_save_start_row + parameters_timestamps_data_max_len),
                (- timestamps_data_temp_save_start_row + (parameters_ws_current_row_number - 1)), (i * 2) + 1)
            parameters_ws.move_range(
                "{}{}:{}{}".format(col, values_data_temp_save_start_row + 1, col,
                                   values_data_temp_save_start_row + parameters_timestamps_data_max_len),
                (- values_data_temp_save_start_row + (parameters_ws_current_row_number - 1)), (i * 2) + 2)

            parameter_current_col_number += 1

        ################################################################################################################
        # parameters chart and parameters table
        ################################################################################################################

        ws['B' + str(current_sheet_parameters_row_number)].font = title_font
        ws['B' + str(current_sheet_parameters_row_number)] = name + ' ' + _('Parameters')

        current_sheet_parameters_row_number += 1

        chart_start_row_number = current_sheet_parameters_row_number

        col_index = 0

        for i in range(0, parameters_names_len):
            line = LineChart()
            data_col = 3 + col_index * 3
            labels_col = 2 + col_index * 3
            col_index += 1
            line.title = _('Parameters') + ' - ' + \
                parameters_ws.cell(row=parameters_table_start_row_number, column=data_col).value
            labels = Reference(parameters_ws, min_col=labels_col, min_row=parameters_table_start_row_number + 1,
                               max_row=(len(parameters_data['timestamps'][i]) + parameters_table_start_row_number))
            line_data = Reference(parameters_ws, min_col=data_col, min_row=parameters_table_start_row_number,
                                  max_row=(len(parameters_data['timestamps'][i]) + parameters_table_start_row_number))
            line.add_data(line_data, titles_from_data=True)
            line.set_categories(labels)
            line_data = line.series[0]
            line_data.marker.symbol = "circle"
            line_data.smooth = True
            line.x_axis.crosses = 'min'
            line.height = 8.25
            line.width = 24
            line.dLbls = DataLabelList()
            line.dLbls.dLblPos = 't'
            line.dLbls.showVal = False
            line.dLbls.showPercent = False
            chart_col = 'B'
            chart_cell = chart_col + str(chart_start_row_number)
            chart_start_row_number += 6
            ws.add_chart(line, chart_cell)

        current_sheet_parameters_row_number = chart_start_row_number

        current_sheet_parameters_row_number += 1

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename


def timestamps_data_all_equal_0(lists):
    for i, value in enumerate(list(lists)):
        if len(value) > 0:
            return False

    return True


def get_parameters_timestamps_lists_max_len(parameters_timestamps_lists):
    max_len = 0
    for i, value in enumerate(list(parameters_timestamps_lists)):
        if len(value) > max_len:
            max_len = len(value)

    return max_len


def timestamps_data_not_equal_0(lists):
    number = 0
    for i, value in enumerate(list(lists)):
        if len(value) > 0:
            number += 1
    return number