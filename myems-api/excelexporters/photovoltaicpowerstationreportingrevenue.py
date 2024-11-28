import base64
from core.utilities import get_translation
import os
import re
import uuid
import openpyxl.utils.cell as format_cell
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from core.utilities import round2


########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file from the report data
# Step 3: Encode the excel file to Base64
########################################################################################################################

def export(report, name, reporting_start_datetime_local, reporting_end_datetime_local, period_type, language):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if report is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(report,
                              name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              period_type,
                              language)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        print(str(ex))

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        print(str(ex))
    return base64_message


def generate_excel(report,
                   name,
                   reporting_start_datetime_local,
                   reporting_end_datetime_local,
                   period_type,
                   language):
    trans = get_translation(language)
    trans.install()
    _ = trans.gettext
    wb = Workbook()
    ws = wb.active
    ws.title = "PhotovoltaicPowerStationReportingRevenue"

    # Row height
    ws.row_dimensions[1].height = 102

    for i in range(2, 2000 + 1):
        ws.row_dimensions[i].height = 42

    # Col width
    ws.column_dimensions['A'].width = 1.5

    ws.column_dimensions['B'].width = 25.0

    for i in range(ord('C'), ord('Z')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Arial', size=15, bold=True)
    title_font = Font(name='Arial', size=15, bold=True)

    table_fill = PatternFill(fill_type='solid', fgColor='90ee90')
    f_border = Border(left=Side(border_style='medium'),
                      right=Side(border_style='medium'),
                      bottom=Side(border_style='medium'),
                      top=Side(border_style='medium')
                      )
    b_border = Border(bottom=Side(border_style='medium'), )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)

    # Img
    img = Image("excelexporters/myems.png")
    ws.add_image(img, 'A1')

    # Title
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = _('Name') + ':'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'] = name

    ws['D3'].alignment = b_r_alignment
    ws['D3'] = _('Period Type') + ':'
    ws['E3'].border = b_border
    ws['E3'].alignment = b_c_alignment
    ws['E3'] = period_type

    ws['B4'].alignment = b_r_alignment
    ws['B4'] = _('Reporting Start Datetime') + ':'
    ws['C4'].border = b_border
    ws['C4'].alignment = b_c_alignment
    ws['C4'] = reporting_start_datetime_local

    ws['D4'].alignment = b_r_alignment
    ws['D4'] = _('Reporting End Datetime') + ':'
    ws['E4'].border = b_border
    ws['E4'].alignment = b_c_alignment
    ws['E4'] = reporting_end_datetime_local

    if "reporting_period" not in report.keys() or \
            "names" not in report['reporting_period'].keys() or len(report['reporting_period']['names']) == 0:
        filename = str(uuid.uuid4()) + '.xlsx'
        wb.save(filename)

        return filename

    ####################################################################################################################
    # Time-Of-Use Data
    # 6: title
    # 7: table title
    # 8~12 table_data
    # Total: 7 rows
    ####################################################################################################################
    reporting_period_data = report['reporting_period']
    print(reporting_period_data)
    if "toppeaks" not in reporting_period_data.keys() or \
            reporting_period_data['toppeaks'] is None or \
            len(reporting_period_data['toppeaks']) == 0:
        for i in range(6, 12 + 1):
            ws.row_dimensions[i].height = 0.1
    else:
        ws['B7'].font = title_font
        ws['B7'] = name

        ws.row_dimensions[7].height = 60
        ws['B7'].fill = table_fill
        ws['B7'].font = name_font
        ws['B7'].alignment = c_c_alignment
        ws['B7'].border = f_border

        ws['B8'].font = title_font
        ws['B8'].alignment = c_c_alignment
        ws['B8'] = _('TopPeak')
        ws['B8'].border = f_border

        ws['B9'].font = title_font
        ws['B9'].alignment = c_c_alignment
        ws['B9'] = _('OnPeak')
        ws['B9'].border = f_border

        ws['B10'].font = title_font
        ws['B10'].alignment = c_c_alignment
        ws['B10'] = _('MidPeak')
        ws['B10'].border = f_border

        ws['B11'].font = title_font
        ws['B11'].alignment = c_c_alignment
        ws['B11'] = _('OffPeak')
        ws['B11'].border = f_border

        ws['B12'].font = title_font
        ws['B12'].alignment = c_c_alignment
        ws['B12'] = _('Subtotal')
        ws['B12'].border = f_border

        for i in range(len(reporting_period_data['names'])):
            col = chr(ord('C') + i)

            ws[col + '7'].fill = table_fill
            ws[col + '7'].font = name_font
            ws[col + '7'].alignment = c_c_alignment
            ws[col + '7'].border = f_border
            ws[col + '7'] = reporting_period_data['names'][i] + '(' + reporting_period_data['units'][i] + ')'

            ws[col + '8'].font = title_font
            ws[col + '8'].alignment = c_c_alignment
            ws[col + '8'].border = f_border
            ws[col + '8'] = round2(reporting_period_data['toppeaks'][i], 2)

            ws[col + '9'].font = title_font
            ws[col + '9'].alignment = c_c_alignment
            ws[col + '9'].border = f_border
            ws[col + '9'] = round2(reporting_period_data['onpeaks'][i], 2)

            ws[col + '10'].font = title_font
            ws[col + '10'].alignment = c_c_alignment
            ws[col + '10'].border = f_border
            ws[col + '10'] = round2(reporting_period_data['midpeaks'][i], 2)

            ws[col + '11'].font = title_font
            ws[col + '11'].alignment = c_c_alignment
            ws[col + '11'].border = f_border
            ws[col + '11'] = round2(reporting_period_data['offpeaks'][i], 2)

            ws[col + '12'].font = title_font
            ws[col + '12'].alignment = c_c_alignment
            ws[col + '12'].border = f_border
            ws[col + '12'] = round2(reporting_period_data['subtotals'][i], 2)

    ####################################################################################################################
    # Details Data
    # 15: table head
    # 16 ~ 16+N table body
    # Total: N+2 rows
    # if has not energy data: set low height for rows
    ####################################################################################################################

    if "names" not in reporting_period_data.keys() or \
            reporting_period_data['names'] is None or \
            len(reporting_period_data['names']) == 0:
        for i in range(14, 16 + 1):
            ws.row_dimensions[i].height = 0.1
    else:
        ws.row_dimensions[8].height = 60
        ws['B15'].fill = table_fill
        ws['B15'].font = name_font
        ws['B15'].border = f_border
        ws['B15'] = name

        for i in range(0, len(reporting_period_data['names'])):
            col = chr(ord('C') + i)
            ws[col + '15'].fill = table_fill
            ws[col + '15'].font = name_font
            ws[col + '15'].alignment = c_c_alignment
            ws[col + '15'] = reporting_period_data['names'][i] + " (" + reporting_period_data['units'][i] + ")"
            ws[col + '15'].border = f_border

            timestamps_length = len(reporting_period_data['timestamps'][0])
            for j in range(0, timestamps_length):
                ws['B' + str(16 + j)].font = name_font
                ws['B' + str(16 + j)].alignment = c_c_alignment
                ws['B' + str(16 + j)] = reporting_period_data['timestamps'][0][j]
                ws['B' + str(16 + j)].border = f_border

                ws[col + str(16+j)].font = name_font
                ws[col + str(16+j)].alignment = c_c_alignment
                ws[col + str(16+j)] = round2(reporting_period_data['values'][i][j], 2)
                ws[col + str(16+j)].border = f_border

            ws[col + str(16 + timestamps_length)].font = name_font
            ws[col + str(16 + timestamps_length)].alignment = c_c_alignment
            ws[col + str(16 + timestamps_length)] = round2(reporting_period_data['subtotals'][i], 2)
            ws[col + str(16 + timestamps_length)].border = f_border

        ws['B' + str(16 + len(reporting_period_data['timestamps'][0]))].font = title_font
        ws['B' + str(16 + len(reporting_period_data['timestamps'][0]))].alignment = c_c_alignment
        ws['B' + str(16 + len(reporting_period_data['timestamps'][0]))] = _('Total')
        ws['B' + str(16 + len(reporting_period_data['timestamps'][0]))].border = f_border

    ####################################################################################################################
    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename
