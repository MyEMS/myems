"""
MyEMS Normalization Service - Offline Meter Processing Module

This module handles the processing of offline meter data from uploaded Excel files.
It processes Excel files containing daily energy consumption data for offline meters
and converts them into normalized hourly consumption data for analysis and reporting.

The offline meter processing performs the following functions:
1. Retrieves new offline meter files from the historical database
2. Parses Excel files to extract daily energy consumption data
3. Validates meter IDs and data ranges against system configuration
4. Converts daily values to hourly consumption data
5. Stores normalized data in the energy database
6. Updates file processing status

This module is essential for incorporating manual meter readings and historical
data from offline sources into the MyEMS system.
"""

import time
from datetime import datetime, timedelta
from decimal import Decimal

import mysql.connector
from openpyxl import load_workbook

import config


################################################################################################################
# Offline Meter Processing Procedures:
# STEP 1: Get all 'new' offline meter files from historical database
# STEP 2: For each new file, iterate all rows and read cell values and store data to energy data list
# STEP 3: Insert or update energy data to table offline meter hourly in energy database
# STEP 4: Update file status to 'done' or 'error'
################################################################################################################


def calculate_hourly(logger):
    """
    Main function for offline meter data processing from uploaded Excel files.

    This function runs continuously, checking for new offline meter files and processing
    them to convert daily energy consumption data into normalized hourly data.

    Args:
        logger: Logger instance for recording processing activities and errors
    """
    while True:
        # The outermost while loop to reconnect to server if there is a connection error
        ################################################################################################################
        # STEP 1: Get all 'new' offline meter files from historical database
        ################################################################################################################
        cnx = None
        cursor = None

        # Connect to historical database to retrieve offline meter files
        try:
            cnx = mysql.connector.connect(**config.myems_historical_db)
            cursor = cnx.cursor()
        except Exception as e:
            logger.error("Error in step 1.1 of offline meter.calculate_hourly " + str(e))
            if cursor:
                cursor.close()
            if cnx:
                cnx.close()
            # Sleep several minutes and continue the outer loop to reconnect the database
            print("Could not connect the MyEMS Historical Database, and go to sleep 60 seconds...")
            time.sleep(60)
            continue

        print("Connected to MyEMS Historical Database")

        print("Getting all new offline meter files")

        # Query for new offline meter files that need processing
        try:
            query = (" SELECT id, file_name, file_object "
                     " FROM tbl_offline_meter_files "
                     " WHERE status = 'new' "
                     " ORDER BY id ")

            cursor.execute(query, )
            rows_files = cursor.fetchall()
        except Exception as e:
            logger.error("Error in step 1.2 of offline meter.calculate_hourly " + str(e))
            time.sleep(60)
            continue
        finally:
            if cursor:
                cursor.close()
            if cnx:
                cnx.close()

        # Build list of Excel files to process
        excel_file_list = list()
        if rows_files is not None and len(rows_files) > 0:
            for row_file in rows_files:
                excel_file_list.append({"id": row_file[0],
                                        "name": row_file[1],
                                        "file_object": row_file[2]})
        else:
            print("there isn't any new files found, and go to sleep 60 seconds...")
            time.sleep(60)
            continue

        ################################################################################################################
        # STEP 2: For each new file, dump file object to local file and then load workbook from the local file
        ################################################################################################################
        for excel_file in excel_file_list:
            print("read data from offline meter file" + excel_file['name'])
            is_valid_file = True
            fw = None

            # Write Excel file object to temporary file for processing
            try:
                fw = open("myems-normalization.blob", 'wb')
                fw.write(excel_file['file_object'])
                fw.close()
            except Exception as e:
                logger.error("Error in step 2.1 of offline meter.calculate_hourly " + str(e))
                if fw:
                    fw.close()
                # Mark as invalid file
                is_valid_file = False

            fr = None
            wb = None

            # Load Excel workbook from temporary file
            try:
                fr = open("myems-normalization.blob", 'rb')
                wb = load_workbook(fr, data_only=True)
                fr.close()
            except Exception as e:
                logger.error("Error in step 2.2 of offline meter.calculate_hourly " + str(e))
                if fr:
                    fr.close()
                # Mark as invalid file
                is_valid_file = False

            energy_data_list = list()

            # Process the active worksheet if file is valid
            if is_valid_file:
                ws = wb.active

                # Get timezone offset in minutes for converting local time to UTC
                timezone_offset = int(config.utc_offset[1:3]) * 60 + int(config.utc_offset[4:6])
                if config.utc_offset[0] == '-':
                    timezone_offset = -timezone_offset

                # Process each row in the Excel file (starting from row 3, columns 1-34)
                for row in ws.iter_rows(min_row=3, max_row=1024, min_col=1, max_col=34):
                    offline_meter_data = dict()
                    offline_meter_data['offline_meter_id'] = None
                    offline_meter_data['offline_meter_name'] = None
                    offline_meter_data['data'] = dict()
                    col_num = 0

                    # Process each cell in the row
                    for cell in row:
                        col_num += 1
                        print(cell.value)

                        if col_num == 1:
                            # Get offline meter ID
                            if cell.value is not None:
                                offline_meter_data['offline_meter_id'] = cell.value
                            else:
                                break
                        elif col_num == 2:
                            # Get offline meter name
                            if cell.value is None:
                                break
                            else:
                                offline_meter_data['offline_meter_name'] = cell.value
                        elif col_num > 3:
                            # Get date of the cell (daily energy consumption data)
                            try:
                                start_datetime_local = datetime(year=ws['A2'].value,
                                                                month=ws['B2'].value,
                                                                day=col_num - 3)
                            except ValueError:
                                # Invalid date and go to next cell in this row until reach max_col
                                continue

                            # Convert local datetime to UTC
                            start_datetime_utc = start_datetime_local - timedelta(minutes=timezone_offset)

                            if cell.value is None:
                                # If the cell is empty then stop at that day
                                break
                            else:
                                offline_meter_data['data'][start_datetime_utc] = Decimal(cell.value)

                    # Add valid offline meter data to the list
                    if len(offline_meter_data['data']) > 0:
                        print("offline_meter_data:" + str(offline_meter_data))
                        energy_data_list.append(offline_meter_data)

            ############################################################################################################
            # STEP 3: Insert or update energy data to table offline meter hourly in energy database
            ############################################################################################################
            print("to valid offline meter id in excel file...")

            # Validate that offline meter data was found in the file
            if len(energy_data_list) == 0:
                print("Could not find any offline meters in the file...")
                print("and go to process the next file...")
                is_valid_file = False
            else:
                # Connect to system database to validate offline meter IDs
                try:
                    cnx = mysql.connector.connect(**config.myems_system_db)
                    cursor = cnx.cursor()
                except Exception as e:
                    logger.error("Error in step 3.1 of offlinemeter.calculate_hourly " + str(e))
                    if cursor:
                        cursor.close()
                    if cnx:
                        cnx.close()
                    time.sleep(60)
                    continue

                # Get all offline meters from system database for validation
                try:
                    cursor.execute(" SELECT id, name, hourly_low_limit, hourly_high_limit"
                                   " FROM tbl_offline_meters ")
                    rows_offline_meters = cursor.fetchall()
                except Exception as e:
                    logger.error("Error in step 3.2 of offlinemeter.calculate_hourly " + str(e))
                    time.sleep(60)
                    continue
                finally:
                    if cursor:
                        cursor.close()
                    if cnx:
                        cnx.close()

                # Validate that offline meters exist in the system database
                if rows_offline_meters is None or len(rows_offline_meters) == 0:
                    print("Could not find any offline meters in the MyEMS System Database...")
                    time.sleep(60)
                    continue
                else:
                    # Build set of valid offline meter IDs from system database
                    offline_meter_id_set = set()
                    for row_offline_meter in rows_offline_meters:
                        # Valid offline meter ID in excel file
                        offline_meter_id_set.add(row_offline_meter[0])

                    # Validate each offline meter ID and check daily consumption limits
                    for energy_data_item in energy_data_list:
                        if energy_data_item['offline_meter_id'] not in offline_meter_id_set:
                            is_valid_file = False
                            break

                        # Check daily consumption values against hourly limits
                        for row_offline_meter in rows_offline_meters:
                            if row_offline_meter[0] == energy_data_item['offline_meter_id']:
                                for key in energy_data_item['data']:
                                    # Convert daily consumption to hourly average and check limits
                                    if row_offline_meter[2] > (energy_data_item['data'][key]/24):
                                        is_valid_file = False
                                        break
                                    elif row_offline_meter[3] < (energy_data_item['data'][key]/24):
                                        is_valid_file = False
                                        break
                                break

                if is_valid_file:
                    ####################################################################################################
                    # Delete possibly existing offline meter hourly data in myems energy database,
                    # and then insert new offline meter hourly data
                    ####################################################################################################
                    try:
                        cnx = mysql.connector.connect(**config.myems_energy_db)
                        cursor = cnx.cursor()
                    except Exception as e:
                        logger.error("Error in step 3.2 of offlinemeter.calculate_hourly " + str(e))
                        if cursor:
                            cursor.close()
                        if cnx:
                            cnx.close()
                        time.sleep(60)
                        continue

                    try:
                        # Process each offline meter's daily consumption data
                        for energy_data_item in energy_data_list:
                            offline_meter_id = energy_data_item['offline_meter_id']
                            print(energy_data_item['data'].items())

                            # Convert daily consumption to hourly consumption data
                            for start_datetime_utc, daily_value in energy_data_item['data'].items():
                                end_datetime_utc = start_datetime_utc + timedelta(hours=24)

                                # Calculate hourly consumption from daily value
                                # Distribute daily consumption evenly across the day
                                actual_value = \
                                    daily_value / (Decimal(24) * Decimal(60) / Decimal(config.minutes_to_count))

                                # Delete existing hourly data for this time period
                                cursor.execute(" DELETE FROM tbl_offline_meter_hourly "
                                               " WHERE offline_meter_id = %s "
                                               "       AND start_datetime_utc >= %s "
                                               "       AND start_datetime_utc < %s ",
                                               (offline_meter_id,
                                                start_datetime_utc.isoformat()[0:19],
                                                end_datetime_utc.isoformat()[0:19]))
                                cnx.commit()

                                # TODO: Check with hourly low limit and hourly high limit
                                # Build INSERT statement for hourly consumption data
                                add_values = (" INSERT INTO tbl_offline_meter_hourly "
                                              "             (offline_meter_id, start_datetime_utc, actual_value) "
                                              " VALUES  ")

                                # Generate hourly consumption records for the entire day
                                while start_datetime_utc < end_datetime_utc:
                                    add_values += " (" + str(offline_meter_id) + ","
                                    add_values += "'" + start_datetime_utc.isoformat()[0:19] + "',"
                                    add_values += str(actual_value) + "), "
                                    start_datetime_utc += timedelta(minutes=config.minutes_to_count)

                                print("add_values:" + add_values)
                                # Trim ", " at the end of string and then execute
                                cursor.execute(add_values[:-2])
                                cnx.commit()
                    except Exception as e:
                        logger.error("Error in step 3.3 of offlinemeter.calculate_hourly " + str(e))
                        time.sleep(60)
                        continue
                    finally:
                        if cursor:
                            cursor.close()
                        if cnx:
                            cnx.close()

            ############################################################################################################
            # STEP 4: Update file status to 'done' or 'error'
            ############################################################################################################
            print("to update offline meter file status to done...")
            try:
                cnx = mysql.connector.connect(**config.myems_historical_db)
                cursor = cnx.cursor()
            except Exception as e:
                logger.error("Error in step 4.1 of offlinemeter.calculate_hourly " + str(e))
                if cursor:
                    cursor.close()
                if cnx:
                    cnx.close()
                time.sleep(60)
                continue

            try:
                # Update file status based on processing success
                update_row = (" UPDATE tbl_offline_meter_files "
                              " SET status = %s "
                              " WHERE id = %s ")
                cursor.execute(update_row, ('done' if is_valid_file else 'error', excel_file['id'],))
                cnx.commit()
            except Exception as e:
                logger.error("Error in step 4.2 of offlinemeter.calculate_hourly " + str(e))
                time.sleep(60)
                continue
            finally:
                if cursor:
                    cursor.close()
                if cnx:
                    cnx.close()

        # End of for excel_file in excel_file_list

        print("go to sleep")
        time.sleep(300)  # Sleep for 5 minutes before checking for new files
        print("wake from sleep, and go to work")
    # End of the outermost while loop
