"""
MyEMS Normalization Service - Data Repair Module

This module handles the repair of historical energy data from uploaded Excel files.
It processes data repair files that contain corrected energy values for specific points
and time periods, replacing or updating the existing data in the historical database.

The data repair process performs the following functions:
1. Retrieves new data repair files from the historical database
2. Parses Excel files to extract point data (ID, name, datetime, value)
3. Validates point IDs and data ranges against system configuration
4. Replaces existing energy values in the historical database
5. Updates file processing status to 'done' or 'error'

This module is essential for correcting data quality issues and maintaining
accurate historical energy consumption records.
"""

import time
from datetime import datetime, timedelta
from decimal import Decimal

import mysql.connector
from openpyxl import load_workbook

import config


################################################################################################################
# Data Repair Procedures:
# STEP 1: Get all 'new' data repair files from historical database
# STEP 2: For each new file, iterate all rows, read cell values and store data to point data list
# STEP 3: Insert or update point data to table tbl_energy_value in historical database
# STEP 4: Update file status to 'done' or 'error'
################################################################################################################


def do(logger):
    """
    Main data repair function that processes uploaded Excel files for data correction.

    This function runs continuously, checking for new data repair files and processing
    them to correct historical energy data in the database.

    Args:
        logger: Logger instance for recording repair activities and errors
    """
    while True:
        # The outermost while loop to reconnect to server if there is a connection error
        ################################################################################################################
        # STEP 1: Get all 'new' data repair files from historical database
        ################################################################################################################
        cnx = None
        cursor = None

        # Connect to historical database to retrieve data repair files
        try:
            cnx = mysql.connector.connect(**config.myems_historical_db)
            cursor = cnx.cursor()
        except Exception as e:
            logger.error("Error in step 1.1 of datarepair.do " + str(e))
            # Clean up database connections in case of error
            if cursor:
                cursor.close()
            if cnx:
                cnx.close()
            # Sleep several minutes and continue the outer loop to reconnect the database
            print("Could not connect the myems historical database, and go to sleep 60 seconds")
            time.sleep(60)
            continue

        print("Connected to myems historical database")
        print("Getting all new data repair files")

        # Query for new data repair files that need processing
        try:
            query = (" SELECT id, file_name, file_object "
                     " FROM tbl_data_repair_files "
                     " WHERE status = 'new' "
                     " ORDER BY id ")
            cursor.execute(query, )
            rows_files = cursor.fetchall()
        except Exception as e:
            logger.error("Error in step 1.2 of datarepair.do " + str(e))
            time.sleep(60)
            continue
        finally:
            # Always clean up database connections
            if cursor:
                cursor.close()
            if cnx:
                cnx.close()

        # Build list of Excel files to process
        excel_file_list = list()
        if rows_files is not None and len(rows_files) > 0:
            for row_file in rows_files:
                excel_file_list.append({"id": row_file[0],
                                        "name": row_file[1],
                                        "file_object": row_file[2]})
        else:
            print("there isn't any new data repair files found, and go to sleep 60 seconds")
            time.sleep(60)
            continue

        ################################################################################################################
        # STEP 2: For each new file, iterate all rows, read cell values and store data to point data list
        ################################################################################################################
        for excel_file in excel_file_list:
            print("reading data from data repair file " + excel_file['name'])
            is_valid_file = True
            fw = None

            # Write Excel file object to temporary file for processing
            try:
                fw = open("myems-data-repair.blob", 'wb')
                fw.write(excel_file['file_object'])
                fw.close()
            except Exception as e:
                logger.error("Error in step 2.1 of datarepair.do " + str(e))
                if fw:
                    fw.close()
                # Mark as invalid file
                is_valid_file = False

            fr = None
            wb = None

            # Load Excel workbook from temporary file
            try:
                fr = open("myems-data-repair.blob", 'rb')
                wb = load_workbook(fr, data_only=True)
                fr.close()
            except Exception as e:
                logger.error("Error in step 2.2 of datarepair.do " + str(e))
                if fr:
                    fr.close()
                # Mark as invalid file
                is_valid_file = False

            energy_data_list = list()

            # Process the active worksheet if file is valid
            if is_valid_file:
                ws = wb.active

                # Get timezone offset in minutes for converting local time to UTC
                timezone_offset = int(config.utc_offset[1:3]) * 60 + int(config.utc_offset[4:6])
                if config.utc_offset[0] == '-':
                    timezone_offset = -timezone_offset

                # Process each row in the Excel file (starting from row 2, columns 1-4)
                for row in ws.iter_rows(min_row=2, max_row=1024, min_col=1, max_col=4):
                    repair_file_data = dict()
                    repair_file_data['point_id'] = None
                    repair_file_data['point_name'] = None
                    repair_file_data['date_time_utc'] = None
                    repair_file_data['actual_value'] = None
                    col_num = 0

                    # Process each cell in the row
                    for cell in row:
                        col_num += 1
                        if not isinstance(cell.value, type(None)):
                            print(cell.value)

                        if col_num == 1:
                            # Get point ID (should exist in myems_system_db.tbl_points)
                            if cell.value is not None:
                                repair_file_data['point_id'] = cell.value
                            else:
                                break
                        elif col_num == 2:
                            # Get point name
                            if cell.value is None:
                                is_valid_file = False
                                break
                            else:
                                repair_file_data['point_name'] = cell.value
                        elif col_num == 3:
                            # Get datetime and convert from local time to UTC
                            if cell.value is None:
                                is_valid_file = False
                                break
                            else:
                                try:
                                    start_datetime_local = datetime.strptime(cell.value, '%Y-%m-%d %H:%M:%S')
                                    start_datetime_utc = start_datetime_local - timedelta(minutes=timezone_offset)
                                    repair_file_data['date_time_utc'] = start_datetime_utc
                                except Exception as e:
                                    print("invalid datetime " + str(e))
                                    is_valid_file = False
                                    break
                        elif col_num == 4:
                            # Get actual value (energy consumption)
                            if cell.value is None:
                                is_valid_file = False
                                break
                            else:
                                repair_file_data['actual_value'] = Decimal(cell.value)

                    # Add valid repair data to the list
                    if not isinstance(repair_file_data['point_id'], type(None)):
                        print("repair_file_data:" + str(repair_file_data))
                        energy_data_list.append(repair_file_data)

            ############################################################################################################
            # STEP 3: Insert or update point data to table tbl_energy_value in historical database
            ############################################################################################################
            print("check point id in excel file")

            # Validate that repair data was found in the file
            if len(energy_data_list) == 0:
                print("Could not find any repair data in the file")
                print("go to process the next file")
                is_valid_file = False
            else:
                # Connect to system database to validate point IDs
                try:
                    cnx = mysql.connector.connect(**config.myems_system_db)
                    cursor = cnx.cursor()
                except Exception as e:
                    logger.error("Error in step 3.1 of datarepair.do " + str(e))
                    if cursor:
                        cursor.close()
                    if cnx:
                        cnx.close()
                    time.sleep(60)
                    continue

                # Get all points from system database for validation
                try:
                    cursor.execute(" SELECT id, name, low_limit, high_limit"
                                   " FROM tbl_points ")
                    rows_points = cursor.fetchall()
                except Exception as e:
                    logger.error("Error in step 3.2 of datarepair.do " + str(e))
                    time.sleep(60)
                    continue
                finally:
                    if cursor:
                        cursor.close()
                    if cnx:
                        cnx.close()

                # Validate that points exist in the system database
                if rows_points is None or len(rows_points) == 0:
                    print("Could not find any points in the myems system database")
                    time.sleep(60)
                    continue
                elif is_valid_file:
                    # Build set of valid point IDs from system database
                    system_point_id_set = set()
                    for rows_point in rows_points:
                        # Valid point ID in excel file
                        system_point_id_set.add(rows_point[0])

                    # Build set of point IDs from the repair file
                    file_point_id_set = set()
                    for energy_data_item in energy_data_list:
                        file_point_id_set.add(energy_data_item['point_id'])

                    # Limit one point ID per file for data integrity
                    if len(file_point_id_set) != 1:
                        is_valid_file = False

                    # Validate each point ID and check value ranges
                    for energy_data_item in energy_data_list:
                        if energy_data_item['point_id'] not in system_point_id_set:
                            is_valid_file = False
                            break
                        # Check actual value with point high limit and low limit
                        for rows_point in rows_points:
                            if energy_data_item['point_id'] == rows_point[0]:
                                if energy_data_item['actual_value'] < rows_point[2]:
                                    is_valid_file = False
                                    break
                                elif energy_data_item['actual_value'] > rows_point[3]:
                                    is_valid_file = False
                                    break
                                break

                if is_valid_file:
                    ####################################################################################################
                    # Delete possibly existing point value data in myems historical database,
                    # and then insert new point data
                    ####################################################################################################
                    try:
                        cnx = mysql.connector.connect(**config.myems_historical_db)
                        cursor = cnx.cursor()
                    except Exception as e:
                        logger.error("Error in step 3.2 of datarepair.do " + str(e))
                        if cursor:
                            cursor.close()
                        if cnx:
                            cnx.close()
                        time.sleep(60)
                        continue

                    try:
                        # Extract datetime range for deletion
                        date_time_utc_list = list()
                        for i in range(len(energy_data_list)):
                            for item in (energy_data_list[i]['date_time_utc'],):
                                date_time_utc_list.append(item)

                        start_date_time_utc = min(date_time_utc_list)
                        end_date_time_utc = max(date_time_utc_list)
                        point_id = energy_data_list[0]['point_id']

                        print("deleted data from %s to %s in table myems_historical_db.tbl_energy_value",
                              start_date_time_utc, end_date_time_utc)

                        # Delete existing energy values in the time range
                        cursor.execute(" DELETE FROM tbl_energy_value "
                                       " WHERE point_id = %s "
                                       "       AND utc_date_time >= %s "
                                       "       AND utc_date_time <= %s ",
                                       (str(point_id),
                                        start_date_time_utc.isoformat()[0:19],
                                        end_date_time_utc.isoformat()[0:19]))
                        cnx.commit()

                        # Insert new repaired energy values
                        for energy_data_item in energy_data_list:
                            add_values = (" INSERT INTO tbl_energy_value "
                                          "             (point_id, utc_date_time, actual_value, is_bad) "
                                          " VALUES  ")
                            add_values += " (" + str(point_id) + ","
                            add_values += "'" + energy_data_item['date_time_utc'].isoformat()[0:19] + "',"
                            add_values += "'" + str(energy_data_item['actual_value']) + "',"
                            add_values += "0" + "), "  # Mark as good data (is_bad = 0)
                            print("add_values:" + add_values)
                            # Trim ", " at the end of string and then execute
                            cursor.execute(add_values[:-2])
                            cnx.commit()
                    except Exception as e:
                        logger.error("Error in step 3.3 of datarepair.do " + str(e))
                        time.sleep(60)
                        continue
                    finally:
                        if cursor:
                            cursor.close()
                        if cnx:
                            cnx.close()

            ############################################################################################################
            # STEP 4: Update file status to 'done' or 'error'
            ############################################################################################################
            print("updating data repair file status")
            try:
                cnx = mysql.connector.connect(**config.myems_historical_db)
                cursor = cnx.cursor()
            except Exception as e:
                logger.error("Error in step 4.1 of datarepair.do " + str(e))
                if cursor:
                    cursor.close()
                if cnx:
                    cnx.close()
                time.sleep(60)
                continue

            try:
                # Update file status based on processing success
                update_row = (" UPDATE tbl_data_repair_files "
                              " SET status = %s "
                              " WHERE id = %s ")
                cursor.execute(update_row, ('done' if is_valid_file else 'error', excel_file['id'],))
                cnx.commit()
            except Exception as e:
                logger.error("Error in step 4.2 of datarepair.do " + str(e))
                time.sleep(60)
                continue
            finally:
                if cursor:
                    cursor.close()
                if cnx:
                    cnx.close()

        # End of for excel_file in excel_file_list

        print("go to sleep")
        time.sleep(300)  # Sleep for 5 minutes before checking for new files
        print("wake from sleep, and go to work")
    # End of the outermost while loop
